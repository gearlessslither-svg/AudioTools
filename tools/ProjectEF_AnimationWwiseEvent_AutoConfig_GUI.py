#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from __future__ import annotations

import json
import math
import os
import queue
import random
import re
import csv
import shutil
import subprocess
import sys
import tempfile
import threading
import tkinter as tk
import traceback
import time
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
from datetime import datetime


APP_DIR = Path(__file__).resolve().parent
CORE_SCRIPT = APP_DIR / "ProjectEF_AnimationWwiseEvent_AutoConfig.py"
if str(APP_DIR) not in sys.path:
    sys.path.insert(0, str(APP_DIR))

import ProjectEF_AnimationWwiseEvent_AutoConfig as core

try:
    import ProjectEF_AudioCodexTaskCard_GUI as task_card
except Exception:
    task_card = None

try:
    import ProjectEF_ActionResourceIndex as action_index
except Exception:
    action_index = None


DEFAULT_ANIMATION = str(
    core.DEFAULT_UNITY_ROOT
    / "Assets/GameProject/ArtAssets/Bird/Clp_Bird_Mallards01_General_Fly_Loop.fbx"
)
DEFAULT_WWISE_EVENT = "Play_Bird_Wing_Flap_Small"
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".bmp", ".webp", ".tif", ".tiff"}

BG = "#12161c"
PANEL = "#1a2028"
PANEL_2 = "#222a34"
CARD = "#28313d"
INK = "#edf3f8"
MUTED = "#a7b3c0"
LINE = "#3a4654"
ACCENT = "#3fb7a4"
ACCENT_HOVER = "#55c8b7"
WARN = "#f0b65a"
BAD = "#f07178"
GOOD = "#65d18f"


class AnimationWwiseEventAutoConfigGui(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title("ProjectEF Animation Wwise Event AutoConfig GUI")
        self.geometry("1180x780")
        self.minsize(1040, 660)
        self.configure(bg=BG)

        self.unity_root_var = tk.StringVar(value=str(core.DEFAULT_UNITY_ROOT))
        self.wwise_root_var = tk.StringVar(value=str(core.DEFAULT_WWISE_ROOT))
        self.animation_var = tk.StringVar(value=DEFAULT_ANIMATION)
        self.wwise_event_var = tk.StringVar(value=DEFAULT_WWISE_EVENT)
        self.prefab_var = tk.StringVar(value="")
        self.mode_var = tk.StringVar(value="auto")
        self.endpoint_regex_var = tk.StringVar(value=r"wing|hand|foot|paw|toe|tail")
        self.event_times_var = tk.StringVar(value="")
        self.sample_fps_var = tk.StringVar(value="60")
        self.strength_ratio_var = tk.StringVar(value="0.30")
        self.min_gap_var = tk.StringVar(value="0.28")
        self.skip_prefab_component_var = tk.BooleanVar(value=False)
        self.audio_aware_spacing_var = tk.BooleanVar(value=True)
        self.p4_reconcile_var = tk.BooleanVar(value=True)
        self.audio_mode_var = tk.StringVar(value="Click")
        self.playback_speed_var = tk.StringVar(value="1.0")
        self.loop_preview_var = tk.BooleanVar(value=True)
        self.ocr_status_var = tk.StringVar(value="OCR: checking")

        self.resolved_animation_var = tk.StringVar(value="未定位")
        self.source_animation_var = tk.StringVar(value="未定位")
        self.resolved_prefab_var = tk.StringVar(value="自动")
        self.wwise_evidence_var = tk.StringVar(value="未校验")
        self.summary_var = tk.StringVar(value="填写动画和 Wwise Event 后，可先定位资源，再直接预览。")
        self.status_var = tk.StringVar(value="Ready")
        self.candidate_filter_var = tk.StringVar(value="")

        self.messages: queue.Queue[tuple[str, object]] = queue.Queue()
        self.worker: threading.Thread | None = None
        self.command_buttons: list[tk.Button] = []
        self.last_report_md: Path | None = None
        self.last_report_json: Path | None = None
        self.last_command: list[str] | None = None
        self.preview_data: dict[str, object] | None = None
        self.last_report: dict[str, object] | None = None
        self.last_task_card = None
        self.last_task_card_md: Path | None = None
        self.last_task_card_json: Path | None = None
        self.last_action_index_json: Path | None = None
        self.last_action_index_xlsx: Path | None = None
        self.task_candidates: list[object] = []
        self.unity_preview_request_path: Path | None = None
        self.auto_open_unity_preview_var = tk.BooleanVar(value=False)
        self.playing = False
        self.play_start_perf = 0.0
        self.play_base_time = 0.0
        self.next_event_index = 0
        self.batch_items: list[dict[str, str]] = []
        self.preview_visual_mode = "animation"
        self.timeline_preview_frames: list[Path] = []
        self.timeline_preview_display_indices: list[int] = []
        self.timeline_preview_manifest: dict[str, object] | None = None
        self.timeline_preview_frame_index = 0
        self.timeline_preview_photo = None
        self.timeline_preview_popup_photo = None
        self.timeline_preview_window: tk.Toplevel | None = None
        self.timeline_preview_popup_canvas: tk.Canvas | None = None
        self.timeline_preview_popup_user_closed = False
        self.timeline_preview_interval_ms = 100
        self.timeline_preview_poll_attempts = 0
        self.timeline_preview_playing = False
        self.timeline_preview_error = ""

        self.configure_style()
        self.build_ui()
        self.candidate_filter_var.trace_add("write", lambda *_args: self.refresh_task_candidate_tree())
        self.refresh_ocr_status()
        self.after(100, self.pump_messages)

    def configure_style(self) -> None:
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass
        style.configure("TFrame", background=BG)
        style.configure("TLabel", background=BG, foreground=INK)
        style.configure("Treeview", background="#10151b", fieldbackground="#10151b", foreground=INK, rowheight=24)
        style.configure("Treeview.Heading", background=PANEL_2, foreground=INK, relief="flat")
        style.map("Treeview", background=[("selected", "#315f58")])
        style.configure("TCombobox", fieldbackground=PANEL_2, background=PANEL_2, foreground=INK)

    def build_ui(self) -> None:
        root = tk.Frame(self, bg=BG)
        root.pack(fill=tk.BOTH, expand=True)

        header = tk.Frame(root, bg=BG)
        header.pack(fill=tk.X, padx=18, pady=(16, 10))
        tk.Label(
            header,
            text="Animation Wwise Event AutoConfig",
            bg=BG,
            fg=INK,
            font=("Segoe UI", 22, "bold"),
        ).pack(anchor="w")
        tk.Label(
            header,
            text="定位 .fbx/.anim 和 Wwise Event，预览自动分析出的打点，再一键写入 Animation Event 与接收器配置。",
            bg=BG,
            fg=MUTED,
            font=("Segoe UI", 10),
        ).pack(anchor="w", pady=(2, 0))

        config = tk.Frame(root, bg=PANEL, highlightbackground=LINE, highlightthickness=1)
        config.pack(fill=tk.X, padx=18, pady=(0, 12))
        self.path_row(config, "Unity Root", self.unity_root_var, self.choose_unity_root, self.open_unity_root).pack(
            fill=tk.X, padx=12, pady=(12, 6)
        )
        self.path_row(config, "Wwise Root", self.wwise_root_var, self.choose_wwise_root, self.open_wwise_root).pack(
            fill=tk.X, padx=12, pady=6
        )
        self.path_row(config, "Animation", self.animation_var, self.choose_animation, self.open_animation_location).pack(
            fill=tk.X, padx=12, pady=6
        )
        self.path_row(config, "Prefab", self.prefab_var, self.choose_prefab, self.open_prefab_location, optional=True).pack(
            fill=tk.X, padx=12, pady=6
        )

        event_row = tk.Frame(config, bg=PANEL)
        event_row.pack(fill=tk.X, padx=12, pady=(6, 12))
        tk.Label(event_row, text="Wwise Event", bg=PANEL, fg=MUTED, font=("Segoe UI", 9, "bold"), width=13, anchor="w").pack(
            side=tk.LEFT
        )
        tk.Entry(
            event_row,
            textvariable=self.wwise_event_var,
            bg=PANEL_2,
            fg=INK,
            insertbackground=INK,
            relief=tk.FLAT,
            font=("Consolas", 9),
        ).pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=6, padx=(4, 8))
        self.command_button(event_row, "校验 / 定位", self.resolve_assets, CARD, INK).pack(side=tk.LEFT, padx=(0, 8))
        self.command_button(event_row, "打开 WWU", self.open_wwise_event_location, CARD, INK).pack(side=tk.LEFT)

        options = tk.Frame(root, bg=PANEL, highlightbackground=LINE, highlightthickness=1)
        options.pack(fill=tk.X, padx=18, pady=(0, 12))
        self.option_row(options).pack(fill=tk.X, padx=12, pady=10)

        actions = tk.Frame(root, bg=PANEL, highlightbackground=LINE, highlightthickness=1)
        actions.pack(fill=tk.X, padx=18, pady=(0, 12))
        self.command_button(actions, "定位资源", self.resolve_assets, CARD, INK).pack(side=tk.LEFT, padx=(12, 8), pady=10)
        self.command_button(actions, "直接预览", self.preview_config, ACCENT, "#06120f").pack(side=tk.LEFT, padx=(0, 8), pady=10)
        self.command_button(actions, "写入 / 修改配置", self.apply_config, WARN, "#161007").pack(side=tk.LEFT, padx=(0, 8), pady=10)
        self.command_button(actions, "锁定可编辑 .anim", self.use_resolved_animation, CARD, INK).pack(side=tk.LEFT, padx=(0, 8), pady=10)
        self.command_button(actions, "Unity Preview", self.open_unity_preview, CARD, INK).pack(side=tk.LEFT, padx=(0, 8), pady=10)
        self.command_button(actions, "Timeline Preview", self.open_timeline_prefab_preview, CARD, INK).pack(
            side=tk.LEFT, padx=(0, 8), pady=10
        )
        self.command_button(actions, "Frame Debug", self.open_timeline_frame_debug_preview, CARD, INK).pack(
            side=tk.LEFT, padx=(0, 8), pady=10
        )
        self.command_button(actions, "Project Audio", self.open_project_audio_preview, CARD, INK).pack(
            side=tk.LEFT, padx=(0, 8), pady=10
        )
        self.command_button(actions, "Unity Edit Keys", self.open_anim_text_editor, CARD, INK).pack(side=tk.LEFT, padx=(0, 8), pady=10)
        self.command_button(actions, "打开报告", self.open_last_report, CARD, INK).pack(side=tk.LEFT, padx=(0, 8), pady=10)
        self.command_button(actions, "报告目录", self.open_report_folder, CARD, INK).pack(side=tk.LEFT, padx=(0, 8), pady=10)
        self.command_button(actions, "复制命令", self.copy_cli_command, CARD, INK).pack(side=tk.LEFT, padx=(0, 8), pady=10)
        self.command_button(actions, "清空日志", self.clear_log, CARD, INK).pack(side=tk.LEFT, padx=(0, 12), pady=10)

        batch = tk.Frame(root, bg=PANEL, highlightbackground=LINE, highlightthickness=1)
        batch.pack(fill=tk.X, padx=18, pady=(0, 12))
        self.build_batch_panel(batch)

        body = tk.PanedWindow(root, orient=tk.HORIZONTAL, sashrelief=tk.FLAT, bg=BG, bd=0, sashwidth=8)
        body.pack(fill=tk.BOTH, expand=True, padx=18, pady=(0, 12))

        left = tk.Frame(body, bg=PANEL, highlightbackground=LINE, highlightthickness=1)
        right = tk.Frame(body, bg=PANEL, highlightbackground=LINE, highlightthickness=1)
        body.add(left, minsize=430)
        body.add(right, minsize=520)

        self.build_location_panel(left)
        self.build_preview_panel(right)

        footer = tk.Frame(root, bg=BG)
        footer.pack(fill=tk.X, padx=18, pady=(0, 12))
        tk.Label(footer, textvariable=self.status_var, bg=BG, fg=MUTED, font=("Segoe UI", 9)).pack(side=tk.LEFT)
        tk.Label(footer, text=str(CORE_SCRIPT), bg=BG, fg="#6f7b88", font=("Segoe UI", 9)).pack(side=tk.RIGHT)

    def path_row(
        self,
        parent: tk.Frame,
        label: str,
        variable: tk.StringVar,
        browse_command,
        locate_command,
        optional: bool = False,
    ) -> tk.Frame:
        row = tk.Frame(parent, bg=PANEL)
        tk.Label(row, text=label, bg=PANEL, fg=MUTED, font=("Segoe UI", 9, "bold"), width=13, anchor="w").pack(
            side=tk.LEFT
        )
        entry = tk.Entry(
            row,
            textvariable=variable,
            bg=PANEL_2,
            fg=INK,
            insertbackground=INK,
            relief=tk.FLAT,
            font=("Consolas", 9),
        )
        entry.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=6, padx=(4, 8))
        if optional:
            self.command_button(row, "自动", lambda: variable.set(""), CARD, INK).pack(side=tk.LEFT, padx=(0, 8))
        self.command_button(row, "浏览", browse_command, CARD, INK).pack(side=tk.LEFT, padx=(0, 8))
        self.command_button(row, "定位", locate_command, CARD, INK).pack(side=tk.LEFT)
        return row

    def option_row(self, parent: tk.Frame) -> tk.Frame:
        row = tk.Frame(parent, bg=PANEL)
        tk.Label(row, text="分析模式", bg=PANEL, fg=MUTED, font=("Segoe UI", 9, "bold")).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Combobox(
            row,
            textvariable=self.mode_var,
            values=("auto", "manual", "contact", "downstroke", "speed"),
            state="readonly",
            width=12,
        ).pack(side=tk.LEFT, padx=(0, 16))
        self.small_entry(row, "端点匹配", self.endpoint_regex_var, 14)
        self.small_entry(row, "Event Times", self.event_times_var, 12)
        self.small_entry(row, "采样 FPS", self.sample_fps_var, 6)
        self.small_entry(row, "强度阈值", self.strength_ratio_var, 6)
        self.small_entry(row, "最小间隔", self.min_gap_var, 6)
        tk.Checkbutton(
            row,
            text="音频感知间隔",
            variable=self.audio_aware_spacing_var,
            bg=PANEL,
            fg=INK,
            activebackground=PANEL,
            activeforeground=INK,
            selectcolor=PANEL_2,
        ).pack(side=tk.LEFT, padx=(0, 12))
        tk.Checkbutton(
            row,
            text="Preview后打开Unity",
            variable=self.auto_open_unity_preview_var,
            bg=PANEL,
            fg=INK,
            activebackground=PANEL,
            activeforeground=INK,
            selectcolor=PANEL_2,
        ).pack(side=tk.LEFT, padx=(0, 12))
        tk.Checkbutton(
            row,
            text="Apply后P4 Reconcile",
            variable=self.p4_reconcile_var,
            bg=PANEL,
            fg=INK,
            activebackground=PANEL,
            activeforeground=INK,
            selectcolor=PANEL_2,
        ).pack(side=tk.LEFT, padx=(0, 12))
        tk.Checkbutton(
            row,
            text="不处理 Prefab 接收器",
            variable=self.skip_prefab_component_var,
            bg=PANEL,
            fg=INK,
            activebackground=PANEL,
            activeforeground=INK,
            selectcolor=PANEL_2,
        ).pack(side=tk.LEFT, padx=(12, 0))
        return row

    def build_batch_panel(self, parent: tk.Frame) -> None:
        top = tk.Frame(parent, bg=PANEL)
        top.pack(fill=tk.X, padx=12, pady=(10, 8))
        tk.Label(top, text="Round Info / Task Card / Batch", bg=PANEL, fg=MUTED, font=("Segoe UI", 10, "bold")).pack(side=tk.LEFT)
        tk.Label(top, textvariable=self.ocr_status_var, bg=PANEL, fg=MUTED, font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=(12, 0))
        self.command_button(top, "Generate Card", self.generate_task_card, GOOD, "#06120f").pack(side=tk.RIGHT, padx=(8, 0))
        self.command_button(top, "Scan Action Index", self.scan_action_index, GOOD, "#06120f").pack(side=tk.RIGHT, padx=(8, 0))
        self.command_button(top, "Load Action Index", self.load_action_index, CARD, INK).pack(side=tk.RIGHT, padx=(8, 0))
        self.command_button(top, "Load Card", self.load_task_card, CARD, INK).pack(side=tk.RIGHT, padx=(8, 0))
        self.command_button(top, "Load Latest", self.load_latest_task_card, CARD, INK).pack(side=tk.RIGHT, padx=(8, 0))
        self.command_button(top, "Preview Selected", self.preview_selected_candidate, ACCENT, "#06120f").pack(side=tk.RIGHT, padx=(8, 0))
        self.command_button(top, "Use Candidate", self.use_selected_task_candidate, CARD, INK).pack(side=tk.RIGHT, padx=(8, 0))
        self.command_button(top, "Copy Card", self.copy_task_card, CARD, INK).pack(side=tk.RIGHT, padx=(8, 0))
        self.command_button(top, "Paste Text", self.paste_batch_text, CARD, INK).pack(side=tk.RIGHT, padx=(8, 0))
        self.command_button(top, "Paste Image OCR", self.paste_batch_image_ocr, CARD, INK).pack(side=tk.RIGHT, padx=(8, 0))
        self.command_button(top, "OCR Files", self.ocr_image_files, CARD, INK).pack(side=tk.RIGHT, padx=(8, 0))
        self.command_button(top, "Parse Batch", self.parse_batch_text, ACCENT, "#06120f").pack(side=tk.RIGHT, padx=(8, 0))
        self.command_button(top, "Batch Preview", self.preview_batch, CARD, INK).pack(side=tk.RIGHT, padx=(8, 0))
        self.command_button(top, "Batch Apply", self.apply_batch, WARN, "#161007").pack(side=tk.RIGHT, padx=(8, 0))

        body = tk.Frame(parent, bg=PANEL)
        body.pack(fill=tk.X, padx=12, pady=(0, 10))
        self.batch_text = tk.Text(
            body,
            bg="#10151b",
            fg=INK,
            insertbackground=INK,
            relief=tk.FLAT,
            wrap=tk.WORD,
            font=("Consolas", 9),
            height=2,
            padx=8,
            pady=8,
        )
        self.batch_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))
        self.batch_text.bind("<<Paste>>", self.on_batch_text_paste)

        self.batch_tree = ttk.Treeview(
            body,
            columns=("index", "animation", "event", "times", "status"),
            show="headings",
            height=2,
        )
        for key, label, width in [
            ("index", "#", 44),
            ("animation", "Animation", 240),
            ("event", "Wwise Event", 220),
            ("times", "Times", 90),
            ("status", "Status", 180),
        ]:
            self.batch_tree.heading(key, text=label)
            self.batch_tree.column(key, width=width, anchor="w", stretch=key in {"animation", "event", "status"})
        self.batch_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.batch_tree.bind("<<TreeviewSelect>>", self.on_batch_tree_select)

        candidate_frame = tk.Frame(parent, bg=PANEL)
        candidate_frame.pack(fill=tk.X, padx=12, pady=(0, 10))
        candidate_header = tk.Frame(candidate_frame, bg=PANEL)
        candidate_header.pack(fill=tk.X, pady=(0, 4))
        tk.Label(
            candidate_header,
            text="Task Card / Action Index Candidates (select one, then Use Candidate or Preview)",
            bg=PANEL,
            fg=MUTED,
            font=("Segoe UI", 9, "bold"),
        ).pack(side=tk.LEFT)
        tk.Label(candidate_header, text="Filter", bg=PANEL, fg=MUTED, font=("Segoe UI", 9, "bold")).pack(
            side=tk.RIGHT, padx=(8, 4)
        )
        tk.Entry(
            candidate_header,
            textvariable=self.candidate_filter_var,
            bg=PANEL_2,
            fg=INK,
            insertbackground=INK,
            relief=tk.FLAT,
            font=("Consolas", 9),
            width=32,
        ).pack(side=tk.RIGHT, ipady=4)
        candidate_list = tk.Frame(candidate_frame, bg=PANEL)
        candidate_list.pack(fill=tk.BOTH, expand=True)
        self.task_candidate_tree = ttk.Treeview(
            candidate_list,
            columns=("index", "kind", "score", "path"),
            show="headings",
            height=8,
        )
        for key, label, width in [
            ("index", "#", 44),
            ("kind", "Kind", 220),
            ("score", "Score", 70),
            ("path", "Unity Path", 760),
        ]:
            self.task_candidate_tree.heading(key, text=label)
            self.task_candidate_tree.column(key, width=width, anchor="w", stretch=key == "path")
        candidate_scroll = ttk.Scrollbar(candidate_list, orient=tk.VERTICAL, command=self.task_candidate_tree.yview)
        self.task_candidate_tree.configure(yscrollcommand=candidate_scroll.set)
        self.task_candidate_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        candidate_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.task_candidate_tree.bind("<Double-1>", lambda _event: self.use_selected_task_candidate())

    def small_entry(self, parent: tk.Frame, label: str, variable: tk.StringVar, width: int) -> None:
        tk.Label(parent, text=label, bg=PANEL, fg=MUTED, font=("Segoe UI", 9, "bold")).pack(side=tk.LEFT, padx=(0, 6))
        tk.Entry(
            parent,
            textvariable=variable,
            bg=PANEL_2,
            fg=INK,
            insertbackground=INK,
            relief=tk.FLAT,
            font=("Consolas", 9),
            width=width,
        ).pack(side=tk.LEFT, ipady=5, padx=(0, 16))

    def command_button(self, parent: tk.Frame, text: str, command, bg: str, fg: str) -> tk.Button:
        button = tk.Button(
            parent,
            text=text,
            command=command,
            bg=bg,
            fg=fg,
            activebackground=ACCENT_HOVER if bg == ACCENT else "#344250",
            activeforeground=fg,
            relief=tk.FLAT,
            padx=12,
            pady=7,
            font=("Segoe UI", 9, "bold"),
        )
        self.command_buttons.append(button)
        return button

    def build_location_panel(self, parent: tk.Frame) -> None:
        tk.Label(parent, text="定位结果", bg=PANEL, fg=MUTED, font=("Segoe UI", 10, "bold")).pack(
            anchor="w", padx=12, pady=(10, 8)
        )
        self.info_line(parent, "Editable .anim", self.resolved_animation_var, GOOD).pack(fill=tk.X, padx=12, pady=4)
        self.info_line(parent, "Source .fbx", self.source_animation_var, MUTED).pack(fill=tk.X, padx=12, pady=4)
        self.info_line(parent, "Prefab", self.resolved_prefab_var, MUTED).pack(fill=tk.X, padx=12, pady=4)
        self.info_line(parent, "Wwise evidence", self.wwise_evidence_var, MUTED).pack(fill=tk.X, padx=12, pady=4)

        tk.Label(parent, text="摘要", bg=PANEL, fg=MUTED, font=("Segoe UI", 10, "bold")).pack(
            anchor="w", padx=12, pady=(14, 6)
        )
        tk.Label(
            parent,
            textvariable=self.summary_var,
            bg="#10151b",
            fg=INK,
            justify=tk.LEFT,
            anchor="nw",
            wraplength=390,
            padx=10,
            pady=10,
        ).pack(fill=tk.X, padx=12, pady=(0, 10))

        tk.Label(parent, text="日志", bg=PANEL, fg=MUTED, font=("Segoe UI", 10, "bold")).pack(
            anchor="w", padx=12, pady=(4, 6)
        )
        self.log = tk.Text(
            parent,
            bg="#10151b",
            fg=INK,
            insertbackground=INK,
            relief=tk.FLAT,
            wrap=tk.WORD,
            font=("Consolas", 9),
            padx=10,
            pady=10,
            height=13,
        )
        self.log.pack(fill=tk.BOTH, expand=True, padx=12, pady=(0, 12))
        self.log.insert(tk.END, "Ready.\n")
        self.log.configure(state=tk.DISABLED)

    def build_preview_panel(self, parent: tk.Frame) -> None:
        tk.Label(parent, text="预览打点", bg=PANEL, fg=MUTED, font=("Segoe UI", 10, "bold")).pack(
            anchor="w", padx=12, pady=(10, 8)
        )
        self.event_tree = ttk.Treeview(parent, columns=("index", "time"), show="headings", height=12)
        self.event_tree.heading("index", text="#")
        self.event_tree.heading("time", text="Time (s)")
        self.event_tree.column("index", width=70, anchor="center", stretch=False)
        self.event_tree.column("time", width=160, anchor="w")
        self.event_tree.pack(fill=tk.X, padx=12, pady=(0, 10))

        controls = tk.Frame(parent, bg=PANEL)
        controls.pack(fill=tk.X, padx=12, pady=(0, 10))
        self.play_button = self.command_button(controls, "播放/暂停", self.toggle_preview_playback, ACCENT, "#06120f")
        self.play_button.pack(side=tk.LEFT, padx=(0, 8))
        self.command_button(controls, "停止", self.stop_preview_playback, CARD, INK).pack(side=tk.LEFT, padx=(0, 12))
        tk.Label(controls, text="音频", bg=PANEL, fg=MUTED, font=("Segoe UI", 9, "bold")).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Combobox(
            controls,
            textvariable=self.audio_mode_var,
            values=("Click", "Wwise Source", "Mute"),
            state="readonly",
            width=12,
        ).pack(side=tk.LEFT, padx=(0, 12))
        tk.Label(controls, text="速度", bg=PANEL, fg=MUTED, font=("Segoe UI", 9, "bold")).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Combobox(
            controls,
            textvariable=self.playback_speed_var,
            values=("0.5", "1.0", "2.0", "4.0"),
            state="readonly",
            width=6,
        ).pack(side=tk.LEFT, padx=(0, 12))
        tk.Checkbutton(
            controls,
            text="Loop",
            variable=self.loop_preview_var,
            bg=PANEL,
            fg=INK,
            activebackground=PANEL,
            activeforeground=INK,
            selectcolor=PANEL_2,
        ).pack(side=tk.LEFT)

        self.preview_canvas = tk.Canvas(parent, bg="#10151b", highlightthickness=0, height=230)
        self.preview_canvas.pack(fill=tk.X, padx=12, pady=(0, 10))
        self.preview_canvas.bind("<Configure>", lambda _event: self.draw_preview(self.current_preview_time()))

        tk.Label(parent, text="分析骨骼 / 端点", bg=PANEL, fg=MUTED, font=("Segoe UI", 10, "bold")).pack(
            anchor="w", padx=12, pady=(2, 6)
        )
        self.endpoint_text = tk.Text(
            parent,
            bg="#10151b",
            fg=INK,
            insertbackground=INK,
            relief=tk.FLAT,
            wrap=tk.NONE,
            font=("Consolas", 9),
            padx=10,
            pady=10,
            height=8,
        )
        self.endpoint_text.pack(fill=tk.BOTH, expand=True, padx=12, pady=(0, 12))
        self.endpoint_text.configure(state=tk.DISABLED)

    def info_line(self, parent: tk.Frame, label: str, variable: tk.StringVar, value_color: str) -> tk.Frame:
        row = tk.Frame(parent, bg=PANEL)
        tk.Label(row, text=label, bg=PANEL, fg=MUTED, font=("Segoe UI", 9, "bold"), width=15, anchor="w").pack(side=tk.LEFT)
        tk.Label(
            row,
            textvariable=variable,
            bg="#10151b",
            fg=value_color,
            anchor="w",
            justify=tk.LEFT,
            wraplength=295,
            padx=8,
            pady=6,
        ).pack(side=tk.LEFT, fill=tk.X, expand=True)
        return row

    def choose_unity_root(self) -> None:
        path = filedialog.askdirectory(initialdir=self.unity_root_var.get() or str(core.DEFAULT_UNITY_ROOT))
        if path:
            self.unity_root_var.set(path)

    def choose_wwise_root(self) -> None:
        path = filedialog.askdirectory(initialdir=self.wwise_root_var.get() or str(core.DEFAULT_WWISE_ROOT))
        if path:
            self.wwise_root_var.set(path)

    def choose_animation(self) -> None:
        initial = self.animation_var.get() or str(Path(self.unity_root_var.get()) / "Assets")
        initial_dir = str(Path(initial).parent if Path(initial).suffix else Path(initial))
        path = filedialog.askopenfilename(
            initialdir=initial_dir,
            title="选择 Animation .anim 或源 .fbx",
            filetypes=(("Unity Animation / FBX", "*.anim *.fbx"), ("Unity Animation", "*.anim"), ("FBX", "*.fbx"), ("All files", "*.*")),
        )
        if path:
            self.animation_var.set(path)

    def choose_prefab(self) -> None:
        initial = self.prefab_var.get() or str(Path(self.unity_root_var.get()) / "Assets")
        initial_dir = str(Path(initial).parent if Path(initial).suffix else Path(initial))
        path = filedialog.askopenfilename(
            initialdir=initial_dir,
            title="选择 Prefab，留空则按动画目录自动推断",
            filetypes=(("Unity Prefab", "*.prefab"), ("All files", "*.*")),
        )
        if path:
            self.prefab_var.set(path)

    def append_batch_text(self, text: str) -> None:
        if not text.strip():
            return
        self.batch_text.insert(tk.END, ("\n\n" if self.batch_text.get("1.0", tk.END).strip() else "") + text.strip())
        self.batch_text.see(tk.END)

    def paste_batch_text(self) -> None:
        try:
            text = self.clipboard_get()
        except tk.TclError:
            messagebox.showinfo("Clipboard", "Clipboard does not contain text.")
            return
        self.append_batch_text(text)

    def on_batch_text_paste(self, _event: object | None = None) -> str | None:
        try:
            text = self.clipboard_get()
        except tk.TclError:
            text = ""
        if text:
            return None

        try:
            images = self.read_clipboard_images()
        except Exception as exc:
            self.append_log(f"Clipboard image OCR unavailable: {exc}")
            return None
        if not images:
            return None

        self.run_background("Running OCR from pasted clipboard image...", lambda: self.ocr_images_worker(images))
        return "break"

    def find_tesseract_executable(self) -> Path | None:
        found = shutil.which("tesseract")
        candidates = [
            Path(found) if found else None,
            Path(r"C:\Program Files\Tesseract-OCR\tesseract.exe"),
            Path(r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe"),
        ]
        for candidate in candidates:
            if candidate and candidate.exists():
                return candidate
        return None

    def prepare_ocr_engine(self):
        try:
            import pytesseract
        except Exception as exc:
            return None, f"pytesseract is not available: {exc}"

        tesseract = self.find_tesseract_executable()
        if not tesseract:
            return None, "tesseract.exe was not found. OCR is unavailable, but batch text/manual entries still work."
        pytesseract.pytesseract.tesseract_cmd = str(tesseract)
        return pytesseract, None

    def windows_ocr_available(self) -> bool:
        command = (
            "Add-Type -AssemblyName System.Runtime.WindowsRuntime; "
            "$null -ne [Windows.Media.Ocr.OcrEngine, Windows.Foundation, ContentType=WindowsRuntime]::TryCreateFromUserProfileLanguages()"
        )
        try:
            completed = subprocess.run(
                ["powershell.exe", "-NoProfile", "-NonInteractive", "-Command", command],
                stdout=subprocess.PIPE,
                stderr=subprocess.DEVNULL,
                text=True,
                encoding="utf-8",
                errors="replace",
                timeout=5,
                creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0) if os.name == "nt" else 0,
            )
        except Exception:
            return False
        return "True" in completed.stdout

    def prepare_image_for_ocr(self, image_like: object) -> tuple[Path, bool]:
        try:
            from PIL import Image, ImageEnhance, ImageFilter
        except Exception:
            if isinstance(image_like, Path):
                return image_like, False
            raise

        if isinstance(image_like, Path):
            image = Image.open(image_like)
        else:
            image = image_like  # type: ignore[assignment]
        image = image.convert("RGB")  # type: ignore[attr-defined]
        scale = 4 if max(image.size) < 1400 else 2
        image = image.resize((image.width * scale, image.height * scale), Image.Resampling.LANCZOS)
        image = ImageEnhance.Contrast(image).enhance(1.8)
        image = ImageEnhance.Sharpness(image).enhance(1.4)
        image = image.filter(ImageFilter.SHARPEN)
        path = Path(tempfile.gettempdir()) / f"ProjectEF_AnimationWwiseEvent_OCR_{time.time_ns()}.png"
        image.save(path)
        return path, True

    def refresh_ocr_status(self) -> None:
        try:
            import pytesseract  # noqa: F401
        except Exception:
            if self.windows_ocr_available():
                self.ocr_status_var.set("OCR: Windows fallback")
            else:
                self.ocr_status_var.set("OCR: unavailable")
            return
        if self.find_tesseract_executable():
            self.ocr_status_var.set("OCR: ready")
        elif self.windows_ocr_available():
            self.ocr_status_var.set("OCR: Windows fallback")
        else:
            self.ocr_status_var.set("OCR: tesseract.exe missing")

    def read_clipboard_images(self) -> list[tuple[str, object]]:
        try:
            from PIL import ImageGrab
        except Exception as exc:
            raise RuntimeError(f"Pillow ImageGrab is not available: {exc}") from exc

        payload = ImageGrab.grabclipboard()
        if payload is None:
            return []
        if isinstance(payload, list):
            images: list[tuple[str, object]] = []
            for item in payload:
                path = Path(str(item))
                if path.suffix.lower() in {".png", ".jpg", ".jpeg", ".bmp", ".webp", ".tif", ".tiff"}:
                    images.append((str(path), path))
            return images
        return [("clipboard image", payload)]

    def paste_batch_image_ocr(self) -> None:
        try:
            images = self.read_clipboard_images()
        except Exception as exc:
            messagebox.showerror("OCR", str(exc))
            return
        if not images:
            messagebox.showinfo("OCR", "Clipboard does not contain an image or copied image files.")
            return
        self.run_background("Running OCR from clipboard image...", lambda: self.ocr_images_worker(images))

    def ocr_image_files(self) -> None:
        paths = filedialog.askopenfilenames(
            title="Select image files for OCR",
            filetypes=(
                ("Images", "*.png *.jpg *.jpeg *.bmp *.webp *.tif *.tiff"),
                ("All files", "*.*"),
            ),
        )
        if not paths:
            return
        images = [(str(path), Path(path)) for path in paths]
        self.run_background("Running OCR from image files...", lambda: self.ocr_images_worker(images))

    def windows_ocr_image(self, image_path: Path) -> str:
        script = r'''
$ErrorActionPreference = 'Stop'
[Console]::OutputEncoding = [System.Text.Encoding]::UTF8
Add-Type -AssemblyName System.Runtime.WindowsRuntime
function AwaitOp($op, [type]$resultType) {
    $methods = [System.WindowsRuntimeSystemExtensions].GetMethods() | Where-Object {
        $_.Name -eq 'AsTask' -and $_.IsGenericMethodDefinition -and $_.GetParameters().Count -eq 1
    }
    $method = $methods | Where-Object { $_.GetParameters()[0].ParameterType.Name -eq 'IAsyncOperation`1' } | Select-Object -First 1
    if ($null -eq $method) { throw 'No generic AsTask(IAsyncOperation<T>) overload found.' }
    $task = $method.MakeGenericMethod($resultType).Invoke($null, @($op))
    $task.Wait()
    return $task.Result
}
$path = $env:PROJECTEF_OCR_IMAGE
$file = AwaitOp ([Windows.Storage.StorageFile, Windows.Storage, ContentType=WindowsRuntime]::GetFileFromPathAsync($path)) ([Windows.Storage.StorageFile, Windows.Storage, ContentType=WindowsRuntime])
$stream = AwaitOp ($file.OpenAsync([Windows.Storage.FileAccessMode, Windows.Storage, ContentType=WindowsRuntime]::Read)) ([Windows.Storage.Streams.IRandomAccessStream, Windows.Storage.Streams, ContentType=WindowsRuntime])
$decoder = AwaitOp ([Windows.Graphics.Imaging.BitmapDecoder, Windows.Graphics.Imaging, ContentType=WindowsRuntime]::CreateAsync($stream)) ([Windows.Graphics.Imaging.BitmapDecoder, Windows.Graphics.Imaging, ContentType=WindowsRuntime])
$bitmap = AwaitOp ($decoder.GetSoftwareBitmapAsync()) ([Windows.Graphics.Imaging.SoftwareBitmap, Windows.Graphics.Imaging, ContentType=WindowsRuntime])
if ($bitmap.BitmapPixelFormat -ne [Windows.Graphics.Imaging.BitmapPixelFormat, Windows.Graphics.Imaging, ContentType=WindowsRuntime]::Bgra8 -or $bitmap.BitmapAlphaMode -ne [Windows.Graphics.Imaging.BitmapAlphaMode, Windows.Graphics.Imaging, ContentType=WindowsRuntime]::Premultiplied) {
    $bitmap = [Windows.Graphics.Imaging.SoftwareBitmap, Windows.Graphics.Imaging, ContentType=WindowsRuntime]::Convert($bitmap, [Windows.Graphics.Imaging.BitmapPixelFormat, Windows.Graphics.Imaging, ContentType=WindowsRuntime]::Bgra8, [Windows.Graphics.Imaging.BitmapAlphaMode, Windows.Graphics.Imaging, ContentType=WindowsRuntime]::Premultiplied)
}
$lang = [Windows.Globalization.Language, Windows.Globalization, ContentType=WindowsRuntime]::new('en-US')
$engine = [Windows.Media.Ocr.OcrEngine, Windows.Foundation, ContentType=WindowsRuntime]::TryCreateFromLanguage($lang)
if ($null -eq $engine) {
    $engine = [Windows.Media.Ocr.OcrEngine, Windows.Foundation, ContentType=WindowsRuntime]::TryCreateFromUserProfileLanguages()
}
if ($null -eq $engine) { throw 'No Windows OCR engine available.' }
$result = AwaitOp ($engine.RecognizeAsync($bitmap)) ([Windows.Media.Ocr.OcrResult, Windows.Foundation, ContentType=WindowsRuntime])
$result.Text
'''
        env = os.environ.copy()
        env["PROJECTEF_OCR_IMAGE"] = str(image_path)
        script_path = Path(tempfile.gettempdir()) / f"ProjectEF_AnimationWwiseEvent_WindowsOCR_{time.time_ns()}.ps1"
        script_path.write_text(script, encoding="utf-8")
        try:
            completed = subprocess.run(
                [
                    "powershell.exe",
                    "-NoProfile",
                    "-NonInteractive",
                    "-ExecutionPolicy",
                    "Bypass",
                    "-File",
                    str(script_path),
                ],
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                encoding="utf-8",
                errors="replace",
                env=env,
                timeout=20,
                creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0) if os.name == "nt" else 0,
            )
        finally:
            try:
                script_path.unlink(missing_ok=True)
            except Exception:
                pass
        if completed.returncode != 0:
            raise RuntimeError((completed.stderr or completed.stdout or "Windows OCR failed").strip())
        return completed.stdout.strip()

    def ocr_images_worker(self, images: list[tuple[str, object]]) -> None:
        pytesseract, error = self.prepare_ocr_engine()
        use_windows_ocr = bool(error and self.windows_ocr_available())
        if error and not use_windows_ocr:
            blocks: list[str] = []
            errors: list[str] = [error]
            attachment_dir = core.REPORT_DIR / "TaskCardImages"
            attachment_dir.mkdir(parents=True, exist_ok=True)
            for label, image_like in images:
                try:
                    ocr_path, is_temp = self.prepare_image_for_ocr(image_like)
                    label_path = Path(str(label))
                    if label_path.suffix.lower() in IMAGE_EXTENSIONS and label_path.exists():
                        evidence_label = str(label_path)
                    else:
                        saved_path = attachment_dir / f"ProjectEF_TaskCardImage_{time.time_ns()}.png"
                        shutil.copy2(ocr_path, saved_path)
                        evidence_label = str(saved_path)
                    if is_temp:
                        ocr_path.unlink(missing_ok=True)
                    blocks.append(f"# OCR Image: {evidence_label}\n(OCR unavailable: {error})")
                except Exception as exc:
                    errors.append(f"{label}: {exc}")
            self.messages.put(("ocr_done", {"text": "\n\n".join(blocks), "count": len(blocks), "errors": errors}))
            return

        try:
            from PIL import Image
        except Exception as exc:
            self.messages.put(("ocr_done", {"text": "", "count": 0, "errors": [str(exc)]}))
            return

        blocks: list[str] = []
        errors: list[str] = []
        temp_paths: list[Path] = []
        attachment_dir = core.REPORT_DIR / "TaskCardImages"
        attachment_dir.mkdir(parents=True, exist_ok=True)
        for label, image_like in images:
            try:
                ocr_path, is_temp = self.prepare_image_for_ocr(image_like)
                if is_temp:
                    temp_paths.append(ocr_path)
                label_path = Path(str(label))
                if label_path.suffix.lower() in IMAGE_EXTENSIONS and label_path.exists():
                    evidence_label = str(label_path)
                else:
                    saved_path = attachment_dir / f"ProjectEF_TaskCardImage_{time.time_ns()}.png"
                    shutil.copy2(ocr_path, saved_path)
                    evidence_label = str(saved_path)
                if use_windows_ocr:
                    text = self.windows_ocr_image(ocr_path)
                else:
                    with Image.open(ocr_path) as image:
                        text = pytesseract.image_to_string(image, config="--psm 6")
                blocks.append(f"# OCR Image: {evidence_label}\n{text.strip()}")
            except Exception as exc:
                errors.append(f"{label}: {exc}")
        for path in temp_paths:
            try:
                path.unlink(missing_ok=True)
            except Exception:
                pass

        self.messages.put(("ocr_done", {"text": "\n\n".join(blocks), "count": len(blocks), "errors": errors}))

    def parse_batch_text(self) -> None:
        text = self.batch_text.get("1.0", tk.END)
        entries = core.parse_batch_entries(text, self.event_times_var.get().strip())
        self.batch_items = []
        for entry in entries:
            item = dict(entry)
            item["status"] = "parsed"
            item["preview_argv"] = []
            self.batch_items.append(item)
        self.refresh_batch_tree()
        if not self.batch_items:
            messagebox.showinfo("Batch", "No complete Animation + Event entries were parsed.")
        else:
            self.append_log(f"Parsed {len(self.batch_items)} batch item(s).")

    def refresh_batch_tree(self) -> None:
        for row in self.batch_tree.get_children():
            self.batch_tree.delete(row)
        for index, item in enumerate(self.batch_items, start=1):
            self.batch_tree.insert(
                "",
                tk.END,
                values=(
                    index,
                    item.get("animation", ""),
                    item.get("event", ""),
                    item.get("event_times", ""),
                    item.get("status", ""),
                ),
            )

    def extract_round_inputs(self, text: str) -> list[str]:
        values: list[str] = []

        def add(value: str) -> None:
            item = value.strip().strip("\"'`，,;；。)）]")
            if item and item not in values:
                values.append(item)

        for match in re.finditer(
            r"[A-Za-z]:\\[^\r\n]+?\.(?:prefab|anim|playable|fbx|asset|controller|overridecontroller)\b",
            text,
            flags=re.IGNORECASE,
        ):
            add(match.group(0))
        for match in re.finditer(
            r"Assets[/\\][^\s\"'`，,;；。)）]+?\.(?:prefab|anim|playable|fbx|asset|controller|overridecontroller)\b",
            text,
            flags=re.IGNORECASE,
        ):
            add(match.group(0).replace("\\", "/"))
        for match in re.finditer(
            r"\b[A-Za-z0-9_.-]+\.(?:prefab|anim|playable|fbx|asset|controller|overridecontroller)\b",
            text,
            flags=re.IGNORECASE,
        ):
            add(match.group(0))

        for raw_line in text.splitlines():
            line = raw_line.strip()
            if not line or len(line) > 180:
                continue
            lowered = line.lower()
            if any(token in lowered for token in ("pfb_", "clp_", "tml_", "vfx", "fishsurfacestrike", "catchingfish")):
                add(line)

        animation = self.animation_var.get().strip()
        if animation and animation != DEFAULT_ANIMATION:
            add(animation)
        prefab = self.prefab_var.get().strip()
        if prefab:
            add(prefab)
        return values

    def extract_round_image_inputs(self, text: str) -> list[str]:
        values: list[str] = []

        def add(value: str) -> None:
            item = value.strip().strip("\"'`，,;；。)）]")
            if item and item not in values:
                values.append(item)

        extension_pattern = "|".join(re.escape(ext.lstrip(".")) for ext in sorted(IMAGE_EXTENSIONS))
        for match in re.finditer(
            rf"[A-Za-z]:[/\\][^\r\n]+?\.(?:{extension_pattern})\b",
            text,
            flags=re.IGNORECASE,
        ):
            add(match.group(0))
        for match in re.finditer(
            rf"[^\s\"'`，,;；。)）]+?\.(?:{extension_pattern})\b",
            text,
            flags=re.IGNORECASE,
        ):
            value = match.group(0)
            if "/" in value or "\\" in value:
                add(value)
        return values

    def extract_round_events(self, text: str) -> list[str]:
        events: list[str] = []
        for event in re.findall(r"\bPlay_[A-Za-z0-9_]+\b", text):
            if event not in events:
                events.append(event)
        current = self.wwise_event_var.get().strip()
        if current and current != DEFAULT_WWISE_EVENT and current not in events:
            events.append(current)
        return events

    def infer_round_intent(self, text: str, inputs: list[str]) -> str:
        combined = " ".join([text, *inputs]).lower()
        if any(token in combined for token in ("watersplash", "water splash", "fishsurfacestrike", "waterinteract")):
            return "Water Splash"
        if any(token in combined for token in ("timeline", ".playable", "cutscene", "tml_")):
            return "Timeline"
        if any(token in combined for token in ("vfx", "particle", "effect")):
            return "VFX"
        if any(token in combined for token in ("ui", "button", "界面")):
            return "UI"
        if any(token in combined for token in ("animation", ".anim", ".fbx", "clp_", "bird", "wing")):
            return "Animation"
        return "Auto"

    def generate_task_card(self) -> None:
        if task_card is None:
            messagebox.showerror("Task Card unavailable", "ProjectEF_AudioCodexTaskCard_GUI.py could not be imported.")
            return
        round_text = self.batch_text.get("1.0", tk.END).strip()
        if not round_text:
            messagebox.showinfo("Round Info", "Paste or type this round's known file names, folders, events, or task notes first.")
            return
        asset_inputs = self.extract_round_inputs(round_text)
        image_inputs = self.extract_round_image_inputs(round_text)
        inputs = [*asset_inputs]
        for item in image_inputs:
            if item not in inputs:
                inputs.append(item)
        events = self.extract_round_events(round_text)
        title_source = events[0] if events else (asset_inputs[0] if asset_inputs else (image_inputs[0] if image_inputs else "request"))
        title = f"Round audio config - {Path(title_source).stem or title_source}"
        intent = self.infer_round_intent(round_text, inputs)
        unity_root = core.normalize_path(self.required_text(self.unity_root_var, "Unity Root"))
        self.run_background(
            "Generating Codex task card...",
            lambda: self.generate_task_card_worker(title, intent, unity_root, inputs, events, round_text),
        )

    def generate_task_card_worker(
        self,
        title: str,
        intent: str,
        unity_root: Path,
        inputs: list[str],
        events: list[str],
        notes: str,
    ) -> None:
        assert task_card is not None
        card = task_card.build_task_card(title, intent, unity_root, inputs, events, notes)
        card = task_card.save_task_card(card)
        self.messages.put(("task_card_done", {"card": card, "md": Path(card.md_path), "json": Path(card.json_path)}))

    def refresh_task_candidate_tree(self) -> None:
        tree = getattr(self, "task_candidate_tree", None)
        if not tree:
            return
        for row in tree.get_children():
            tree.delete(row)
        filter_text = ""
        if hasattr(self, "candidate_filter_var"):
            filter_text = self.candidate_filter_var.get().strip().lower()
        for index, candidate in enumerate(self.task_candidates, start=1):
            unity_path = self.candidate_value(candidate, "unity_path") or self.candidate_value(candidate, "path")
            if filter_text:
                haystack = " ".join(
                    [
                        self.candidate_value(candidate, "kind"),
                        self.candidate_value(candidate, "name"),
                        self.candidate_value(candidate, "system"),
                        self.candidate_value(candidate, "group"),
                        unity_path,
                        self.candidate_value(candidate, "prefab_path"),
                        self.candidate_value(candidate, "timeline_path"),
                        self.candidate_value(candidate, "animation_path"),
                        self.candidate_value(candidate, "notes"),
                    ]
                ).lower()
                if filter_text not in haystack:
                    continue
            tree.insert(
                "",
                tk.END,
                values=(
                    index,
                    self.candidate_value(candidate, "kind"),
                    self.candidate_value(candidate, "score"),
                    unity_path,
                ),
            )

    def auto_select_preferred_task_candidate(self) -> None:
        tree = getattr(self, "task_candidate_tree", None)
        if not tree:
            return
        preferred_item = ""
        fallback_item = ""
        for item_id in tree.get_children():
            values = tree.item(item_id, "values")
            if not values:
                continue
            if not fallback_item:
                fallback_item = item_id
            kind = str(values[1] if len(values) > 1 else "")
            if kind in {"Timeline Prefab", "Timeline"}:
                preferred_item = item_id
                break
        target = preferred_item or fallback_item
        if target:
            tree.selection_set(target)
            tree.focus(target)
            tree.see(target)

    def show_task_card_summary_in_round_text(self, source_path: Path) -> None:
        if not hasattr(self, "batch_text"):
            return
        counts: dict[str, int] = {}
        for candidate in self.task_candidates:
            kind = self.candidate_value(candidate, "kind") or "Unknown"
            counts[kind] = counts.get(kind, 0) + 1
        lines = [
            "Loaded latest task card:",
            str(source_path),
            "",
            f"Candidates: {len(self.task_candidates)}",
            "Kinds: " + (", ".join(f"{key}={value}" for key, value in sorted(counts.items())) or "none"),
            "",
            "First candidates:",
        ]
        for index, candidate in enumerate(self.task_candidates[:12], start=1):
            lines.append(
                f"{index}. {self.candidate_value(candidate, 'kind')} "
                f"[{self.candidate_value(candidate, 'score')}] "
                f"{self.candidate_value(candidate, 'unity_path') or self.candidate_value(candidate, 'path')}"
            )
        if any(self.candidate_value(candidate, "kind") == "Timeline Prefab" for candidate in self.task_candidates):
            lines.extend(["", "Auto-selected the first Timeline Prefab. Click Timeline Preview to render it."])
        self.batch_text.delete("1.0", tk.END)
        self.batch_text.insert(tk.END, "\n".join(lines))
        self.batch_text.see("1.0")

    def candidate_value(self, candidate: object, key: str) -> str:
        if isinstance(candidate, dict):
            return str(candidate.get(key) or "")
        return str(getattr(candidate, key, "") or "")

    def apply_task_card_done(self, payload: dict[str, object]) -> None:
        card = payload.get("card")
        self.last_task_card = card
        md = payload.get("md")
        json_path = payload.get("json")
        if isinstance(md, Path):
            self.last_task_card_md = md
        if isinstance(json_path, Path):
            self.last_task_card_json = json_path
        candidates = getattr(card, "candidates", []) if card is not None else []
        self.task_candidates = list(candidates) if isinstance(candidates, list) else []
        self.refresh_task_candidate_tree()
        self.auto_select_preferred_task_candidate()
        if isinstance(self.last_task_card_json, Path):
            self.show_task_card_summary_in_round_text(self.last_task_card_json)
        self.summary_var.set(
            f"Task card generated.\n"
            f"Candidates: {len(self.task_candidates)}\n"
            f"Card: {self.last_task_card_md or 'None'}\n"
            "Select a candidate and click Use Candidate. Then run Preview/Apply as usual."
        )
        self.save_current_round_index("task_card_generated")
        self.append_log(f"Task card generated: {self.last_task_card_md}")
        for index, candidate in enumerate(self.task_candidates[:8], start=1):
            self.append_log(
                f"Candidate {index}: {self.candidate_value(candidate, 'kind')} / "
                f"{self.candidate_value(candidate, 'score')} / {self.candidate_value(candidate, 'unity_path')}"
            )

    def scan_action_index(self) -> None:
        if action_index is None:
            messagebox.showerror("Action Index unavailable", "ProjectEF_ActionResourceIndex.py could not be imported.")
            return
        unity_root = core.normalize_path(self.required_text(self.unity_root_var, "Unity Root"))
        self.run_background(
            "Scanning ProjectEF action resources. This is read-only and may take a minute...",
            lambda: self.scan_action_index_worker(unity_root),
        )

    def scan_action_index_worker(self, unity_root: Path) -> None:
        assert action_index is not None
        result = action_index.build_action_resource_index(unity_root=unity_root)
        self.messages.put(("action_index_done", result))

    def normalize_action_index_rows(self, rows: list[dict[str, object]]) -> list[dict[str, str]]:
        normalized: list[dict[str, str]] = []
        for row in rows:
            item = {str(key): str(value or "") for key, value in row.items()}
            if not item.get("unity_path"):
                for key in ("timeline_path", "animation_path", "prefab_path", "path"):
                    if item.get(key):
                        item["unity_path"] = item[key]
                        break
            if not item.get("path"):
                item["path"] = item.get("unity_path", "")
            if not item.get("kind"):
                preview_kind = item.get("preview_kind", "")
                item["kind"] = preview_kind.title() if preview_kind else "Action Resource"
            if not item.get("score"):
                item["score"] = "0"
            item["source"] = item.get("source") or "action_index"
            normalized.append(item)
        return normalized

    def load_action_index_path(self, path: Path, save_index: bool = True) -> None:
        suffix = path.suffix.lower()
        rows: list[dict[str, object]] = []
        if suffix == ".json":
            data = json.loads(path.read_text(encoding="utf-8-sig"))
            if isinstance(data, dict):
                rows = list(data.get("rows", [])) if isinstance(data.get("rows"), list) else []
            elif isinstance(data, list):
                rows = data
            self.last_action_index_json = path
            xlsx_candidate = path.with_suffix(".xlsx")
            self.last_action_index_xlsx = xlsx_candidate if xlsx_candidate.exists() else None
        elif suffix == ".csv":
            with path.open("r", encoding="utf-8-sig", newline="") as handle:
                rows = list(csv.DictReader(handle))
            self.last_action_index_json = None
            self.last_action_index_xlsx = None
        elif suffix == ".xlsx":
            try:
                from openpyxl import load_workbook
            except Exception as exc:
                raise RuntimeError("openpyxl is required to load xlsx action indexes.") from exc
            workbook = load_workbook(path, read_only=True, data_only=True)
            sheet = workbook.active
            iterator = sheet.iter_rows(values_only=True)
            headers = [str(value or "").strip() for value in next(iterator)]
            for values in iterator:
                rows.append({headers[index]: values[index] if index < len(values) else "" for index in range(len(headers))})
            self.last_action_index_json = None
            self.last_action_index_xlsx = path
        else:
            raise ValueError(f"Unsupported action index file: {path}")

        self.task_candidates = self.normalize_action_index_rows(rows)
        self.refresh_task_candidate_tree()
        self.auto_select_preferred_task_candidate()
        self.summary_var.set(
            f"Action resource index loaded.\n"
            f"Rows: {len(self.task_candidates)}\n"
            f"Index: {path}\n"
            "Select a row, then click Timeline Preview or Direct Preview depending on the kind."
        )
        self.batch_text.delete("1.0", tk.END)
        self.batch_text.insert(
            tk.END,
            "\n".join(
                [
                    "Loaded action resource index:",
                    str(path),
                    "",
                    f"Rows: {len(self.task_candidates)}",
                    "Select a row in the candidate list below.",
                    "Timeline rows use Timeline Preview. AnimationClip rows use Use Candidate -> Direct Preview/Unity Preview.",
                ]
            ),
        )
        self.append_log(f"Loaded action resource index: {path}")
        if save_index:
            self.save_current_round_index("action_index_loaded")

    def apply_action_index_done(self, payload: dict[str, object]) -> None:
        rows = payload.get("rows", [])
        json_path = Path(str(payload.get("json_path") or ""))
        xlsx_path = Path(str(payload.get("xlsx_path") or ""))
        self.last_action_index_json = json_path if json_path.exists() else None
        self.last_action_index_xlsx = xlsx_path if xlsx_path.exists() else None
        self.task_candidates = self.normalize_action_index_rows(rows if isinstance(rows, list) else [])
        self.refresh_task_candidate_tree()
        self.auto_select_preferred_task_candidate()
        self.summary_var.set(
            f"Action resource index generated.\n"
            f"Rows: {len(self.task_candidates)}\n"
            f"Excel: {xlsx_path}\n"
            f"JSON: {json_path}"
        )
        self.batch_text.delete("1.0", tk.END)
        self.batch_text.insert(
            tk.END,
            "\n".join(
                [
                    "Generated action resource index:",
                    f"Excel: {xlsx_path}",
                    f"JSON: {json_path}",
                    "",
                    f"Rows: {len(self.task_candidates)}",
                    "Select a row below and click Timeline Preview / Direct Preview.",
                ]
            ),
        )
        self.save_current_round_index("action_index_generated")
        self.append_log(f"Action index generated: {xlsx_path}")
        self.append_log(f"Action index rows: {len(self.task_candidates)}")

    def load_action_index(self) -> None:
        initial = action_index.REPORT_DIR if action_index is not None else core.REPORT_DIR
        path_text = filedialog.askopenfilename(
            initialdir=str(initial),
            title="Load action resource index",
            filetypes=(
                ("Action Index", "*.json *.xlsx *.csv"),
                ("JSON", "*.json"),
                ("Excel", "*.xlsx"),
                ("CSV", "*.csv"),
                ("All files", "*.*"),
            ),
        )
        if not path_text:
            return
        try:
            self.load_action_index_path(Path(path_text), save_index=True)
        except Exception as exc:
            messagebox.showerror("Load Action Index failed", str(exc))

    def task_card_report_dir(self) -> Path:
        if task_card is not None and hasattr(task_card, "REPORT_DIR"):
            return Path(task_card.REPORT_DIR)
        return Path(core.REPORT_DIR) / "CodexTaskCards"

    def current_round_index_path(self) -> Path:
        return self.task_card_report_dir() / "ProjectEF_AudioTool_CurrentRoundIndex.json"

    def save_current_round_index(self, reason: str = "updated") -> None:
        try:
            index_path = self.current_round_index_path()
            index_path.parent.mkdir(parents=True, exist_ok=True)
            payload = {
                "version": 1,
                "updated_at": datetime.now().isoformat(timespec="seconds"),
                "reason": reason,
                "unity_root": self.unity_root_var.get(),
                "wwise_root": self.wwise_root_var.get(),
                "animation": self.animation_var.get(),
                "prefab": self.prefab_var.get(),
                "wwise_event": self.wwise_event_var.get(),
                "last_task_card_md": str(self.last_task_card_md or ""),
                "last_task_card_json": str(self.last_task_card_json or ""),
                "last_action_index_json": str(self.last_action_index_json or ""),
                "last_action_index_xlsx": str(self.last_action_index_xlsx or ""),
                "last_report_md": str(self.last_report_md or ""),
                "last_report_json": str(self.last_report_json or ""),
                "last_unity_preview_request": str(self.unity_preview_request_path or ""),
                "timeline_preview_request": str(Path(self.unity_root_var.get()) / "Temp" / "ProjectEF_TimelinePrefabPreviewRequest.json"),
                "timeline_preview_frames_dir": str(Path(self.unity_root_var.get()) / "Temp" / "ProjectEF_TimelinePrefabPreviewFrames"),
                "task_candidate_count": len(self.task_candidates),
            }
            index_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception as exc:
            self.append_log(f"Current round index save failed: {exc}")

    def load_task_card_path(self, path: Path, save_index: bool = True) -> None:
        if path.suffix.lower() == ".json":
            data = json.loads(path.read_text(encoding="utf-8-sig"))
            self.last_task_card = data
            self.last_task_card_json = path
            md_path = data.get("md_path") if isinstance(data, dict) else None
            sibling_md = path.with_suffix(".md")
            if md_path and Path(str(md_path)).exists():
                self.last_task_card_md = Path(str(md_path))
            elif sibling_md.exists():
                self.last_task_card_md = sibling_md
            else:
                self.last_task_card_md = None
            self.task_candidates = list(data.get("candidates", [])) if isinstance(data, dict) else []
            self.refresh_task_candidate_tree()
            self.auto_select_preferred_task_candidate()
            self.show_task_card_summary_in_round_text(path)
            self.summary_var.set(
                f"Task card loaded.\n"
                f"Candidates: {len(self.task_candidates)}\n"
                f"Card: {path}\n"
                "Select a candidate, then run Timeline Preview or the matching config flow."
            )
            self.append_log(f"Loaded task card JSON: {path}")
        else:
            self.last_task_card_md = path
            self.last_task_card_json = None
            self.last_task_card = None
            self.task_candidates = []
            self.refresh_task_candidate_tree()
            self.summary_var.set(f"Task card markdown loaded.\nCard: {path}")
            self.append_log(f"Loaded task card markdown: {path}")
            self.append_log(path.read_text(encoding="utf-8-sig", errors="replace")[:4000])
        if save_index:
            self.save_current_round_index("task_card_loaded")

    def find_latest_valid_task_card_json(self) -> Path | None:
        report_dir = self.task_card_report_dir()
        if not report_dir.exists():
            return None
        for path in sorted(report_dir.glob("ProjectEF_AudioCodexTaskCard_*.json"), key=lambda item: item.stat().st_mtime, reverse=True):
            try:
                json.loads(path.read_text(encoding="utf-8-sig"))
                return path
            except Exception:
                self.append_log(f"Skipped invalid task card JSON: {path}")
        return None

    def find_latest_action_index_file(self) -> Path | None:
        if action_index is None:
            return None
        report_dir = Path(action_index.REPORT_DIR)
        if not report_dir.exists():
            return None
        candidates: list[Path] = []
        for pattern in ("ProjectEF_ActionResourceIndex_*.json", "ProjectEF_ActionResourceIndex_*.xlsx", "ProjectEF_ActionResourceIndex_*.csv"):
            candidates.extend(report_dir.glob(pattern))
        if not candidates:
            return None
        return max(candidates, key=lambda item: item.stat().st_mtime)

    def load_latest_task_card(self) -> None:
        index_path = self.current_round_index_path()
        card_path: Path | None = None
        action_index_path: Path | None = None
        if index_path.exists():
            try:
                index = json.loads(index_path.read_text(encoding="utf-8-sig"))
                candidate = Path(str(index.get("last_task_card_json") or ""))
                if candidate.exists():
                    card_path = candidate
                else:
                    candidate = Path(str(index.get("last_task_card_md") or ""))
                    if candidate.exists():
                        card_path = candidate
                if card_path is None:
                    candidate = Path(str(index.get("last_action_index_json") or ""))
                    if candidate.exists():
                        action_index_path = candidate
                    else:
                        candidate = Path(str(index.get("last_action_index_xlsx") or ""))
                        if candidate.exists():
                            action_index_path = candidate
            except Exception as exc:
                self.append_log(f"Current round index read failed: {exc}")

        latest_card_path = self.find_latest_valid_task_card_json()
        latest_action_index_path = self.find_latest_action_index_file()

        if card_path is None:
            card_path = latest_card_path
        if action_index_path is None:
            action_index_path = latest_action_index_path

        if card_path is not None and action_index_path is not None:
            try:
                if action_index_path.stat().st_mtime > card_path.stat().st_mtime:
                    card_path = None
            except OSError:
                pass

        if card_path is None:
            if action_index_path is None:
                messagebox.showinfo("Load Latest", "No valid task card or action index was found. Generate or load one first.")
                return
            try:
                self.load_action_index_path(action_index_path, save_index=True)
                self.append_log(f"Current round index: {index_path}")
            except Exception as exc:
                messagebox.showerror("Load Latest failed", str(exc))
            return

        try:
            self.load_task_card_path(card_path, save_index=True)
            self.append_log(f"Current round index: {index_path}")
        except Exception as exc:
            messagebox.showerror("Load Latest failed", str(exc))

    def load_task_card(self) -> None:
        initial = str((task_card.REPORT_DIR if task_card is not None else core.REPORT_DIR))
        path_text = filedialog.askopenfilename(
            initialdir=initial,
            title="Load Codex task card JSON/MD",
            filetypes=(("Task Card", "*.json *.md"), ("JSON", "*.json"), ("Markdown", "*.md"), ("All files", "*.*")),
        )
        if not path_text:
            return
        self.load_task_card_path(Path(path_text), save_index=True)

    def copy_task_card(self) -> None:
        if self.last_task_card_md and self.last_task_card_md.exists():
            text = self.last_task_card_md.read_text(encoding="utf-8-sig", errors="replace")
        elif self.last_task_card is not None and task_card is not None and not isinstance(self.last_task_card, dict):
            text = task_card.render_markdown(self.last_task_card)
        elif self.last_task_card_json and self.last_task_card_json.exists():
            sibling_md = self.last_task_card_json.with_suffix(".md")
            if sibling_md.exists():
                text = sibling_md.read_text(encoding="utf-8-sig", errors="replace")
            else:
                text = self.last_task_card_json.read_text(encoding="utf-8-sig", errors="replace")
        else:
            messagebox.showinfo("Task Card", "Generate or load a task card first.")
            return
        self.clipboard_clear()
        self.clipboard_append(text)
        self.append_log("Task card copied to clipboard.")

    def selected_task_candidate(self) -> object | None:
        tree = getattr(self, "task_candidate_tree", None)
        if not tree:
            return None
        selection = tree.selection()
        if not selection:
            return None
        values = tree.item(selection[0], "values")
        try:
            index = int(values[0]) - 1
        except Exception:
            return None
        if 0 <= index < len(self.task_candidates):
            return self.task_candidates[index]
        return None

    def use_selected_task_candidate(self) -> None:
        candidate = self.selected_task_candidate()
        if candidate is None:
            messagebox.showinfo("Task Card", "Select a task-card candidate first.")
            return
        path_value = self.candidate_value(candidate, "path") or self.candidate_value(candidate, "unity_path")
        prefab_path = self.candidate_value(candidate, "prefab_path")
        timeline_path = self.candidate_value(candidate, "timeline_path")
        animation_path = self.candidate_value(candidate, "animation_path")
        kind = self.candidate_value(candidate, "kind")
        applied_animation = animation_path or (path_value if Path(path_value).suffix.lower() in {".anim", ".fbx"} else "")
        applied_prefab = prefab_path or (path_value if Path(path_value).suffix.lower() == ".prefab" else "")

        if applied_animation:
            self.animation_var.set(applied_animation)
            self.append_log(f"Candidate applied to Animation: {applied_animation}")
        if applied_prefab:
            self.prefab_var.set(applied_prefab)
            self.append_log(f"Candidate applied to Prefab: {applied_prefab}")
        if timeline_path:
            self.append_log(f"Candidate Timeline ready for Timeline Preview: {timeline_path}")

        if timeline_path and applied_prefab and not applied_animation:
            self.status_var.set("Timeline candidate selected. Click Timeline Preview.")
            self.save_current_round_index("timeline_candidate_selected")
            return
        if applied_prefab and not applied_animation and not timeline_path:
            self.status_var.set("Prefab candidate selected. Use the matching preview/config flow for this asset type.")
            self.save_current_round_index("prefab_candidate_selected")
            return
        elif applied_animation:
            pass
        elif applied_prefab:
            pass
        else:
            self.append_log(f"Candidate selected for Codex analysis, but this animation tool cannot write {kind}: {path_value}")

        card = self.last_task_card
        events = []
        if isinstance(card, dict):
            events = card.get("events", []) if isinstance(card.get("events"), list) else []
        elif card is not None:
            events = getattr(card, "events", [])
        for event in events:
            event_text = str(event)
            if event_text and not event_text.startswith("TODO"):
                self.wwise_event_var.set(event_text)
                break
        self.resolve_assets()

    def preview_selected_candidate(self) -> None:
        candidate = self.selected_task_candidate()
        if candidate is None:
            messagebox.showinfo("Preview Selected", "Select a task-card or action-index candidate first.")
            return

        path_value = self.candidate_value(candidate, "unity_path") or self.candidate_value(candidate, "path")
        prefab_path = self.candidate_value(candidate, "prefab_path")
        timeline_path = self.candidate_value(candidate, "timeline_path")
        animation_path = self.candidate_value(candidate, "animation_path")
        preview_kind = self.candidate_value(candidate, "preview_kind").lower()
        suffix = Path(path_value).suffix.lower()

        if timeline_path or preview_kind == "timeline" or suffix == ".playable":
            self.open_timeline_prefab_preview()
            return

        if animation_path or suffix in {".anim", ".fbx"}:
            target_animation = animation_path or path_value
            self.animation_var.set(target_animation)
            if prefab_path:
                self.prefab_var.set(prefab_path)
            self.append_log(f"Preview selected animation: {target_animation}")
            self.preview_config()
            return

        if prefab_path or suffix == ".prefab":
            target_prefab = prefab_path or path_value
            self.prefab_var.set(target_prefab)
            messagebox.showinfo(
                "Preview Selected",
                "This row is a prefab/VFX candidate without a matched Timeline or AnimationClip.\n\n"
                "For now it is selected in the Prefab field. Use Unity's prefab/VFX preview or provide the owning Timeline/Animation.",
            )
            self.append_log(f"Selected prefab/VFX candidate: {target_prefab}")
            return

        messagebox.showinfo("Preview Selected", f"No supported preview route for this row: {path_value}")

    def ensure_batch_items(self) -> bool:
        if not self.batch_items:
            self.parse_batch_text()
        return bool(self.batch_items)

    def preview_batch(self) -> None:
        if not self.ensure_batch_items():
            return
        for item in self.batch_items:
            item["status"] = "preview pending"
            item["preview_argv"] = []
        self.refresh_batch_tree()
        self.run_background("Running batch preview...", lambda: self.run_batch_worker(apply=False))

    def apply_batch(self) -> None:
        if not self.ensure_batch_items():
            return
        stale_items: list[int] = []
        not_previewed = [
            item
            for item in self.batch_items
            if item.get("status") not in {"preview ok", "applied"} or not item.get("preview_argv")
        ]
        for index, item in enumerate(self.batch_items, start=1):
            preview_argv = item.get("preview_argv")
            if not preview_argv:
                continue
            if preview_argv != self.build_batch_tool_argv(item, apply=False):
                item["status"] = "preview stale"
                stale_items.append(index)
        if not_previewed:
            self.refresh_batch_tree()
            messagebox.showinfo(
                "Batch preflight required",
                "Run Batch Preview first. Apply is enabled only after every parsed item previews successfully.",
            )
            return
        if stale_items:
            self.refresh_batch_tree()
            messagebox.showinfo(
                "Batch preview stale",
                "Batch settings changed after preview. Run Batch Preview again before applying.",
            )
            return
        if not messagebox.askyesno("Batch Apply", f"Write Animation Events for {len(self.batch_items)} batch item(s)?"):
            return
        for item in self.batch_items:
            item["status"] = "apply pending"
        self.refresh_batch_tree()
        self.run_background("Running batch apply...", lambda: self.run_batch_worker(apply=True))

    def build_batch_tool_argv(self, item: dict[str, str], apply: bool) -> list[str]:
        sample_fps = float(self.required_text(self.sample_fps_var, "采样 FPS"))
        strength_ratio = float(self.required_text(self.strength_ratio_var, "强度阈值"))
        min_gap = float(self.required_text(self.min_gap_var, "最小间隔"))
        argv = [
            "--unity-root",
            self.required_text(self.unity_root_var, "Unity Root"),
            "--wwise-root",
            self.required_text(self.wwise_root_var, "Wwise Root"),
            "--animation",
            item["animation"],
            "--wwise-event",
            item["event"],
            "--mode",
            item.get("mode") or self.mode_var.get(),
            "--endpoint-regex",
            self.required_text(self.endpoint_regex_var, "端点匹配"),
            "--sample-fps",
            str(sample_fps),
            "--strength-ratio",
            str(strength_ratio),
            "--min-gap",
            str(min_gap),
        ]
        event_times = item.get("event_times", "").strip() or self.event_times_var.get().strip()
        if event_times:
            argv.extend(["--event-times", event_times])
        prefab = item.get("prefab", "").strip() or self.prefab_var.get().strip()
        if prefab:
            argv.extend(["--prefab", prefab])
        if self.skip_prefab_component_var.get():
            argv.append("--skip-prefab-component")
        if not self.audio_aware_spacing_var.get():
            argv.append("--disable-audio-aware-spacing")
        if apply:
            argv.append("--apply")
            if self.p4_reconcile_var.get():
                argv.append("--p4-reconcile")
        return argv

    def build_apply_argv_from_preview(self, item: dict[str, object]) -> list[str]:
        preview_argv = item.get("preview_argv")
        if not isinstance(preview_argv, list) or not preview_argv:
            return self.build_batch_tool_argv(item, apply=True)  # type: ignore[arg-type]
        argv = [str(value) for value in preview_argv]
        argv.append("--apply")
        if self.p4_reconcile_var.get():
            argv.append("--p4-reconcile")
        return argv

    def run_batch_worker(self, apply: bool) -> None:
        results: list[dict[str, object]] = []
        creationflags = getattr(subprocess, "CREATE_NO_WINDOW", 0) if os.name == "nt" else 0
        env = os.environ.copy()
        env["PYTHONIOENCODING"] = "utf-8"
        env["PYTHONUTF8"] = "1"

        for index, item in enumerate(list(self.batch_items)):
            argv = self.build_apply_argv_from_preview(item) if apply else self.build_batch_tool_argv(item, apply=False)
            command = [sys.executable, "-B", str(CORE_SCRIPT), *argv]
            completed = subprocess.run(
                command,
                cwd=str(APP_DIR),
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                encoding="utf-8",
                errors="replace",
                creationflags=creationflags,
                env=env,
            )
            output = completed.stdout
            if completed.stderr:
                output = output.rstrip() + "\n\n[stderr]\n" + completed.stderr
            ok = completed.returncode == 0
            paths = self.extract_report_paths(output) if ok else {}
            result = {
                "index": index,
                "ok": ok,
                "apply": apply,
                "output": output,
                "md": paths.get("md"),
                "json": paths.get("json"),
                "command": command,
                "argv": argv,
            }
            results.append(result)
            self.messages.put(("batch_progress", result))

        self.messages.put(("batch_done", {"apply": apply, "results": results}))

    def resolve_assets(self) -> None:
        self.run_background("正在定位资源...", self.resolve_worker)

    def resolve_worker(self) -> None:
        data = self.resolve_current()
        self.messages.put(("resolved", data))

    def preview_config(self) -> None:
        self.run_background("正在预览，不写入 Unity 文件...", lambda: self.run_tool_worker(apply=False))

    def apply_config(self) -> None:
        if not messagebox.askyesno(
            "确认写入",
            "这会修改目标 .anim，并按需要确保 AnimationWwiseEventReceiver 脚本和 Prefab 接收器配置存在。\n\n继续写入吗？",
        ):
            return
        self.run_background("正在写入 / 修改配置...", lambda: self.run_tool_worker(apply=True))

    def resolve_current(self) -> dict[str, object]:
        unity_root = core.normalize_path(self.required_text(self.unity_root_var, "Unity Root"))
        wwise_root = core.normalize_path(self.required_text(self.wwise_root_var, "Wwise Root"))
        animation_query = self.required_text(self.animation_var, "Animation")
        event_name = self.required_text(self.wwise_event_var, "Wwise Event")
        prefab_query = self.prefab_var.get().strip() or None

        animation_path, source_animation_path = core.resolve_animation_asset(unity_root, animation_query)
        prefab_path = core.resolve_prefab(unity_root, animation_path, prefab_query)
        wwise_validation = core.validate_wwise_event(wwise_root, event_name)

        return {
            "animation": animation_path,
            "source_animation": source_animation_path,
            "prefab": prefab_path,
            "wwise_valid": wwise_validation[0],
            "wwise_evidence": wwise_validation[1],
        }

    def build_tool_argv(self, apply: bool) -> list[str]:
        sample_fps = float(self.required_text(self.sample_fps_var, "采样 FPS"))
        strength_ratio = float(self.required_text(self.strength_ratio_var, "强度阈值"))
        min_gap = float(self.required_text(self.min_gap_var, "最小间隔"))
        argv = [
            "--unity-root",
            self.required_text(self.unity_root_var, "Unity Root"),
            "--wwise-root",
            self.required_text(self.wwise_root_var, "Wwise Root"),
            "--animation",
            self.required_text(self.animation_var, "Animation"),
            "--wwise-event",
            self.required_text(self.wwise_event_var, "Wwise Event"),
            "--mode",
            self.mode_var.get(),
            "--endpoint-regex",
            self.required_text(self.endpoint_regex_var, "端点匹配"),
            "--sample-fps",
            str(sample_fps),
            "--strength-ratio",
            str(strength_ratio),
            "--min-gap",
            str(min_gap),
        ]
        event_times = self.event_times_var.get().strip()
        if event_times:
            argv.extend(["--event-times", event_times])
        prefab = self.prefab_var.get().strip()
        if prefab:
            argv.extend(["--prefab", prefab])
        if self.skip_prefab_component_var.get():
            argv.append("--skip-prefab-component")
        if not self.audio_aware_spacing_var.get():
            argv.append("--disable-audio-aware-spacing")
        if apply:
            argv.append("--apply")
            if self.p4_reconcile_var.get():
                argv.append("--p4-reconcile")
        return argv

    def run_tool_worker(self, apply: bool) -> None:
        argv = self.build_tool_argv(apply)
        command = [sys.executable, "-B", str(CORE_SCRIPT), *argv]
        self.last_command = command
        creationflags = getattr(subprocess, "CREATE_NO_WINDOW", 0) if os.name == "nt" else 0
        env = os.environ.copy()
        env["PYTHONIOENCODING"] = "utf-8"
        env["PYTHONUTF8"] = "1"
        completed = subprocess.run(
            command,
            cwd=str(APP_DIR),
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            encoding="utf-8",
            errors="replace",
            creationflags=creationflags,
            env=env,
        )
        output = completed.stdout
        if completed.stderr:
            output = output.rstrip() + "\n\n[stderr]\n" + completed.stderr
        if completed.returncode != 0:
            raise RuntimeError(output or f"Tool failed with exit code {completed.returncode}")

        paths = self.extract_report_paths(output)
        report = self.load_report(paths.get("json"))
        preview_data = self.build_preview_data(report) if report else None
        self.messages.put(
            (
                "run_done",
                {
                    "apply": apply,
                    "output": output,
                    "report": report,
                    "preview_data": preview_data,
                    "md": paths.get("md"),
                    "json": paths.get("json"),
                    "command": command,
                },
            )
        )

    def required_text(self, variable: tk.StringVar, label: str) -> str:
        value = variable.get().strip()
        if not value:
            raise ValueError(f"{label} 不能为空。")
        return value

    def extract_report_paths(self, output: str) -> dict[str, Path]:
        result: dict[str, Path] = {}
        report_match = re.search(r"^Report:\s*(.+)$", output, re.M)
        json_match = re.search(r"^JSON:\s*(.+)$", output, re.M)
        if report_match:
            result["md"] = Path(report_match.group(1).strip())
        if json_match:
            result["json"] = Path(json_match.group(1).strip())
        return result

    def load_report(self, json_path: Path | None) -> dict[str, object] | None:
        if json_path and json_path.exists():
            return json.loads(json_path.read_text(encoding="utf-8-sig"))
        event_name = self.wwise_event_var.get().strip()
        safe_event = re.sub(r"[^A-Za-z0-9_.-]+", "_", event_name)
        candidates = sorted(
            core.REPORT_DIR.glob(f"ProjectEF_AnimationWwiseEvent_AutoConfig_{safe_event}_*.json"),
            key=lambda path: path.stat().st_mtime,
            reverse=True,
        )
        if candidates:
            return json.loads(candidates[0].read_text(encoding="utf-8-sig"))
        return None

    def build_preview_data(self, report: dict[str, object] | None) -> dict[str, object] | None:
        if not report:
            return None
        analysis = report.get("analysis") if isinstance(report.get("analysis"), dict) else {}
        animation_path = Path(str(report.get("animation") or ""))
        prefab_value = report.get("prefab")
        prefab_path = Path(str(prefab_value)) if prefab_value else None
        event_times = analysis.get("event_times", []) if isinstance(analysis, dict) else []
        endpoint_paths = analysis.get("endpoint_paths", []) if isinstance(analysis, dict) else []
        audio_files = self.find_wwise_audio_files(Path(self.wwise_root_var.get()), str(report.get("wwise_event") or ""))

        samples: list[tuple[float, float]] = []
        clip_length = float(analysis.get("clip_length") or 0.0) if isinstance(analysis, dict) else 0.0
        try:
            _text, parsed_clip_length, curves = core.parse_animation_clip(animation_path)
            clip_length = parsed_clip_length or clip_length
            if prefab_path and prefab_path.exists() and isinstance(endpoint_paths, list) and endpoint_paths:
                root_names = {curve.path.split("/")[0] for curve in curves if curve.path}
                transforms, _prefab_text = core.parse_prefab_transforms(prefab_path, root_names)
                curve_map = {curve.path: curve for curve in curves}
                max_samples = min(max(int(clip_length * 30), 120), 2400)
                for index in range(max_samples + 1):
                    t = clip_length * index / max_samples
                    values = []
                    for endpoint in endpoint_paths:
                        endpoint_text = str(endpoint)
                        if endpoint_text in transforms:
                            values.append(core.world_pos_at(endpoint_text, t, transforms, curve_map)[1])
                    if values:
                        samples.append((t, sum(values) / len(values)))
        except Exception:
            samples = []

        return {
            "clip_length": clip_length,
            "event_times": [float(value) for value in event_times] if isinstance(event_times, list) else [],
            "samples": samples,
            "audio_files": audio_files,
        }

    def find_wwise_audio_files(self, wwise_root: Path, event_name: str) -> list[Path]:
        if not wwise_root.exists() or not event_name:
            return []

        target_names: set[str] = set()
        event_pattern = re.compile(rf'<Event\s+Name="{re.escape(event_name)}"(?=[\s>/]).*?</Event>', re.S)
        events_root = wwise_root / "Events"
        search_roots = [events_root] if events_root.exists() else [wwise_root]
        for root in search_roots:
            for path in root.rglob("*.wwu"):
                try:
                    text = core.read_text(path)
                except Exception:
                    continue
                for event_match in event_pattern.finditer(text):
                    target_names.update(re.findall(r'<ObjectRef\s+Name="([^"]+)"', event_match.group(0)))

        audio_names: set[str] = set()
        for path in wwise_root.rglob("*.wwu"):
            try:
                text = core.read_text(path)
            except Exception:
                continue
            for target_name in target_names:
                marker = f'Name="{target_name}"'
                index = text.find(marker)
                if index >= 0:
                    block = text[index : index + 300000]
                    audio_names.update(re.findall(r"<AudioFile>([^<]+\.wav)</AudioFile>", block, re.I))

        originals = wwise_root / "Originals"
        if not originals.exists():
            return []

        matches: list[Path] = []
        if audio_names:
            lower_names = {name.lower() for name in audio_names}
            matches = [path for path in originals.rglob("*.wav") if path.name.lower() in lower_names]

        if not matches:
            base = re.sub(r"^Play_", "", event_name, flags=re.I).lower()
            matches = [path for path in originals.rglob("*.wav") if base in path.stem.lower()]

        return sorted(matches)[:32]

    def run_background(self, status: str, target) -> None:
        if self.worker and self.worker.is_alive():
            messagebox.showinfo("正在执行", "当前任务还没结束，请稍等。")
            return
        self.append_log(status)
        self.set_running(True, status)

        def wrapped() -> None:
            try:
                target()
            except Exception as exc:
                self.messages.put(("error", {"error": exc, "trace": traceback.format_exc()}))
            finally:
                self.messages.put(("idle", None))

        self.worker = threading.Thread(target=wrapped, daemon=True)
        self.worker.start()

    def pump_messages(self) -> None:
        try:
            while True:
                kind, payload = self.messages.get_nowait()
                if kind == "resolved":
                    self.apply_resolved(payload)  # type: ignore[arg-type]
                elif kind == "run_done":
                    self.apply_run_done(payload)  # type: ignore[arg-type]
                elif kind == "ocr_done":
                    self.apply_ocr_done(payload)  # type: ignore[arg-type]
                elif kind == "batch_progress":
                    self.apply_batch_progress(payload)  # type: ignore[arg-type]
                elif kind == "batch_done":
                    self.apply_batch_done(payload)  # type: ignore[arg-type]
                elif kind == "task_card_done":
                    self.apply_task_card_done(payload)  # type: ignore[arg-type]
                elif kind == "action_index_done":
                    self.apply_action_index_done(payload)  # type: ignore[arg-type]
                elif kind == "error":
                    self.apply_error(payload)  # type: ignore[arg-type]
                elif kind == "idle":
                    self.set_running(False, "Ready")
        except queue.Empty:
            pass
        self.after(100, self.pump_messages)

    def apply_resolved(self, payload: dict[str, object]) -> None:
        animation = payload.get("animation")
        source_animation = payload.get("source_animation")
        prefab = payload.get("prefab")
        evidence = payload.get("wwise_evidence")
        valid = bool(payload.get("wwise_valid"))
        self.resolved_animation_var.set(str(animation) if animation else "未定位")
        self.source_animation_var.set(str(source_animation) if source_animation else "无源 FBX 或直接使用 .anim")
        self.resolved_prefab_var.set(str(prefab) if prefab else "未找到 / 未指定")
        self.wwise_evidence_var.set(str(evidence) if evidence else "未校验")
        self.summary_var.set(
            "资源定位完成。\n"
            f"Wwise Event: {'有效' if valid else '未找到'}\n"
            "下一步可以点“直接预览”查看打点，或调整分析参数后再预览。"
        )
        self.append_log("定位完成。")
        self.append_log(f"Editable .anim: {animation}")
        if source_animation:
            self.append_log(f"Source .fbx: {source_animation}")
        self.append_log(f"Prefab: {prefab or 'None'}")
        self.append_log(f"Wwise: {'OK' if valid else 'Missing'} - {evidence}")

    def apply_run_done(self, payload: dict[str, object]) -> None:
        output = str(payload.get("output") or "")
        report = payload.get("report")
        preview_data = payload.get("preview_data")
        md = payload.get("md")
        json_path = payload.get("json")
        apply = bool(payload.get("apply"))
        if isinstance(md, Path):
            self.last_report_md = md
        if isinstance(json_path, Path):
            self.last_report_json = json_path
        self.append_log(output)

        if isinstance(report, dict):
            self.last_report = report
            self.update_preview_from_report(report, apply, preview_data if isinstance(preview_data, dict) else None)
            self.write_unity_preview_request(report)
            if self.auto_open_unity_preview_var.get() and not apply:
                self.open_unity_preview()
        else:
            self.update_preview_from_stdout(output)
            self.summary_var.set("执行完成，但没有读到 JSON 报告；已从命令输出里提取基础打点。请查看日志。")
        self.status_var.set("写入完成" if apply else "预览完成")
        if apply:
            messagebox.showinfo("完成", "Animation Wwise Event 配置已写入。")

    def apply_ocr_done(self, payload: dict[str, object]) -> None:
        text = str(payload.get("text") or "")
        count = int(payload.get("count") or 0)
        errors = payload.get("errors") if isinstance(payload.get("errors"), list) else []
        if text.strip():
            self.append_batch_text(text)
            self.append_log(f"OCR added text from {count} image(s).")
            self.parse_batch_text()
        if errors:
            message = "\n".join(str(error) for error in errors)
            self.append_log("OCR issue:\n" + message)
            if not text.strip():
                messagebox.showinfo("OCR unavailable", message)

    def apply_batch_progress(self, payload: dict[str, object]) -> None:
        index = int(payload.get("index") or 0)
        ok = bool(payload.get("ok"))
        apply = bool(payload.get("apply"))
        status = ("applied" if apply else "preview ok") if ok else "failed"
        if 0 <= index < len(self.batch_items):
            self.batch_items[index]["status"] = status
            if not apply:
                self.batch_items[index]["preview_argv"] = payload.get("argv") if ok else []
            if isinstance(payload.get("json"), Path):
                self.batch_items[index]["report_json"] = str(payload.get("json"))
            if isinstance(payload.get("md"), Path):
                self.batch_items[index]["report_md"] = str(payload.get("md"))
        self.refresh_batch_tree()

        output = str(payload.get("output") or "")
        item = self.batch_items[index] if 0 <= index < len(self.batch_items) else {}
        self.append_log(f"Batch {index + 1}: {status} - {item.get('animation', '')} / {item.get('event', '')}")
        if not ok and output:
            self.append_log(output)

        json_path = payload.get("json")
        md_path = payload.get("md")
        if isinstance(md_path, Path):
            self.last_report_md = md_path
        if isinstance(json_path, Path):
            self.last_report_json = json_path
            report = self.load_report(json_path)
            if isinstance(report, dict):
                self.last_report = report
                self.update_preview_from_report(report, apply, self.build_preview_data(report))

    def on_batch_tree_select(self, _event: object | None = None) -> None:
        selection = self.batch_tree.selection()
        if not selection:
            return
        values = self.batch_tree.item(selection[0], "values")
        try:
            index = int(values[0]) - 1
        except Exception:
            return
        if not (0 <= index < len(self.batch_items)):
            return
        item = self.batch_items[index]
        json_value = item.get("report_json")
        if not json_value:
            return
        json_path = Path(str(json_value))
        if not json_path.exists():
            return
        report = self.load_report(json_path)
        if not isinstance(report, dict):
            return
        self.last_report = report
        self.last_report_json = json_path
        md_value = item.get("report_md")
        if md_value:
            self.last_report_md = Path(str(md_value))
        self.update_preview_from_report(report, item.get("status") == "applied", self.build_preview_data(report))
        self.append_log(f"Loaded preview for batch row {index + 1}: {item.get('animation', '')}")

    def apply_batch_done(self, payload: dict[str, object]) -> None:
        apply = bool(payload.get("apply"))
        results = payload.get("results") if isinstance(payload.get("results"), list) else []
        ok_count = sum(1 for result in results if isinstance(result, dict) and result.get("ok"))
        fail_count = len(results) - ok_count
        summary_md, summary_json = self.write_batch_summary_report(apply, results)
        self.last_report_md = summary_md
        self.last_report_json = summary_json
        self.append_log(f"Batch complete: {ok_count} ok, {fail_count} failed.")
        self.append_log(f"Batch summary: {summary_md}")
        self.status_var.set("Batch apply complete" if apply else "Batch preview complete")
        if apply:
            if fail_count:
                messagebox.showwarning("Batch complete", f"Applied {ok_count}; {fail_count} failed. Check the log before testing.")
            else:
                messagebox.showinfo("Batch complete", f"Configured {ok_count} Animation Wwise Event item(s). Ready to test.")

    def write_batch_summary_report(self, apply: bool, results: list[object]) -> tuple[Path, Path]:
        core.REPORT_DIR.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y-%m-%d_%H%M%S_%f")
        mode = "Apply" if apply else "Preview"
        json_path = core.REPORT_DIR / f"ProjectEF_AnimationWwiseEvent_Batch_{mode}_{stamp}.json"
        md_path = json_path.with_suffix(".md")
        rows: list[dict[str, object]] = []
        for result in results:
            if not isinstance(result, dict):
                continue
            index = int(result.get("index") or 0)
            item = self.batch_items[index] if 0 <= index < len(self.batch_items) else {}
            rows.append(
                {
                    "index": index + 1,
                    "ok": bool(result.get("ok")),
                    "status": item.get("status", "unknown"),
                    "animation": item.get("animation", ""),
                    "wwise_event": item.get("event", ""),
                    "event_times": item.get("event_times", ""),
                    "report_md": str(result.get("md") or ""),
                    "report_json": str(result.get("json") or ""),
                    "command": subprocess.list2cmdline([str(value) for value in result.get("command", [])])
                    if isinstance(result.get("command"), list)
                    else "",
                }
            )
        summary = {
            "timestamp": datetime.now().isoformat(timespec="seconds"),
            "mode": mode.lower(),
            "ok_count": sum(1 for row in rows if row["ok"]),
            "fail_count": sum(1 for row in rows if not row["ok"]),
            "items": rows,
        }
        json_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
        md_lines = [
            f"# ProjectEF Animation Wwise Event Batch {mode}",
            "",
            f"- Timestamp: {summary['timestamp']}",
            f"- OK: {summary['ok_count']}",
            f"- Failed: {summary['fail_count']}",
            "",
            "| # | Status | Animation | Wwise Event | Times | Report |",
            "|---:|---|---|---|---|---|",
        ]
        for row in rows:
            report = row["report_md"] or row["report_json"]
            md_lines.append(
                f"| {row['index']} | {'OK' if row['ok'] else 'FAILED'} / {row['status']} | "
                f"`{row['animation']}` | `{row['wwise_event']}` | `{row['event_times']}` | `{report}` |"
            )
        md_path.write_text("\n".join(md_lines) + "\n", encoding="utf-8")
        return md_path, json_path

    def update_preview_from_report(
        self,
        report: dict[str, object],
        apply: bool,
        preview_data: dict[str, object] | None,
    ) -> None:
        analysis = report.get("analysis") if isinstance(report.get("analysis"), dict) else {}
        wwise_design = report.get("wwise_design") if isinstance(report.get("wwise_design"), dict) else {}
        animation = str(report.get("animation") or "")
        prefab = str(report.get("prefab") or "")
        evidence = str(report.get("wwise_event_evidence") or "")
        event_name = str(report.get("wwise_event") or "")
        event_times = analysis.get("event_times", []) if isinstance(analysis, dict) else []
        endpoint_paths = analysis.get("endpoint_paths", []) if isinstance(analysis, dict) else []
        analysis_notes = analysis.get("notes", []) if isinstance(analysis, dict) else []
        changed_files = report.get("changed_files", [])

        self.resolved_animation_var.set(animation or "未定位")
        self.source_animation_var.set(str(report.get("source_animation") or "无源 FBX 或直接使用 .anim"))
        self.resolved_prefab_var.set(prefab or "未找到 / 未指定")
        self.wwise_evidence_var.set(evidence)

        for row in self.event_tree.get_children():
            self.event_tree.delete(row)
        if isinstance(event_times, list):
            for index, time_value in enumerate(event_times, start=1):
                self.event_tree.insert("", tk.END, values=(index, f"{float(time_value):.3f}"))

        endpoint_lines = []
        if isinstance(analysis_notes, list) and analysis_notes:
            endpoint_lines.append("Analysis notes:")
            endpoint_lines.extend(f"- {note}" for note in analysis_notes)
            endpoint_lines.append("")
        if isinstance(endpoint_paths, list):
            endpoint_lines.extend(str(path) for path in endpoint_paths)
        if isinstance(changed_files, list):
            endpoint_lines.append("")
            endpoint_lines.append("Changed files:" if apply else "Planned changed files:")
            endpoint_lines.extend(str(path) for path in changed_files)
        if isinstance(wwise_design, dict):
            endpoint_lines.append("")
            endpoint_lines.append("Wwise design:")
            endpoint_lines.append(f"target: {wwise_design.get('target_type')} / {', '.join(wwise_design.get('target_names', []))}")
            endpoint_lines.append(f"audio sources: {wwise_design.get('audio_source_count')}")
            endpoint_lines.append(
                f"duration range: {self.format_seconds(wwise_design.get('min_duration'))} - "
                f"{self.format_seconds(wwise_design.get('max_duration'))}"
            )
            endpoint_lines.append(f"effective min gap: {self.format_seconds(wwise_design.get('effective_min_gap'))}")
            for note in wwise_design.get("notes", []) if isinstance(wwise_design.get("notes"), list) else []:
                endpoint_lines.append(f"- {note}")
        self.set_text(self.endpoint_text, "\n".join(endpoint_lines) or "No endpoint data.")
        self.preview_data = preview_data
        self.preview_visual_mode = "animation"
        self.stop_preview_playback(redraw=False)
        self.draw_preview(0.0)

        count = len(event_times) if isinstance(event_times, list) else 0
        valid = bool(report.get("wwise_event_valid"))
        action = "已写入" if apply else "预览"
        self.summary_var.set(
            f"{action}: {event_name}\n"
            f"打点数量: {count}\n"
            f"Wwise 校验: {'有效' if valid else '未找到'}\n"
            f"模式: {analysis.get('mode')} / {analysis.get('metric') if isinstance(analysis, dict) else ''}\n"
            f"选择策略: {analysis.get('selection_policy') if isinstance(analysis, dict) else ''}\n"
            f"实际间隔: {self.format_seconds(wwise_design.get('effective_min_gap') if isinstance(wwise_design, dict) else None)}\n"
            f"报告: {self.last_report_md or 'None'}"
        )

    def format_seconds(self, value: object) -> str:
        try:
            if value is None:
                return "unknown"
            return f"{float(value):.3f}s"
        except Exception:
            return str(value)

    def update_preview_from_stdout(self, output: str) -> None:
        event_line = re.search(r"Event times:\s*\n([0-9.,\s]+)", output, re.M)
        if not event_line:
            return
        times = [float(value) for value in re.findall(r"[0-9]+(?:\.[0-9]+)?", event_line.group(1))]
        for row in self.event_tree.get_children():
            self.event_tree.delete(row)
        for index, time_value in enumerate(times, start=1):
            self.event_tree.insert("", tk.END, values=(index, f"{time_value:.3f}"))
        clip_length = max(times) if times else 0.0
        self.preview_data = {"clip_length": clip_length, "event_times": times, "samples": [], "audio_files": []}
        self.preview_visual_mode = "animation"
        self.draw_preview(0.0)

    def toggle_preview_playback(self) -> None:
        if self.preview_visual_mode == "timeline_frames" and self.timeline_preview_frames:
            self.timeline_preview_playing = not self.timeline_preview_playing
            self.status_var.set("Timeline preview playing" if self.timeline_preview_playing else "Timeline preview paused")
            return
        if not self.preview_data:
            messagebox.showinfo("暂无预览", "请先点击“直接预览”。")
            return
        if self.playing:
            self.play_base_time = self.current_preview_time()
            self.playing = False
            self.status_var.set("预览暂停")
            return
        self.playing = True
        self.play_start_perf = time.perf_counter()
        self.next_event_index = self.next_event_index_for_time(self.play_base_time)
        self.status_var.set("正在播放预览")
        self.playback_tick()

    def stop_preview_playback(self, redraw: bool = True) -> None:
        if self.preview_visual_mode == "timeline_frames" and self.timeline_preview_frames:
            self.timeline_preview_playing = False
            self.timeline_preview_frame_index = 0
            self.stop_async_audio()
            if redraw:
                self.draw_timeline_preview_frame()
                self.status_var.set("Timeline preview stopped")
            return
        self.playing = False
        self.play_base_time = 0.0
        self.next_event_index = 0
        self.stop_async_audio()
        if redraw:
            self.draw_preview(0.0)
            self.status_var.set("预览停止")

    def current_preview_time(self) -> float:
        if not self.preview_data:
            return 0.0
        if not self.playing:
            return self.play_base_time
        speed = float(self.playback_speed_var.get() or "1.0")
        return self.play_base_time + (time.perf_counter() - self.play_start_perf) * speed

    def next_event_index_for_time(self, current_time: float) -> int:
        if not self.preview_data:
            return 0
        times = self.preview_data.get("event_times", [])
        if not isinstance(times, list):
            return 0
        index = 0
        while index < len(times) and float(times[index]) < current_time:
            index += 1
        return index

    def playback_tick(self) -> None:
        if not self.playing or not self.preview_data:
            return
        clip_length = float(self.preview_data.get("clip_length") or 0.0)
        current_time = self.current_preview_time()
        if clip_length > 0.0 and current_time > clip_length:
            if self.loop_preview_var.get():
                self.play_base_time = 0.0
                self.play_start_perf = time.perf_counter()
                self.next_event_index = 0
                current_time = 0.0
            else:
                self.stop_preview_playback()
                return

        times = self.preview_data.get("event_times", [])
        if isinstance(times, list):
            while self.next_event_index < len(times) and float(times[self.next_event_index]) <= current_time:
                self.play_marker_audio(self.next_event_index)
                self.next_event_index += 1

        self.draw_preview(current_time)
        self.after(33, self.playback_tick)

    def play_marker_audio(self, marker_index: int) -> None:
        mode = self.audio_mode_var.get()
        if mode == "Mute":
            return
        if mode == "Wwise Source" and self.preview_data:
            audio_files = self.preview_data.get("audio_files", [])
            if isinstance(audio_files, list) and audio_files:
                path = Path(random.choice(audio_files))
                try:
                    import winsound

                    winsound.PlaySound(str(path), winsound.SND_FILENAME | winsound.SND_ASYNC)
                    return
                except Exception:
                    pass
        self.play_click(marker_index)

    def play_click(self, marker_index: int) -> None:
        def worker() -> None:
            try:
                import winsound

                freq = 950 + (marker_index % 3) * 120
                winsound.Beep(freq, 45)
            except Exception:
                self.bell()

        threading.Thread(target=worker, daemon=True).start()

    def stop_async_audio(self) -> None:
        try:
            import winsound

            winsound.PlaySound(None, winsound.SND_PURGE)
        except Exception:
            pass

    def draw_preview(self, current_time: float) -> None:
        canvas = getattr(self, "preview_canvas", None)
        if not canvas:
            return
        if self.preview_visual_mode == "timeline_frames":
            self.draw_timeline_preview_frame()
            return
        canvas.delete("all")
        width = max(canvas.winfo_width(), 600)
        height = max(canvas.winfo_height(), 220)
        canvas.create_rectangle(0, 0, width, height, fill="#10151b", outline="")

        if self.preview_visual_mode == "timeline_pending":
            canvas.create_text(
                width / 2,
                height / 2,
                text="Unity is rendering Timeline preview frames...",
                fill=MUTED,
                font=("Segoe UI", 11, "bold"),
            )
            return

        if self.preview_visual_mode == "timeline_error":
            self.draw_timeline_preview_error_to_canvas(canvas)
            return

        if not self.preview_data:
            canvas.create_text(
                width / 2,
                height / 2,
                text="点击“直接预览”后，这里会显示运动曲线、打点和播放光标。",
                fill=MUTED,
                font=("Segoe UI", 11, "bold"),
            )
            return

        clip_length = float(self.preview_data.get("clip_length") or 0.0)
        samples = self.preview_data.get("samples", [])
        event_times = self.preview_data.get("event_times", [])
        audio_files = self.preview_data.get("audio_files", [])
        if clip_length <= 0.0:
            clip_length = max([float(value) for value in event_times], default=1.0) if isinstance(event_times, list) else 1.0

        graph_left = 36
        graph_right = width - 18
        graph_top = 88
        graph_bottom = height - 34
        graph_width = graph_right - graph_left
        graph_height = graph_bottom - graph_top

        current_time = max(0.0, min(current_time, clip_length))
        current_y_norm = self.sample_normalized_y(samples, current_time)
        self.draw_bird_preview(canvas, width, current_y_norm)

        canvas.create_rectangle(graph_left, graph_top, graph_right, graph_bottom, outline="#2b3542", fill="#0d1218")
        canvas.create_text(
            graph_left,
            20,
            text=f"t={current_time:.2f}s / {clip_length:.2f}s",
            anchor="w",
            fill=INK,
            font=("Consolas", 10, "bold"),
        )
        audio_count = len(audio_files) if isinstance(audio_files, list) else 0
        canvas.create_text(
            graph_right,
            20,
            text=f"markers: {len(event_times) if isinstance(event_times, list) else 0}   wav: {audio_count}",
            anchor="e",
            fill=MUTED,
            font=("Consolas", 9),
        )

        if isinstance(samples, list) and samples:
            values = [float(item[1]) for item in samples]
            y_min = min(values)
            y_max = max(values)
            if abs(y_max - y_min) < 0.000001:
                y_min -= 1.0
                y_max += 1.0
            points: list[float] = []
            for item in samples:
                t = float(item[0])
                y = float(item[1])
                x = graph_left + (t / clip_length) * graph_width
                py = graph_bottom - ((y - y_min) / (y_max - y_min)) * graph_height
                points.extend([x, py])
            if len(points) >= 4:
                canvas.create_line(*points, fill=ACCENT, width=2, smooth=True)
        else:
            canvas.create_text(
                (graph_left + graph_right) / 2,
                (graph_top + graph_bottom) / 2,
                text="没有可绘制的端点曲线；仍可用时间轴检查打点。",
                fill=MUTED,
                font=("Segoe UI", 10),
            )

        if isinstance(event_times, list):
            for index, time_value in enumerate(event_times):
                x = graph_left + (float(time_value) / clip_length) * graph_width
                color = WARN if abs(float(time_value) - current_time) < 0.06 else "#6d4d28"
                canvas.create_line(x, graph_top, x, graph_bottom, fill=color, width=1)
                if abs(float(time_value) - current_time) < 0.06:
                    canvas.create_oval(x - 5, graph_top - 7, x + 5, graph_top + 3, fill=WARN, outline="")

        cursor_x = graph_left + (current_time / clip_length) * graph_width
        canvas.create_line(cursor_x, graph_top - 10, cursor_x, graph_bottom + 8, fill=BAD, width=2)
        canvas.create_text(graph_left, graph_bottom + 18, text="0", anchor="w", fill=MUTED, font=("Consolas", 8))
        canvas.create_text(
            graph_right,
            graph_bottom + 18,
            text=f"{clip_length:.1f}s",
            anchor="e",
            fill=MUTED,
            font=("Consolas", 8),
        )

    def ensure_timeline_preview_window(self, pending: bool = False) -> None:
        if self.timeline_preview_window is not None and self.timeline_preview_window.winfo_exists():
            self.timeline_preview_window.deiconify()
            self.timeline_preview_window.lift()
            self.timeline_preview_window.focus_force()
            if pending and self.timeline_preview_popup_canvas is not None:
                self.draw_timeline_preview_pending(self.timeline_preview_popup_canvas)
            return

        window = tk.Toplevel(self)
        window.title("Timeline Preview")
        window.geometry("980x620")
        window.configure(bg=BG)
        window.transient(self)
        window.protocol("WM_DELETE_WINDOW", self.close_timeline_preview_window)
        window.bind("<Escape>", lambda _event: self.close_timeline_preview_window())

        toolbar = tk.Frame(window, bg=PANEL_2)
        toolbar.pack(fill=tk.X)
        self.command_button(toolbar, "Play/Pause", self.toggle_playback, CARD, INK).pack(side=tk.LEFT, padx=6, pady=6)
        self.command_button(toolbar, "Restart", self.stop_preview, CARD, INK).pack(side=tk.LEFT, padx=6, pady=6)
        self.command_button(toolbar, "Reload Frames", lambda: self.load_latest_timeline_preview_frames(), CARD, INK).pack(
            side=tk.LEFT, padx=6, pady=6
        )
        self.command_button(toolbar, "Close", self.close_timeline_preview_window, BAD, INK).pack(
            side=tk.RIGHT, padx=6, pady=6
        )

        canvas = tk.Canvas(window, bg="#05070a", highlightthickness=0)
        canvas.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)
        canvas.bind("<Configure>", lambda _event: self.draw_timeline_preview_frame())
        self.timeline_preview_window = window
        self.timeline_preview_popup_canvas = canvas
        if pending:
            self.draw_timeline_preview_pending(canvas)

    def close_timeline_preview_window(self) -> None:
        self.timeline_preview_playing = False
        self.timeline_preview_popup_user_closed = True
        window = self.timeline_preview_window
        try:
            if window is not None and window.winfo_exists():
                window.withdraw()
                window.destroy()
        except tk.TclError:
            pass
        self.timeline_preview_window = None
        self.timeline_preview_popup_canvas = None
        self.timeline_preview_popup_photo = None

    def draw_timeline_preview_pending(self, canvas: tk.Canvas) -> None:
        canvas.delete("all")
        width = max(canvas.winfo_width(), 600)
        height = max(canvas.winfo_height(), 300)
        canvas.create_rectangle(0, 0, width, height, fill="#10151b", outline="")
        canvas.create_text(
            width / 2,
            height / 2,
            text="Unity is rendering Timeline preview frames...",
            fill=INK,
            font=("Segoe UI", 13, "bold"),
        )
        canvas.create_text(
            width / 2,
            height / 2 + 30,
            text="If this stays here, Unity has not recompiled or loaded the latest editor preview script.",
            fill=MUTED,
            font=("Segoe UI", 9),
        )

    def draw_timeline_preview_error_to_canvas(self, canvas: tk.Canvas) -> None:
        canvas.delete("all")
        width = max(canvas.winfo_width(), 600)
        height = max(canvas.winfo_height(), 260)
        canvas.create_rectangle(0, 0, width, height, fill="#10151b", outline="")
        message = self.timeline_preview_error or "Timeline preview failed."
        canvas.create_text(
            width / 2,
            height / 2 - 12,
            text=message,
            fill=BAD,
            font=("Segoe UI", 11, "bold"),
            width=max(width - 90, 300),
            justify=tk.CENTER,
        )
        canvas.create_text(
            width / 2,
            height / 2 + 46,
            text="This is a tool/editor preview problem, not a change to game audio runtime.",
            fill=MUTED,
            font=("Segoe UI", 9),
            width=max(width - 90, 300),
            justify=tk.CENTER,
        )

    def load_latest_timeline_preview_frames(self) -> None:
        unity_root = Path(self.unity_root_var.get()).resolve()
        manifest_path = unity_root / "Temp" / "ProjectEF_TimelinePrefabPreviewFrames" / "manifest.json"
        try:
            if not manifest_path.exists():
                messagebox.showinfo("Timeline Preview", "No Timeline preview manifest found yet.")
                return
            manifest = json.loads(manifest_path.read_text(encoding="utf-8-sig", errors="replace"))
            self.load_timeline_preview_manifest(manifest)
        except Exception as exc:
            messagebox.showerror("Timeline Preview", str(exc))

    def score_timeline_preview_frame(self, frame_path: Path) -> int:
        try:
            from PIL import Image

            image = Image.open(frame_path).convert("L")
            mask = image.point(lambda pixel: 255 if pixel > 5 else 0)
            bbox = mask.getbbox()
            if not bbox:
                return 0
            return max(0, bbox[2] - bbox[0]) * max(0, bbox[3] - bbox[1])
        except Exception:
            return 1

    def choose_timeline_preview_display_indices(self, frames: list[Path]) -> list[int]:
        if not frames:
            return []
        scores = [self.score_timeline_preview_frame(frame) for frame in frames]
        max_score = max(scores) if scores else 0
        if max_score <= 0:
            return list(range(len(frames)))
        threshold = max(500, int(max_score * 0.15))
        indices = [index for index, score in enumerate(scores) if score >= threshold]
        return indices or [scores.index(max_score)]

    def current_timeline_preview_frame_number(self) -> int:
        if self.timeline_preview_display_indices:
            display_index = self.timeline_preview_frame_index % len(self.timeline_preview_display_indices)
            return self.timeline_preview_display_indices[display_index]
        return self.timeline_preview_frame_index % max(len(self.timeline_preview_frames), 1)

    def draw_timeline_preview_frame_to_canvas(self, canvas: tk.Canvas, photo_attr: str) -> None:
        if not canvas:
            return
        canvas.delete("all")
        width = max(canvas.winfo_width(), 600)
        height = max(canvas.winfo_height(), 220)
        canvas.create_rectangle(0, 0, width, height, fill="#10151b", outline="")

        if not self.timeline_preview_frames:
            canvas.create_text(
                width / 2,
                height / 2,
                text="No Timeline preview frames loaded.",
                fill=MUTED,
                font=("Segoe UI", 11, "bold"),
            )
            return

        frame_number = self.current_timeline_preview_frame_number()
        frame_path = self.timeline_preview_frames[frame_number]
        try:
            from PIL import Image, ImageTk

            image = Image.open(frame_path).convert("RGB")
            gray = image.convert("L")
            mask = gray.point(lambda pixel: 255 if pixel > 5 else 0)
            bbox = mask.getbbox()
            if bbox:
                pad_x = max(12, int((bbox[2] - bbox[0]) * 0.28))
                pad_y = max(12, int((bbox[3] - bbox[1]) * 0.28))
                crop_box = (
                    max(0, bbox[0] - pad_x),
                    max(0, bbox[1] - pad_y),
                    min(image.width, bbox[2] + pad_x),
                    min(image.height, bbox[3] + pad_y),
                )
                crop_w = crop_box[2] - crop_box[0]
                crop_h = crop_box[3] - crop_box[1]
                if crop_w > 4 and crop_h > 4 and crop_w * crop_h < image.width * image.height * 0.85:
                    image = image.crop(crop_box)
            scale = min(width / max(image.width, 1), height / max(image.height, 1))
            target_size = (max(1, int(image.width * scale)), max(1, int(image.height * scale)))
            image = image.resize(target_size, Image.Resampling.LANCZOS)
            setattr(self, photo_attr, ImageTk.PhotoImage(image))
        except Exception:
            setattr(self, photo_attr, tk.PhotoImage(file=str(frame_path)))

        canvas.create_image(width / 2, height / 2, image=getattr(self, photo_attr), anchor="center")
        manifest = self.timeline_preview_manifest or {}
        label = str(manifest.get("prefabAssetPath") or frame_path.name)
        if len(label) > 90:
            label = "..." + label[-87:]
        message = str(manifest.get("message") or "")
        if len(message) > 120:
            message = message[:117] + "..."
        header_height = 42 if message else 24
        canvas.create_rectangle(0, 0, width, header_height, fill="#10151b", outline="")
        canvas.create_text(
            10,
            12,
            text=f"Timeline {frame_number + 1}/{len(self.timeline_preview_frames)}  {label}",
            anchor="w",
            fill=INK,
            font=("Consolas", 9, "bold"),
        )
        if message:
            canvas.create_text(
                10,
                30,
                text=message,
                anchor="w",
                fill=ACCENT if "fallback" in message.lower() or "black" in message.lower() else MUTED,
                font=("Segoe UI", 8, "bold"),
            )
        max_frame_score = int(manifest.get("_maxFrameScore") or 0)
        if max_frame_score <= 0:
            canvas.create_rectangle(0, height / 2 - 45, width, height / 2 + 45, fill="#10151b", outline=BAD)
            canvas.create_text(
                width / 2,
                height / 2 - 10,
                text="Timeline preview frames are black.",
                fill=BAD,
                font=("Segoe UI", 13, "bold"),
            )
            canvas.create_text(
                width / 2,
                height / 2 + 18,
                text="Refresh Unity scripts, then render again. The latest renderer auto-falls back to a tool camera.",
                fill=INK,
                font=("Segoe UI", 9),
            )

    def draw_timeline_preview_frame(self) -> None:
        if self.preview_visual_mode == "timeline_error":
            main_canvas = getattr(self, "preview_canvas", None)
            if main_canvas:
                self.draw_timeline_preview_error_to_canvas(main_canvas)
            popup_canvas = self.timeline_preview_popup_canvas
            if popup_canvas is not None and popup_canvas.winfo_exists():
                self.draw_timeline_preview_error_to_canvas(popup_canvas)
            return
        main_canvas = getattr(self, "preview_canvas", None)
        if main_canvas:
            self.draw_timeline_preview_frame_to_canvas(main_canvas, "timeline_preview_photo")
        popup_canvas = self.timeline_preview_popup_canvas
        if popup_canvas is not None and popup_canvas.winfo_exists():
            self.draw_timeline_preview_frame_to_canvas(popup_canvas, "timeline_preview_popup_photo")

    def sample_normalized_y(self, samples: object, current_time: float) -> float:
        if not isinstance(samples, list) or not samples:
            return 0.5
        values = [float(item[1]) for item in samples]
        y_min = min(values)
        y_max = max(values)
        if abs(y_max - y_min) < 0.000001:
            return 0.5
        previous = samples[0]
        for item in samples[1:]:
            if float(item[0]) >= current_time:
                span = float(item[0]) - float(previous[0])
                ratio = 0.0 if span <= 0.0 else (current_time - float(previous[0])) / span
                y = float(previous[1]) + (float(item[1]) - float(previous[1])) * ratio
                return (y - y_min) / (y_max - y_min)
            previous = item
        return (float(samples[-1][1]) - y_min) / (y_max - y_min)

    def draw_bird_preview(self, canvas: tk.Canvas, width: int, y_norm: float) -> None:
        cx = width / 2
        cy = 48
        body_w = 58
        body_h = 24
        wing_len = 92
        wing_drop = (y_norm - 0.5) * 72
        wing_lift = math.sin((y_norm - 0.5) * math.pi) * 10
        canvas.create_oval(cx - body_w / 2, cy - body_h / 2, cx + body_w / 2, cy + body_h / 2, fill="#d7dde4", outline="")
        canvas.create_oval(cx + 20, cy - 12, cx + 42, cy + 4, fill="#d7dde4", outline="")
        canvas.create_polygon(cx + 40, cy - 5, cx + 58, cy - 1, cx + 40, cy + 4, fill="#f0c36a", outline="")
        canvas.create_line(cx - 10, cy, cx - wing_len, cy + wing_drop - wing_lift, fill=ACCENT, width=5, capstyle=tk.ROUND)
        canvas.create_line(cx + 4, cy, cx + wing_len, cy + wing_drop - wing_lift, fill=ACCENT, width=5, capstyle=tk.ROUND)

    def apply_error(self, payload: dict[str, object]) -> None:
        error = payload.get("error")
        trace = str(payload.get("trace") or "")
        self.append_log(f"ERROR: {error}")
        if trace:
            self.append_log(trace)
        self.status_var.set("执行失败")
        messagebox.showerror("执行失败", str(error))

    def set_running(self, running: bool, status: str) -> None:
        self.status_var.set(status)
        state = tk.DISABLED if running else tk.NORMAL
        for button in self.command_buttons:
            button.configure(state=state)

    def append_log(self, text: str) -> None:
        self.log.configure(state=tk.NORMAL)
        self.log.insert(tk.END, text.rstrip() + "\n")
        self.log.see(tk.END)
        self.log.configure(state=tk.DISABLED)

    def clear_log(self) -> None:
        self.log.configure(state=tk.NORMAL)
        self.log.delete("1.0", tk.END)
        self.log.configure(state=tk.DISABLED)

    def set_text(self, widget: tk.Text, text: str) -> None:
        widget.configure(state=tk.NORMAL)
        widget.delete("1.0", tk.END)
        widget.insert(tk.END, text)
        widget.configure(state=tk.DISABLED)

    def use_resolved_animation(self) -> None:
        value = self.resolved_animation_var.get().strip()
        if value and value not in {"未定位"} and Path(value).exists():
            self.animation_var.set(value)
            self.append_log(f"Animation 输入已切换到可编辑 .anim: {value}")
        else:
            messagebox.showinfo("未定位", "请先点击“定位资源”或“直接预览”。")

    def copy_cli_command(self) -> None:
        try:
            argv = self.build_tool_argv(apply=False)
            command = ["python", "-B", str(CORE_SCRIPT), *argv]
            text = subprocess.list2cmdline(command)
            self.clipboard_clear()
            self.clipboard_append(text)
            self.append_log("已复制预览命令。")
        except Exception as exc:
            messagebox.showerror("无法复制命令", str(exc))

    def write_unity_preview_request(self, report: dict[str, object], command: str = "preview") -> Path | None:
        try:
            unity_root = Path(self.unity_root_var.get()).resolve()
            animation = Path(str(report.get("animation") or "")).resolve()
            prefab_value = report.get("prefab")
            prefab = Path(str(prefab_value)).resolve() if prefab_value else None
            analysis = report.get("analysis") if isinstance(report.get("analysis"), dict) else {}
            event_times = analysis.get("event_times", []) if isinstance(analysis, dict) else []
            if not isinstance(event_times, list) or not event_times:
                self.append_log("Unity preview request blocked: no event times. Run Direct Preview or Write Config first.")
                messagebox.showerror(
                    "No event times",
                    "No analyzed event times are available. Run Direct Preview or Write Config first, then open Unity Preview/Edit Keys.",
                )
                return None

            request = {
                "animationAssetPath": self.to_unity_asset_path(unity_root, animation),
                "prefabAssetPath": self.to_unity_asset_path(unity_root, prefab) if prefab else "",
                "wwiseEvent": str(report.get("wwise_event") or ""),
                "reportPath": str(self.last_report_md or self.last_report_json or ""),
                "eventTimes": [float(value) for value in event_times] if isinstance(event_times, list) else [],
                "command": command,
            }

            request_path = unity_root / "Temp" / "ProjectEF_AnimationWwiseEventPreviewRequest.json"
            request_path.parent.mkdir(parents=True, exist_ok=True)
            request_path.write_text(json.dumps(request, ensure_ascii=True, indent=2), encoding="utf-8")
            self.unity_preview_request_path = request_path
            self.append_log(f"Unity preview request: {request_path}")
            return request_path
        except Exception as exc:
            self.append_log(f"Unity preview request failed: {exc}")
            return None

    def to_unity_asset_path(self, unity_root: Path, path: Path | None) -> str:
        if path is None:
            return ""
        relative = path.resolve().relative_to(unity_root.resolve())
        return relative.as_posix()

    def text_to_unity_asset_path(self, unity_root: Path, value: object) -> str:
        text = str(value or "").strip().replace("\\", "/")
        if not text:
            return ""
        assets_index = text.lower().find("assets/")
        if assets_index >= 0:
            return text[assets_index:]
        return self.to_unity_asset_path(unity_root, Path(text))

    def open_unity_preview(self) -> None:
        if self.last_report:
            report = self.last_report
        else:
            try:
                resolved = self.resolve_current()
            except Exception as exc:
                messagebox.showerror("Unity Preview failed", str(exc))
                return
            animation = resolved.get("animation")
            if not isinstance(animation, Path) or not animation.exists():
                messagebox.showerror("Missing .anim", str(animation))
                return
            messagebox.showinfo(
                "No preview data",
                "Run Direct Preview or Write Config first. Unity Preview will not send an empty event request.",
            )
            return

        request_path = self.write_unity_preview_request(report, command="preview")
        if not request_path:
            return

        if self.is_unity_project_open(Path(self.unity_root_var.get()).resolve()):
            self.append_log("Unity project is already open. Sent preview request to the Unity editor window.")
            self.status_var.set("Unity preview request sent")
            return

        self.launch_unity_execute_method("AnimationWwiseEventPreviewWindow.OpenFromExternalRequest")

    def read_unity_meta_guid(self, unity_asset_path: str) -> str:
        try:
            unity_root = Path(self.unity_root_var.get()).resolve()
            asset_path = self.text_to_unity_asset_path(unity_root, unity_asset_path)
            if not asset_path:
                return ""
            meta_path = unity_root / (asset_path + ".meta")
            if not meta_path.exists():
                return ""
            for line in meta_path.read_text(encoding="utf-8", errors="ignore").splitlines():
                if line.startswith("guid:"):
                    return line.split(":", 1)[1].strip()
        except Exception as exc:
            self.append_log(f"Could not read Unity meta guid for {unity_asset_path}: {exc}")
        return ""

    def timeline_preview_match_score(self, playable_path: str, prefab_path: str, playable_guid: str) -> int:
        unity_root = Path(self.unity_root_var.get()).resolve()
        playable_norm = playable_path.replace("\\", "/")
        prefab_norm = prefab_path.replace("\\", "/")
        playable_lower = playable_norm.lower()
        prefab_lower = prefab_norm.lower()
        score = 0

        if playable_guid:
            try:
                prefab_asset_path = self.text_to_unity_asset_path(unity_root, prefab_path)
                prefab_file = unity_root / prefab_asset_path
                if prefab_file.exists() and playable_guid in prefab_file.read_text(encoding="utf-8", errors="ignore"):
                    score += 200
            except Exception as exc:
                self.append_log(f"Could not inspect prefab Timeline references: {prefab_path}: {exc}")

        playable_parts = playable_lower.split("/")
        prefab_parts = prefab_lower.split("/")
        try:
            marker_index = playable_parts.index("catchingfish")
            family = playable_parts[marker_index + 1]
        except (ValueError, IndexError):
            family = ""

        if family:
            if family in prefab_parts:
                score += 120
            if family in Path(prefab_lower).stem:
                score += 35

        playable_name = Path(playable_lower).stem
        prefab_compact = re.sub(r"[^a-z0-9]+", "", prefab_lower)
        timeline_match = re.search(r"catchingfish[_-]?([a-z]+)[_-]?(\d+)", playable_name)
        if timeline_match:
            exact_token = timeline_match.group(1) + timeline_match.group(2)
            if exact_token in prefab_compact:
                score += 90

        if "pfb_cutscene_catchingfish" in Path(prefab_lower).stem:
            score += 20
        if "pfb_tool" in Path(prefab_lower).stem:
            score -= 100
        if "_new" in Path(prefab_lower).stem:
            score -= 15
        if "variant" in Path(prefab_lower).stem:
            score -= 10

        return score

    def resolve_prefabs_for_selected_timeline(self, playable_path: str, candidate_prefabs: list[str]) -> list[str]:
        playable_guid = self.read_unity_meta_guid(playable_path)
        scored: list[tuple[int, str]] = []
        for prefab_path in candidate_prefabs:
            score = self.timeline_preview_match_score(playable_path, prefab_path, playable_guid)
            if score > 0:
                scored.append((score, prefab_path))

        scored.sort(key=lambda item: (-item[0], item[1].lower()))
        if scored:
            top_score = scored[0][0]
            if top_score >= 100:
                threshold = max(100, int(top_score * 0.6))
                filtered = [(score, prefab_path) for score, prefab_path in scored if score >= threshold]
            else:
                filtered = scored[:1]
            self.append_log(
                "Timeline playable matched to prefab: "
                f"{scored[0][1]} (score {scored[0][0]}, candidates {len(scored)}, using {len(filtered)})"
            )
            return [prefab_path for _, prefab_path in filtered]
        return []

    def collect_timeline_preview_targets(self) -> tuple[list[str], str]:
        selected = self.selected_task_candidate()
        selected_path = ""
        selected_suffix = ""
        selected_prefab_path = ""
        selected_timeline_path = ""
        if selected is not None:
            selected_path = self.candidate_value(selected, "unity_path") or self.candidate_value(selected, "path")
            selected_suffix = Path(selected_path).suffix.lower()
            selected_prefab_path = self.candidate_value(selected, "prefab_path")
            selected_timeline_path = self.candidate_value(selected, "timeline_path")

        prefab_paths: list[str] = []
        timeline_path = ""
        candidate_prefabs: list[str] = []
        other_prefabs: list[str] = []
        for candidate in self.task_candidates:
            path_value = self.candidate_value(candidate, "unity_path") or self.candidate_value(candidate, "path")
            if Path(path_value).suffix.lower() != ".prefab":
                continue
            kind = self.candidate_value(candidate, "kind")
            if kind == "Timeline Prefab":
                candidate_prefabs.append(path_value)
            else:
                other_prefabs.append(path_value)

        if selected_timeline_path and selected_prefab_path:
            prefab_paths.append(selected_prefab_path)
            timeline_path = selected_timeline_path
        elif selected_path and selected_suffix == ".prefab":
            prefab_paths.append(selected_path)
            if selected_timeline_path:
                timeline_path = selected_timeline_path
        elif selected_path and selected_suffix == ".playable":
            if selected_prefab_path:
                prefab_paths.append(selected_prefab_path)
            else:
                prefab_paths.extend(self.resolve_prefabs_for_selected_timeline(selected_path, candidate_prefabs))
            if prefab_paths:
                timeline_path = selected_path
            else:
                self.append_log(
                    "Timeline asset selected, but no matching Timeline Prefab was found. "
                    "Falling back to the first Timeline Prefab candidate."
                )

        if not prefab_paths:
            prefab_paths.extend(candidate_prefabs)

        if not prefab_paths:
            prefab_text = self.prefab_var.get().strip()
            if prefab_text:
                prefab_paths.append(prefab_text)

        if not prefab_paths:
            prefab_paths.extend(other_prefabs)

        unique_prefabs: list[str] = []
        seen: set[str] = set()
        for value in prefab_paths:
            key = value.replace("\\", "/").lower()
            if value and key not in seen:
                seen.add(key)
                unique_prefabs.append(value)

        return unique_prefabs, timeline_path

    def write_timeline_preview_request(
        self,
        prefab_paths: list[str],
        timeline_path: str = "",
        command: str = "open_native",
    ) -> Path | None:
        try:
            unity_root = Path(self.unity_root_var.get()).resolve()
            prefab_asset_paths = [
                self.text_to_unity_asset_path(unity_root, value)
                for value in prefab_paths
                if str(value or "").strip()
            ]
            prefab_asset_paths = [value for value in prefab_asset_paths if value]
            if not prefab_asset_paths:
                messagebox.showinfo(
                    "Timeline Preview",
                    "Select a prefab candidate, or fill the Prefab field first.",
                )
                return None

            timeline_asset_path = self.text_to_unity_asset_path(unity_root, timeline_path) if timeline_path else ""
            output_dir = unity_root / "Temp" / "ProjectEF_TimelinePrefabPreviewFrames"
            request = {
                "prefabAssetPath": prefab_asset_paths[0],
                "prefabAssetPaths": prefab_asset_paths,
                "timelineAssetPath": timeline_asset_path,
                "reportPath": str(self.last_task_card_md or self.last_report_md or self.last_task_card_json or ""),
                "command": command,
                "outputDirectory": str(output_dir),
                "frameCount": 48,
                "width": 640,
                "height": 360,
                "prepareProjectTimelineRuntime": False,
                "requestNonce": int(time.time() * 1000),
            }

            request_path = unity_root / "Temp" / "ProjectEF_TimelinePrefabPreviewRequest.json"
            request_path.parent.mkdir(parents=True, exist_ok=True)
            if command == "render_frames":
                output_dir.mkdir(parents=True, exist_ok=True)
                manifest_path = output_dir / "manifest.json"
                try:
                    if manifest_path.exists():
                        manifest_path.unlink()
                except OSError as exc:
                    self.append_log(f"Could not remove stale Timeline manifest: {exc}")
            request_path.write_text(json.dumps(request, ensure_ascii=True, indent=2), encoding="utf-8")
            self.append_log(f"Timeline preview request: {request_path}")
            self.append_log(f"Timeline preview command: {command}")
            self.append_log(f"Timeline preview prefabs: {len(prefab_asset_paths)}")
            self.append_log(f"Timeline preview target: {prefab_asset_paths[0]}")
            if timeline_asset_path:
                self.append_log(f"Timeline preview timeline: {timeline_asset_path}")
            self.save_current_round_index("timeline_preview_request")
            return request_path
        except Exception as exc:
            self.append_log(f"Timeline preview request failed: {exc}")
            messagebox.showerror("Timeline Preview failed", str(exc))
            return None

    def open_timeline_prefab_preview(self) -> None:
        prefab_paths, timeline_path = self.collect_timeline_preview_targets()
        request_path = self.write_timeline_preview_request(prefab_paths, timeline_path, command="open_native")
        if not request_path:
            return

        unity_root = Path(self.unity_root_var.get()).resolve()
        if self.is_unity_project_open(unity_root):
            stale_message = self.timeline_preview_stale_script_message(unity_root)
            if stale_message:
                self.preview_visual_mode = "timeline_error"
                self.timeline_preview_error = stale_message
                self.ensure_timeline_preview_window(pending=False)
                self.draw_preview(0.0)
                self.append_log(stale_message)
                messagebox.showwarning("Timeline Preview needs Unity refresh", stale_message)
                return
            self.append_log("Unity project is already open. Sent native Timeline preview request to the Unity editor.")
            self.status_var.set("Native Timeline preview request sent")
            return

        self.launch_unity_execute_method("TimelinePrefabPreviewWindow.OpenNativeTimelinePreviewFromExternalRequest")

    def open_timeline_frame_debug_preview(self) -> None:
        prefab_paths, timeline_path = self.collect_timeline_preview_targets()
        request_path = self.write_timeline_preview_request(prefab_paths, timeline_path, command="render_frames")
        if not request_path:
            return

        unity_root = Path(self.unity_root_var.get()).resolve()
        if self.is_unity_project_open(unity_root):
            stale_message = self.timeline_preview_stale_script_message(unity_root)
            if stale_message:
                self.preview_visual_mode = "timeline_error"
                self.timeline_preview_error = stale_message
                self.ensure_timeline_preview_window(pending=False)
                self.draw_preview(0.0)
                self.append_log(stale_message)
                messagebox.showwarning("Timeline Frame Debug needs Unity refresh", stale_message)
                return
            self.append_log("Unity project is already open. Sent in-tool Timeline frame debug request to the Unity editor.")
            self.status_var.set("Timeline frame debug request sent")
            self.start_timeline_preview_poll()
            return

        self.launch_unity_execute_method("TimelinePrefabPreviewWindow.RenderFramesFromExternalRequest")
        self.start_timeline_preview_poll()

    def timeline_preview_stale_script_message(self, unity_root: Path) -> str:
        source = unity_root / "Assets" / "GameProject" / "Scripts" / "Editor" / "Animation" / "TimelinePrefabPreviewWindow.cs"
        editor_dll = unity_root / "Library" / "ScriptAssemblies" / "Assembly-CSharp-Editor.dll"
        try:
            if not source.exists() or not editor_dll.exists():
                return ""
            if source.stat().st_mtime > editor_dll.stat().st_mtime + 2:
                source_time = datetime.fromtimestamp(source.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S")
                dll_time = datetime.fromtimestamp(editor_dll.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S")
                return (
                    "Unity has not loaded the latest Timeline preview script yet.\n\n"
                    f"Script source time: {source_time}\n"
                    f"Loaded editor assembly time: {dll_time}\n\n"
                    "Stop Unity Play Mode if it is running, then use Unity menu Assets > Refresh "
                    "and wait for scripts to compile. After that, click Timeline Preview again."
                )
        except OSError:
            return ""
        return ""

    def start_timeline_preview_poll(self) -> None:
        unity_root = Path(self.unity_root_var.get()).resolve()
        manifest_path = unity_root / "Temp" / "ProjectEF_TimelinePrefabPreviewFrames" / "manifest.json"
        self.preview_visual_mode = "timeline_pending"
        self.timeline_preview_frames = []
        self.timeline_preview_display_indices = []
        self.timeline_preview_manifest = None
        self.timeline_preview_frame_index = 0
        self.timeline_preview_poll_attempts = 0
        self.timeline_preview_error = ""
        self.timeline_preview_popup_user_closed = False
        self.ensure_timeline_preview_window(pending=True)
        self.draw_preview(0.0)
        self.after(500, lambda: self.poll_timeline_preview_manifest(manifest_path))

    def poll_timeline_preview_manifest(self, manifest_path: Path) -> None:
        self.timeline_preview_poll_attempts += 1
        try:
            if manifest_path.exists():
                manifest = json.loads(manifest_path.read_text(encoding="utf-8-sig", errors="replace"))
                status = str(manifest.get("status") or "")
                if status == "ready":
                    self.load_timeline_preview_manifest(manifest)
                    return
                if status == "error":
                    self.preview_visual_mode = "timeline_error"
                    message = str(manifest.get("message") or "Unknown Timeline render error.")
                    self.timeline_preview_error = message
                    self.append_log(f"Timeline preview render failed: {message}")
                    self.status_var.set("Timeline preview failed")
                    self.draw_preview(0.0)
                    return
        except Exception as exc:
            self.append_log(f"Timeline preview manifest read failed: {exc}")

        if self.timeline_preview_poll_attempts >= 120:
            self.preview_visual_mode = "timeline_error"
            self.timeline_preview_error = "Unity did not finish rendering Timeline preview frames. Check Unity Console, then use ProjectEF/Audio/Render Timeline Preview Frames."
            self.append_log("Timeline preview render timed out.")
            self.status_var.set("Timeline preview timed out")
            self.draw_preview(0.0)
            return

        self.after(500, lambda: self.poll_timeline_preview_manifest(manifest_path))

    def load_timeline_preview_manifest(self, manifest: dict[str, object]) -> None:
        frames_value = manifest.get("frames")
        frames = [Path(str(value)) for value in frames_value] if isinstance(frames_value, list) else []
        frames = [path for path in frames if path.exists()]
        if not frames:
            self.preview_visual_mode = "timeline_error"
            self.timeline_preview_error = "Unity rendered no preview frames."
            self.append_log("Timeline preview manifest is ready, but no frame files were found.")
            self.status_var.set("Timeline preview failed")
            self.draw_preview(0.0)
            return

        diagnostics = str(manifest.get("diagnostics") or "")
        tool_revision = str(manifest.get("toolRevision") or "")
        if not tool_revision:
            stale_message = self.timeline_preview_stale_script_message(Path(self.unity_root_var.get()).resolve())
            if stale_message:
                self.preview_visual_mode = "timeline_error"
                self.timeline_preview_error = (
                    stale_message
                    + "\n\nThe loaded manifest was produced by the old renderer and is ignored to avoid black preview playback."
                )
                self.timeline_preview_frames = []
                self.timeline_preview_display_indices = []
                self.append_log("Timeline preview blocked: old manifest has no toolRevision and Unity scripts are stale.")
                self.status_var.set("Timeline preview needs Unity refresh")
                if not self.timeline_preview_popup_user_closed:
                    self.ensure_timeline_preview_window(pending=False)
                self.draw_timeline_preview_frame()
                return
            manifest["message"] = (
                "Unity is still running an older Timeline preview renderer; "
                "refresh Unity scripts, then render again."
            )
            self.append_log("Timeline preview warning: manifest has no toolRevision, so Unity has not loaded the latest preview script.")
        first_candidate_match = re.search(r"\[0\]\s+(.*?)\s+score=(\d+)", diagnostics)
        if first_candidate_match:
            first_candidate = first_candidate_match.group(1).strip()
            first_score = int(first_candidate_match.group(2))
            shown_candidate = str(manifest.get("prefabAssetPath") or "").strip()
            current_message = str(manifest.get("message") or "")
            if first_score <= 8 and shown_candidate and first_candidate and shown_candidate != first_candidate:
                manifest["message"] = (
                    "Selected Timeline prefab rendered almost black; "
                    "showing best visible fallback candidate."
                )
                self.append_log(
                    "Timeline preview fallback: selected candidate rendered black, "
                    f"showing {shown_candidate}. Original message: {current_message}"
                )

        frame_scores = [self.score_timeline_preview_frame(frame) for frame in frames]
        manifest["_maxFrameScore"] = max(frame_scores) if frame_scores else 0
        if manifest["_maxFrameScore"] <= 0:
            manifest["message"] = (
                "Timeline frames are black. Refresh Unity scripts and render again; "
                "the updated renderer uses a tool-camera fallback."
            )
            self.append_log("Timeline preview frames are black; Unity needs the latest preview script loaded.")

        self.timeline_preview_manifest = manifest
        self.timeline_preview_frames = frames
        self.timeline_preview_display_indices = self.choose_timeline_preview_display_indices(frames)
        self.timeline_preview_frame_index = 0
        duration = float(manifest.get("duration") or 0.0)
        if duration > 0 and frames:
            self.timeline_preview_interval_ms = max(33, min(250, int(duration * 1000 / len(frames))))
        else:
            self.timeline_preview_interval_ms = 100
        self.preview_visual_mode = "timeline_frames"
        self.timeline_preview_playing = True
        if not self.timeline_preview_popup_user_closed:
            self.ensure_timeline_preview_window(pending=False)
        self.append_log(
            f"Timeline preview loaded in tool: {len(frames)} frames / {duration:.2f}s / "
            f"{manifest.get('prefabAssetPath') or ''}"
        )
        self.status_var.set("Timeline preview loaded")
        self.draw_timeline_preview_frame()
        self.after(self.timeline_preview_interval_ms, self.timeline_preview_tick)

    def timeline_preview_tick(self) -> None:
        if self.preview_visual_mode != "timeline_frames" or not self.timeline_preview_frames:
            return
        if self.timeline_preview_playing:
            frame_count = len(self.timeline_preview_display_indices) or len(self.timeline_preview_frames)
            self.timeline_preview_frame_index = (self.timeline_preview_frame_index + 1) % frame_count
            self.draw_timeline_preview_frame()
        self.after(self.timeline_preview_interval_ms, self.timeline_preview_tick)

    def launch_unity_execute_method(self, method_name: str) -> None:
        unity_root = Path(self.unity_root_var.get()).resolve()
        unity_exe = self.find_unity_exe(unity_root)
        if not unity_exe:
            selected = filedialog.askopenfilename(
                title="Select Unity.exe",
                filetypes=(("Unity Editor", "Unity.exe"), ("Executable", "*.exe"), ("All files", "*.*")),
            )
            unity_exe = Path(selected) if selected else None
        if not unity_exe:
            messagebox.showerror(
                "Unity not found",
                "Could not find Unity.exe. Open the project manually, then use menu ProjectEF/Audio/Animation Wwise Event Preview.",
            )
            return

        project_version = self.read_unity_version(unity_root)
        if project_version:
            unity_text = str(unity_exe)
            if project_version not in unity_text:
                messagebox.showerror(
                    "Unity version mismatch",
                    f"Blocked for safety.\n\n"
                    f"Project version is {project_version}, but the selected Unity is:\n{unity_exe}\n\n"
                    f"Please select the exact Unity version for this project.",
                )
                return

        command = [
            str(unity_exe),
            "-projectPath",
            str(unity_root),
            "-executeMethod",
            method_name,
        ]
        try:
            subprocess.Popen(command, cwd=str(unity_root))
            self.append_log("Launched Unity preview command:")
            self.append_log(subprocess.list2cmdline(command))
        except Exception as exc:
            messagebox.showerror("Unity launch failed", str(exc))

    def is_unity_project_open(self, unity_root: Path) -> bool:
        unity_root_text = self.normalize_for_compare(unity_root)
        try:
            command = [
                "powershell.exe",
                "-NoProfile",
                "-NonInteractive",
                "-Command",
                "Get-CimInstance Win32_Process -Filter \"name = 'Unity.exe'\" | "
                "Select-Object -ExpandProperty CommandLine",
            ]
            creationflags = getattr(subprocess, "CREATE_NO_WINDOW", 0) if os.name == "nt" else 0
            completed = subprocess.run(
                command,
                stdout=subprocess.PIPE,
                stderr=subprocess.DEVNULL,
                text=True,
                encoding="utf-8",
                errors="replace",
                timeout=5,
                creationflags=creationflags,
            )
        except Exception:
            return False

        for line in completed.stdout.splitlines():
            normalized = self.normalize_for_compare(line)
            if unity_root_text in normalized and "assetimportworker" not in normalized.lower():
                return True
        return False

    def normalize_for_compare(self, value: object) -> str:
        return str(value).replace("\\", "/").replace('"', "").lower()

    def find_unity_exe(self, unity_root: Path) -> Path | None:
        version = self.read_unity_version(unity_root)
        candidates: list[Path] = []
        if version:
            version_without_suffix = re.sub(r"c\d+$", "", version)
            candidates.extend(
                [
                    Path(r"D:\Unity") / version / "Editor" / "Unity.exe",
                    Path(r"C:\Unity") / version / "Editor" / "Unity.exe",
                    Path(r"E:\Unity") / version / "Editor" / "Unity.exe",
                    Path(r"C:\Program Files\Unity\Hub\Editor") / version / "Editor" / "Unity.exe",
                    Path(r"D:\Program Files\Unity\Hub\Editor") / version / "Editor" / "Unity.exe",
                    Path(r"E:\Program Files\Unity\Hub\Editor") / version / "Editor" / "Unity.exe",
                    Path(r"C:\Program Files") / f"Unity {version}" / "Editor" / "Unity.exe",
                    Path(r"C:\Program Files") / f"Unity {version_without_suffix}" / "Editor" / "Unity.exe",
                ]
            )

        for base in [
            Path(r"C:\Program Files\Unity\Hub\Editor"),
            Path(r"D:\Program Files\Unity\Hub\Editor"),
            Path(r"E:\Program Files\Unity\Hub\Editor"),
        ]:
            if base.exists():
                candidates.extend(sorted(base.glob("*/Editor/Unity.exe"), reverse=True))

        for base in [Path(r"C:\Program Files"), Path(r"D:\Program Files"), Path(r"E:\Program Files")]:
            if base.exists():
                candidates.extend(sorted(base.glob("Unity*/Editor/Unity.exe"), reverse=True))

        for base in [Path(r"C:\Unity"), Path(r"D:\Unity"), Path(r"E:\Unity")]:
            if base.exists():
                candidates.extend(sorted(base.glob("*/Editor/Unity.exe"), reverse=True))

        which_unity = shutil.which("Unity") or shutil.which("Unity.exe")
        if which_unity:
            candidates.append(Path(which_unity))

        seen: set[str] = set()
        for candidate in candidates:
            key = str(candidate).lower()
            if key in seen:
                continue
            seen.add(key)
            if candidate.exists():
                return candidate
        return None

    def read_unity_version(self, unity_root: Path) -> str | None:
        path = unity_root / "ProjectSettings" / "ProjectVersion.txt"
        if not path.exists():
            return None
        match = re.search(r"m_EditorVersion:\s*(\S+)", path.read_text(encoding="utf-8", errors="replace"))
        return match.group(1) if match else None

    def current_animation_file(self) -> Path:
        if self.last_report and self.last_report.get("animation"):
            return Path(str(self.last_report["animation"]))

        candidate = self.resolved_animation_var.get().strip()
        if candidate and Path(candidate).exists():
            return Path(candidate)

        resolved = self.resolve_current()
        animation = resolved.get("animation")
        if isinstance(animation, Path):
            return animation
        return Path(str(animation))

    def open_project_audio_preview(self) -> None:
        try:
            if self.last_report:
                report = self.last_report
            else:
                resolved = self.resolve_current()
                animation = resolved.get("animation")
                if not isinstance(animation, Path) or not animation.exists():
                    messagebox.showerror("Missing .anim", str(animation))
                    return
                messagebox.showinfo(
                    "No preview data",
                    "Run Direct Preview or Write Config first. Project Audio will not send an empty event request.",
                )
                return

            request_path = self.write_unity_preview_request(report, command="project_audio_preview")
            if not request_path:
                return

            if self.is_unity_project_open(Path(self.unity_root_var.get()).resolve()):
                self.append_log("Unity project is already open. Sent project-audio preview request to the Unity editor window.")
                self.status_var.set("Project audio preview request sent")
                return

            self.launch_unity_execute_method("AnimationWwiseEventPreviewWindow.OpenFromExternalRequest")
        except Exception as exc:
            messagebox.showerror("Project Audio Preview failed", str(exc))

    def open_anim_text_editor(self) -> None:
        try:
            if self.last_report:
                report = self.last_report
            else:
                resolved = self.resolve_current()
                animation = resolved.get("animation")
                if not isinstance(animation, Path) or not animation.exists():
                    messagebox.showerror("Missing .anim", str(animation))
                    return
                messagebox.showinfo(
                    "No preview data",
                    "Run Direct Preview or Write Config first. Unity Edit Keys will not send an empty event request.",
                )
                return

            request_path = self.write_unity_preview_request(report, command="edit_keys")
            if not request_path:
                return

            if self.is_unity_project_open(Path(self.unity_root_var.get()).resolve()):
                self.append_log("Unity project is already open. Sent edit-key request to the Unity editor window.")
                self.status_var.set("Unity edit-key request sent")
                return

            self.launch_unity_execute_method("AnimationWwiseEventPreviewWindow.OpenFromExternalRequest")
        except Exception as exc:
            messagebox.showerror("Open Animation Event keys failed", str(exc))

    def open_unity_root(self) -> None:
        self.open_path(Path(self.unity_root_var.get()))

    def open_wwise_root(self) -> None:
        self.open_path(Path(self.wwise_root_var.get()))

    def open_animation_location(self) -> None:
        candidate = self.resolved_animation_var.get()
        if not candidate or candidate == "未定位":
            candidate = self.animation_var.get()
        self.open_path(Path(candidate), select_file=True)

    def open_prefab_location(self) -> None:
        candidate = self.resolved_prefab_var.get()
        if not candidate or candidate.startswith("未") or candidate == "自动":
            candidate = self.prefab_var.get()
        if candidate:
            self.open_path(Path(candidate), select_file=True)
        else:
            messagebox.showinfo("未指定", "Prefab 当前是自动推断，请先定位资源。")

    def open_wwise_event_location(self) -> None:
        text = self.wwise_evidence_var.get()
        path = Path(text)
        if path.exists():
            self.open_path(path, select_file=True)
        else:
            messagebox.showinfo("未定位", "请先点击“校验 / 定位”。")

    def open_last_report(self) -> None:
        if self.last_report_md and self.last_report_md.exists():
            self.open_path(self.last_report_md, select_file=False)
        elif self.last_report_json and self.last_report_json.exists():
            self.open_path(self.last_report_json, select_file=False)
        else:
            messagebox.showinfo("暂无报告", "请先点击“直接预览”或“写入 / 修改配置”。")

    def open_report_folder(self) -> None:
        self.open_path(core.REPORT_DIR)

    def open_path(self, path: Path, select_file: bool = False) -> None:
        try:
            path = path.expanduser()
            if select_file and path.is_file() and os.name == "nt":
                subprocess.Popen(["explorer", f"/select,{str(path)}"])
            elif path.exists():
                os.startfile(str(path)) if os.name == "nt" else subprocess.Popen(["xdg-open", str(path)])
            else:
                messagebox.showinfo("路径不存在", str(path))
        except Exception as exc:
            messagebox.showerror("无法打开", str(exc))


def main() -> int:
    if not CORE_SCRIPT.exists():
        messagebox.showerror("Missing script", f"Missing core tool:\n{CORE_SCRIPT}")
        return 1
    app = AnimationWwiseEventAutoConfigGui()
    app.mainloop()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
