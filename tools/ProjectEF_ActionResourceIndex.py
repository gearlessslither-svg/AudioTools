#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import re
from dataclasses import dataclass, asdict
from pathlib import Path


APP_DIR = Path(__file__).resolve().parent
ROOT = APP_DIR.parent
DEFAULT_UNITY_ROOT = Path(r"D:\EF New\Client\TargetProject")
REPORT_DIR = ROOT / "Reports" / "ActionResourceIndex"

ACTION_EXTENSIONS = {".playable", ".anim", ".prefab", ".fbx", ".controller", ".overridecontroller"}
TEXT_EXTENSIONS = {".playable", ".anim", ".prefab", ".controller", ".overridecontroller", ".asset"}
SCAN_ROOTS = [
    "Assets/GameProject/RuntimeAssets",
    "Assets/GameProject/ArtAssets",
]
GUID_RE = re.compile(r"\bguid:\s*([0-9a-fA-F]{32})\b")


@dataclass
class ActionResourceRow:
    index: int
    kind: str
    preview_kind: str
    preview_ready: str
    preview_route: str
    name: str
    unity_path: str
    prefab_path: str
    timeline_path: str
    animation_path: str
    fbx_path: str
    owner_count: int
    system: str
    group: str
    score: int
    guid: str
    notes: str


def now_stamp() -> str:
    return dt.datetime.now().strftime("%Y%m%d_%H%M%S")


def to_unity_path(path: Path, unity_root: Path) -> str:
    try:
        return path.resolve().relative_to(unity_root.resolve()).as_posix()
    except Exception:
        return str(path).replace("\\", "/")


def abs_from_unity(unity_root: Path, unity_path: str) -> Path:
    return unity_root / unity_path.replace("/", "\\")


def read_text(path: Path, max_bytes: int = 4_000_000) -> str:
    if path.suffix.lower() not in TEXT_EXTENSIONS:
        return ""
    try:
        data = path.read_bytes()
        if len(data) > max_bytes:
            data = data[:max_bytes]
        return data.decode("utf-8", errors="ignore")
    except Exception:
        return ""


def read_guid(path: Path) -> str:
    meta = Path(str(path) + ".meta")
    if not meta.exists():
        return ""
    try:
        for line in meta.read_text(encoding="utf-8", errors="ignore").splitlines():
            if line.startswith("guid:"):
                return line.split(":", 1)[1].strip()
    except Exception:
        return ""
    return ""


def iter_scan_files(unity_root: Path, scan_roots: list[str]) -> list[Path]:
    files: list[Path] = []
    for scan_root in scan_roots:
        root = abs_from_unity(unity_root, scan_root) if scan_root.startswith("Assets/") else Path(scan_root)
        if not root.exists():
            continue
        for path in root.rglob("*"):
            if not path.is_file():
                continue
            suffix = path.suffix.lower()
            if suffix == ".meta" or suffix not in ACTION_EXTENSIONS:
                continue
            files.append(path)
    files.sort(key=lambda item: to_unity_path(item, unity_root).lower())
    return files


def path_system_and_group(unity_path: str) -> tuple[str, str]:
    parts = unity_path.split("/")
    lowered = [part.lower() for part in parts]
    for marker in ("runtimeassets", "artassets"):
        if marker in lowered:
            idx = lowered.index(marker)
            system = parts[idx + 1] if idx + 1 < len(parts) else ""
            group = parts[idx + 2] if idx + 2 < len(parts) else ""
            return system, group
    if len(parts) >= 3:
        return parts[1], parts[2]
    return "", ""


def compact_tokens(value: str) -> set[str]:
    tokens = set(re.findall(r"[a-z0-9]+", value.lower()))
    return {token for token in tokens if len(token) >= 2}


def score_owner(asset_path: str, owner_path: str, owner_text: str) -> int:
    asset = asset_path.lower()
    owner = owner_path.lower()
    score = 0
    if "/prefabs/" in owner:
        score += 20
    if "playabledirector" in owner_text.lower() or "timelinedirector" in owner_text.lower():
        score += 40
    if "pfb_tool" in owner:
        score -= 80
    asset_tokens = compact_tokens(Path(asset).stem)
    owner_tokens = compact_tokens(Path(owner).stem)
    score += min(60, len(asset_tokens & owner_tokens) * 12)
    asset_parts = asset.split("/")
    owner_parts = owner.split("/")
    score += min(40, len(set(asset_parts) & set(owner_parts)) * 4)
    return score


def choose_best_owner(asset_path: str, owners: list[str], owner_texts: dict[str, str]) -> str:
    if not owners:
        return ""
    scored = [(score_owner(asset_path, owner, owner_texts.get(owner, "")), owner) for owner in owners]
    scored.sort(key=lambda item: (-item[0], item[1].lower()))
    return scored[0][1]


def classify_prefab(unity_path: str, text: str, referenced_timelines: list[str], referenced_animations: list[str]) -> tuple[str, str, int, str]:
    lowered = text.lower()
    path_lower = unity_path.lower()
    if referenced_timelines or "playabledirector:" in lowered or "timelinedirector" in lowered:
        return "Timeline Prefab", "timeline", 75, "Prefab owns or references a Timeline/PlayableDirector."
    if "particlesystem:" in lowered or "visualeffect" in lowered or "pfb_vfx" in path_lower or "/vfx" in path_lower:
        score = 45
        note = "Prefab looks like VFX/particle content."
        if "wwise" in lowered or "akevent" in lowered or "audio" in lowered:
            score += 20
            note += " It also has audio-like text."
        return "VFX Prefab", "prefab", score, note
    if referenced_animations or "animator:" in lowered:
        return "Animated Prefab", "animation", 40, "Prefab owns an Animator or references animation assets."
    return "Prefab", "prefab", 10, "General prefab context."


def row_from_asset(
    index: int,
    kind: str,
    preview_kind: str,
    unity_path: str,
    guid: str,
    prefab_path: str = "",
    timeline_path: str = "",
    animation_path: str = "",
    fbx_path: str = "",
    owner_count: int = 0,
    score: int = 0,
    notes: str = "",
) -> ActionResourceRow:
    system, group = path_system_and_group(unity_path)
    if preview_kind == "timeline" and prefab_path and timeline_path:
        preview_ready = "Yes"
        preview_route = "Timeline Preview"
    elif preview_kind == "animation" and animation_path:
        preview_ready = "Yes"
        preview_route = "Animation Preview"
    elif preview_kind == "prefab":
        preview_ready = "Manual"
        preview_route = "Prefab/VFX owner trace"
    else:
        preview_ready = "No"
        preview_route = "Needs owner"
    return ActionResourceRow(
        index=index,
        kind=kind,
        preview_kind=preview_kind,
        preview_ready=preview_ready,
        preview_route=preview_route,
        name=Path(unity_path).stem,
        unity_path=unity_path,
        prefab_path=prefab_path,
        timeline_path=timeline_path,
        animation_path=animation_path,
        fbx_path=fbx_path,
        owner_count=owner_count,
        system=system,
        group=group,
        score=score,
        guid=guid,
        notes=notes,
    )


def scan_action_resources(unity_root: Path, scan_roots: list[str] | None = None) -> list[ActionResourceRow]:
    scan_roots = scan_roots or SCAN_ROOTS
    files = iter_scan_files(unity_root, scan_roots)
    unity_paths = {path: to_unity_path(path, unity_root) for path in files}

    timelines = [path for path in files if path.suffix.lower() == ".playable"]
    animations = [path for path in files if path.suffix.lower() == ".anim"]
    fbx_sources = [path for path in files if path.suffix.lower() == ".fbx"]
    prefabs = [path for path in files if path.suffix.lower() == ".prefab"]
    controllers = [path for path in files if path.suffix.lower() in {".controller", ".overridecontroller"}]

    guid_to_asset: dict[str, str] = {}
    asset_guid: dict[str, str] = {}
    for path in [*timelines, *animations, *fbx_sources]:
        guid = read_guid(path)
        if not guid:
            continue
        unity_path = unity_paths[path]
        guid_to_asset[guid] = unity_path
        asset_guid[unity_path] = guid

    timeline_owner_prefabs: dict[str, list[str]] = {}
    animation_owner_prefabs: dict[str, list[str]] = {}
    animation_owner_controllers: dict[str, list[str]] = {}
    owner_texts: dict[str, str] = {}
    prefab_rows: list[tuple[str, str, int, str, list[str], list[str]]] = []

    for prefab in prefabs:
        unity_path = unity_paths[prefab]
        text = read_text(prefab)
        owner_texts[unity_path] = text
        refs = [guid_to_asset[guid] for guid in GUID_RE.findall(text) if guid in guid_to_asset]
        timeline_refs = [ref for ref in refs if ref.endswith(".playable")]
        animation_refs = [ref for ref in refs if ref.endswith(".anim") or ref.endswith(".fbx")]
        for ref in timeline_refs:
            timeline_owner_prefabs.setdefault(ref, []).append(unity_path)
        for ref in animation_refs:
            animation_owner_prefabs.setdefault(ref, []).append(unity_path)
        kind, preview_kind, score, note = classify_prefab(unity_path, text, timeline_refs, animation_refs)
        if kind != "Prefab":
            prefab_rows.append((unity_path, kind, score, note, timeline_refs, animation_refs))

    for controller in controllers:
        unity_path = unity_paths[controller]
        text = read_text(controller)
        refs = [guid_to_asset[guid] for guid in GUID_RE.findall(text) if guid in guid_to_asset]
        for ref in refs:
            if ref.endswith(".anim") or ref.endswith(".fbx"):
                animation_owner_controllers.setdefault(ref, []).append(unity_path)

    rows: list[ActionResourceRow] = []

    def next_index() -> int:
        return len(rows) + 1

    for timeline in timelines:
        unity_path = unity_paths[timeline]
        owners = sorted(set(timeline_owner_prefabs.get(unity_path, [])))
        best_prefab = choose_best_owner(unity_path, owners, owner_texts)
        score = 90 if best_prefab else 60
        note = "Timeline asset."
        if best_prefab:
            note += f" Best preview prefab: {best_prefab}."
        else:
            note += " No referencing prefab found in static scan."
        rows.append(
            row_from_asset(
                next_index(),
                "Timeline",
                "timeline",
                unity_path,
                asset_guid.get(unity_path, ""),
                prefab_path=best_prefab,
                timeline_path=unity_path,
                owner_count=len(owners),
                score=score,
                notes=note,
            )
        )

    for animation in animations:
        unity_path = unity_paths[animation]
        prefab_owners = sorted(set(animation_owner_prefabs.get(unity_path, [])))
        controller_owners = sorted(set(animation_owner_controllers.get(unity_path, [])))
        best_prefab = choose_best_owner(unity_path, prefab_owners, owner_texts)
        score = 65 if best_prefab else 45
        note = "AnimationClip asset."
        if best_prefab:
            note += f" Referencing prefab: {best_prefab}."
        if controller_owners:
            note += f" Referenced by {len(controller_owners)} controller(s)."
        rows.append(
            row_from_asset(
                next_index(),
                "AnimationClip",
                "animation",
                unity_path,
                asset_guid.get(unity_path, ""),
                prefab_path=best_prefab,
                animation_path=unity_path,
                owner_count=len(prefab_owners) + len(controller_owners),
                score=score,
                notes=note,
            )
        )

    for prefab_path, kind, score, note, timeline_refs, animation_refs in prefab_rows:
        timeline_path = timeline_refs[0] if timeline_refs else ""
        animation_path = animation_refs[0] if animation_refs else ""
        rows.append(
            row_from_asset(
                next_index(),
                kind,
                "timeline" if timeline_path else ("animation" if animation_path else "prefab"),
                prefab_path,
                "",
                prefab_path=prefab_path,
                timeline_path=timeline_path,
                animation_path=animation_path,
                owner_count=len(timeline_refs) + len(animation_refs),
                score=score,
                notes=note,
            )
        )

    rows.sort(key=lambda row: (row.system.lower(), row.group.lower(), row.kind.lower(), row.name.lower(), row.unity_path.lower()))
    for index, row in enumerate(rows, start=1):
        row.index = index
    return rows


def write_json(rows: list[ActionResourceRow], out_path: Path, unity_root: Path, scan_roots: list[str]) -> Path:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "version": 1,
        "generated_at": dt.datetime.now().isoformat(timespec="seconds"),
        "unity_root": str(unity_root),
        "scan_roots": scan_roots,
        "row_count": len(rows),
        "rows": [asdict(row) for row in rows],
    }
    out_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return out_path


def write_csv(rows: list[ActionResourceRow], out_path: Path) -> Path:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = list(asdict(rows[0]).keys()) if rows else [field.name for field in ActionResourceRow.__dataclass_fields__.values()]
    with out_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(asdict(row))
    return out_path


def write_xlsx(rows: list[ActionResourceRow], out_path: Path) -> Path:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception as exc:
        raise RuntimeError("openpyxl is required to write xlsx") from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "ActionResources"
    fieldnames = list(asdict(rows[0]).keys()) if rows else [field.name for field in ActionResourceRow.__dataclass_fields__.values()]
    sheet.append(fieldnames)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="263445")
    for row in rows:
        sheet.append([asdict(row).get(field, "") for field in fieldnames])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = {
        "kind": 18,
        "preview_kind": 14,
        "preview_ready": 14,
        "preview_route": 22,
        "name": 34,
        "unity_path": 88,
        "prefab_path": 88,
        "timeline_path": 88,
        "animation_path": 88,
        "system": 18,
        "group": 28,
        "notes": 90,
    }
    for index, field in enumerate(fieldnames, start=1):
        width = widths.get(field, 14)
        sheet.column_dimensions[get_column_letter(index)].width = width
    workbook.save(out_path)
    return out_path


def build_action_resource_index(
    unity_root: Path = DEFAULT_UNITY_ROOT,
    scan_roots: list[str] | None = None,
    out_dir: Path = REPORT_DIR,
) -> dict[str, object]:
    scan_roots = scan_roots or SCAN_ROOTS
    rows = scan_action_resources(unity_root, scan_roots)
    stamp = now_stamp()
    stem = f"ProjectEF_ActionResourceIndex_{stamp}"
    out_dir.mkdir(parents=True, exist_ok=True)
    json_path = write_json(rows, out_dir / f"{stem}.json", unity_root, scan_roots)
    csv_path = write_csv(rows, out_dir / f"{stem}.csv")
    xlsx_path = write_xlsx(rows, out_dir / f"{stem}.xlsx")
    return {
        "rows": [asdict(row) for row in rows],
        "json_path": str(json_path),
        "csv_path": str(csv_path),
        "xlsx_path": str(xlsx_path),
        "row_count": len(rows),
        "scan_roots": scan_roots,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Scan ProjectEF action resources and export Excel/CSV/JSON index.")
    parser.add_argument("--unity-root", default=str(DEFAULT_UNITY_ROOT))
    parser.add_argument("--scan-root", action="append", dest="scan_roots")
    parser.add_argument("--out-dir", default=str(REPORT_DIR))
    args = parser.parse_args()

    result = build_action_resource_index(
        unity_root=Path(args.unity_root),
        scan_roots=args.scan_roots or SCAN_ROOTS,
        out_dir=Path(args.out_dir),
    )
    print(json.dumps({key: value for key, value in result.items() if key != "rows"}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
