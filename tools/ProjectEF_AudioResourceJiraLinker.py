#!/usr/bin/env python3
"""
Build a local, read-only mapping between ProjectEF audio-relevant resources and Jira issues.

Inputs:
  - action/audio resource index exported by ProjectEF_ActionResourceIndex.py / audio candidate report
  - Jira issue cache exported by ProjectEF_AudioRequirementJiraTriage_GUI.py
  - optional P4 changelist learning cache and recent P4 metadata

Outputs:
  - Excel report with resource -> Jira and Jira -> resource views
  - JSON report for future GUI integration

This script does not edit Unity, Wwise, Jira, or Perforce state.
"""

from __future__ import annotations

import argparse
import ast
import datetime as _dt
import hashlib
import json
import os
import re
import subprocess
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import Workbook, load_workbook
try:
    from openpyxl.drawing.image import Image as XLImage
except Exception:  # pragma: no cover - optional Excel image support
    XLImage = None
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_TOOLS_DIR = Path(r"G:\AI\Material\Wwise\Tools")
DEFAULT_REPORT_DIR = Path(r"G:\AI\Material\Wwise\Reports\AudioResourceJiraLinks")
DEFAULT_ACTION_INDEX_DIR = Path(r"G:\AI\Material\Wwise\Reports\ActionResourceIndex")
DEFAULT_THUMBNAIL_DIR = Path(r"G:\AI\Material\Wwise\Reports\ResourceThumbnails")
DEFAULT_P4_PORT = "ef.p4.blackjack-local.com:1666"
DEFAULT_P4_USER = "yupeng"
DEFAULT_P4_CLIENT = "yupeng_ADMIN-V9BNJMS5N"
DEFAULT_P4_SCOPE = "//GameProjectEF/ProjectEF_Trunk/Client/TargetProject/..."
DEFAULT_DESIGN_INDEX = DEFAULT_TOOLS_DIR / "audio_requirement_jira_index.json"
DEFAULT_QA_INDEX = DEFAULT_TOOLS_DIR / "audio_requirement_qa_index.json"

JIRA_KEY_RE = re.compile(r"\bPROEF-\d+\b", re.IGNORECASE)
WORD_RE = re.compile(r"[A-Za-z0-9]+|[\u4e00-\u9fff]+")

STOP_TOKENS = {
    "assets",
    "gameproject",
    "runtimeassets",
    "artassets",
    "prefabs",
    "prefab",
    "animation",
    "animations",
    "timeline",
    "material",
    "materials",
    "texture",
    "textures",
    "mesh",
    "meshes",
    "common",
    "template",
    "abs",
    "clp",
    "pfb",
    "tml",
    "vfx",
    "asset",
    "meta",
    "windows",
    "media",
    "event",
    "projectef",
    "project",
    "game",
    "runtime",
    "runtimeassets",
    "client",
    "targetproject",
    "target",
    "type",
    "new",
    "alpha",
    "alpha1",
    "alpha0",
    "version",
    "版本",
    "资源",
    "功能",
    "系统",
    "通用",
}

GENERIC_MATCH_TOKENS = {
    "list",
    "path",
    "content",
    "row",
    "col",
    "data",
    "id",
    "name",
    "desc",
    "string",
    "float",
    "int32",
    "xlsx",
    "\u5b57\u6bb5",
    "\u8def\u5f84",
    "\u5185\u5bb9",
    "\u6570\u636e",
    "\u540d\u79f0",
    "\u63cf\u8ff0",
    "\u529f\u80fd",
    "\u7cfb\u7edf",
    "\u97f3\u9891",
    "\u97f3\u6548",
    "\u7b56\u5212\u6587\u6863",
    "\u529f\u80fd\u6587\u6863",
}

AUDIO_PATH_HINTS = [
    "wwise",
    "wwisebanks",
    "audio",
    "sound",
    "music",
    "voice",
    "akbank",
    "akambient",
    "akunity",
    "soundbank",
    "wwisescriptableobjects",
    "animationwwiseevent",
    "uiaudio",
]

AUDIO_TEXT_HINTS = [
    "audio",
    "sound",
    "wwise",
    "sfx",
    "music",
    "音频",
    "音效",
    "声音",
    "配音",
    "声效",
    "bgm",
]

RESOURCE_ACTION_HINTS = [
    "splash",
    "surface",
    "strike",
    "hit",
    "impact",
    "water",
    "outwater",
    "intowater",
    "unlock",
    "reward",
    "levelup",
    "popup",
    "click",
    "fly",
    "wing",
    "flap",
    "catch",
    "hook",
    "rod",
    "lure",
    "bait",
    "dipnet",
    "walk",
    "swim",
    "loop",
    "open",
    "close",
]


def now_stamp() -> str:
    return _dt.datetime.now().strftime("%Y%m%d_%H%M%S")


def read_json(path: Path) -> Any:
    with path.open("r", encoding="utf-8", errors="replace") as f:
        return json.load(f)


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def latest_file(directory: Path, patterns: Sequence[str]) -> Optional[Path]:
    files: List[Path] = []
    for pattern in patterns:
        files.extend(directory.glob(pattern))
    files = [p for p in files if p.is_file()]
    if not files:
        return None
    return max(files, key=lambda p: p.stat().st_mtime)


def latest_file_recursive(directory: Path, patterns: Sequence[str]) -> Optional[Path]:
    files: List[Path] = []
    for pattern in patterns:
        files.extend(directory.rglob(pattern))
    files = [p for p in files if p.is_file()]
    if not files:
        return None
    return max(files, key=lambda p: p.stat().st_mtime)


def thumbnail_file_name(asset_path: str) -> str:
    normalized = normalize_path(asset_path)
    digest = hashlib.sha1(normalized.lower().encode("utf-8")).hexdigest()[:12]
    stem = Path(normalized).stem or "asset"
    stem = re.sub(r"[^A-Za-z0-9_.-]+", "_", stem)[:48] or "asset"
    return f"{digest}_{stem}.png"


def thumbnail_path_for_asset(thumbnail_dir: Path, asset_path: str) -> Path:
    return thumbnail_dir / thumbnail_file_name(asset_path)


def normalize_path(value: str) -> str:
    return (value or "").replace("\\", "/").strip()


def lower_path(value: str) -> str:
    return normalize_path(value).lower()


def split_name_tokens(text: str) -> List[str]:
    if not text:
        return []
    # Insert breaks for CamelCase while keeping the original text useful for Chinese.
    text = re.sub(r"([a-z0-9])([A-Z])", r"\1 \2", text)
    text = text.replace("\\", "/")
    text = re.sub(r"[_\-/.:()\[\]{}]+", " ", text)
    tokens = []
    for m in WORD_RE.finditer(text):
        token = m.group(0).lower()
        if len(token) <= 1:
            continue
        if token.isdigit():
            continue
        if token in STOP_TOKENS:
            continue
        tokens.append(token)
    return tokens


def compact_tokens(tokens: Iterable[str], limit: int = 80) -> List[str]:
    seen = set()
    out: List[str] = []
    for token in tokens:
        if token in seen:
            continue
        seen.add(token)
        out.append(token)
        if len(out) >= limit:
            break
    return out


def infer_resource_audio_metadata(row: Dict[str, Any]) -> Dict[str, Any]:
    name = str(row.get("name", ""))
    kind = str(row.get("kind", ""))
    system = str(row.get("system", ""))
    group = str(row.get("group", ""))
    path = str(row.get("unity_path", ""))
    hay = " ".join([name, kind, system, group, path]).lower()
    hits = [kw for kw in RESOURCE_ACTION_HINTS + AUDIO_TEXT_HINTS if kw.lower() in hay]
    hits = compact_tokens(hits, limit=12)
    score = 0
    reasons: List[str] = []
    if "Timeline" in kind:
        score += 80
        layer = "Timeline"
        reasons.append("Timeline/Timeline Prefab，常是成段演出或动作集合")
    elif "VFX" in kind or "/fx/" in hay or "vfx" in hay:
        score += 78
        layer = "VFX Owner"
        reasons.append("VFX/FX 资源，常对应瞬态音效，需要找引用者")
    elif system == "UI" or group in {"UIPrefab", "UIAnimation_ABS"} or "/ui/" in hay:
        score += 68
        layer = "UI Audio Config"
        reasons.append("UI 资源，常有点击、弹窗、状态切换声音")
    elif kind == "AnimationClip":
        score += 45
        layer = "AnimationEvent"
        reasons.append("AnimationClip，可按视觉关键帧打点")
    elif "Prefab" in kind:
        score += 40
        layer = "Prefab Component/Owner"
        reasons.append("Prefab 可能是实例化对象或声音挂载点")
    else:
        layer = "Manual Review"
        reasons.append("需要人工确认播放方式")
    if hits:
        score += min(30, len(hits) * 6)
        reasons.append("命中动作/音频关键词: " + ", ".join(hits))
    if score >= 78:
        priority = "High"
    elif score >= 55:
        priority = "Medium"
    elif score >= 30:
        priority = "Low"
    else:
        priority = "Reference"
    return {
        "audio_priority": priority,
        "audio_candidate_score": score,
        "recommended_config_layer": layer,
        "audio_keyword_hits": ", ".join(hits),
        "why_audio_candidate": "；".join(reasons),
    }


def extract_jira_keys(text: str) -> List[str]:
    seen = set()
    keys = []
    for m in JIRA_KEY_RE.finditer(text or ""):
        key = m.group(0).upper()
        if key not in seen:
            seen.add(key)
            keys.append(key)
    return keys


def parse_listish(value: Any) -> List[str]:
    if value is None:
        return []
    if isinstance(value, list):
        out: List[str] = []
        for item in value:
            if isinstance(item, dict):
                key = item.get("key")
                summary = item.get("summary")
                text = " ".join(str(x) for x in [key, summary] if x)
                if text:
                    out.append(text)
            elif item is not None:
                out.append(str(item))
        return out
    text = str(value).strip()
    if not text:
        return []
    if text.startswith("[") and text.endswith("]"):
        try:
            parsed = ast.literal_eval(text)
            return parse_listish(parsed)
        except (SyntaxError, ValueError):
            pass
    return [text]


def text_has_any(text: str, hints: Sequence[str]) -> bool:
    low = (text or "").lower()
    return any(h.lower() in low for h in hints)


def p4_run(args: Sequence[str], timeout: int = 60) -> Tuple[int, str, str]:
    env = os.environ.copy()
    env["P4PORT"] = env.get("P4PORT") or DEFAULT_P4_PORT
    env["P4USER"] = env.get("P4USER") or DEFAULT_P4_USER
    cmd = ["p4", "-c", env.get("P4CLIENT") or DEFAULT_P4_CLIENT, *args]
    proc = subprocess.run(
        cmd,
        cwd=str(DEFAULT_TOOLS_DIR),
        env=env,
        text=True,
        encoding="utf-8",
        errors="replace",
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        timeout=timeout,
    )
    return proc.returncode, proc.stdout, proc.stderr


@dataclass
class ResourceRow:
    name: str
    kind: str
    unity_path: str
    system: str = ""
    group: str = ""
    audio_priority: str = ""
    audio_candidate_score: int = 0
    recommended_config_layer: str = ""
    audio_keyword_hits: str = ""
    why_audio_candidate: str = ""
    preview_ready: str = ""
    tokens: List[str] = field(default_factory=list)

    @property
    def path_lc(self) -> str:
        return lower_path(self.unity_path)


@dataclass
class JiraIssue:
    key: str
    summary: str
    description: str = ""
    status: str = ""
    assignee: str = ""
    reporter: str = ""
    creator: str = ""
    fix_versions: str = ""
    system: str = ""
    design_area: str = ""
    design_doc: str = ""
    dependency_label: str = ""
    audio_required: str = ""
    ready_state: str = ""
    sound_type: str = ""
    url: str = ""
    related_keys: List[str] = field(default_factory=list)
    qa_doc_refs: str = ""
    design_evidence: str = ""
    qa_evidence: str = ""
    doc_tokens: List[str] = field(default_factory=list)
    tokens: List[str] = field(default_factory=list)


@dataclass
class P4Change:
    change: str
    date: str = ""
    owner: str = ""
    summary: str = ""
    files: List[str] = field(default_factory=list)
    jira_keys: List[str] = field(default_factory=list)
    audio_related: bool = False
    reason: str = ""


def load_resources(action_index: Optional[Path], audio_candidate_xlsx: Optional[Path]) -> List[ResourceRow]:
    resources: List[ResourceRow] = []
    if audio_candidate_xlsx and audio_candidate_xlsx.exists():
        wb = load_workbook(audio_candidate_xlsx, read_only=True, data_only=True)
        sheet = "All" if "All" in wb.sheetnames else wb.sheetnames[-1]
        ws = wb[sheet]
        header = [str(c.value or "") for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {name: i for i, name in enumerate(header)}
        for row in ws.iter_rows(min_row=2, values_only=True):
            def get(name: str, default: str = "") -> str:
                i = idx.get(name)
                if i is None or i >= len(row):
                    return default
                value = row[i]
                return "" if value is None else str(value)

            unity_path = get("unity_path")
            if not unity_path:
                continue
            score_text = get("audio_candidate_score", "0")
            try:
                score = int(float(score_text))
            except ValueError:
                score = 0
            resource = ResourceRow(
                name=get("name"),
                kind=get("kind"),
                unity_path=normalize_path(unity_path),
                system=get("system"),
                group=get("group"),
                audio_priority=get("audio_priority"),
                audio_candidate_score=score,
                recommended_config_layer=get("recommended_config_layer"),
                audio_keyword_hits=get("audio_keyword_hits"),
                why_audio_candidate=get("why_audio_candidate"),
                preview_ready=get("preview_ready"),
            )
            resource.tokens = compact_tokens(
                split_name_tokens(" ".join([
                    resource.name,
                    resource.unity_path,
                    resource.system,
                    resource.group,
                    resource.audio_keyword_hits,
                    resource.recommended_config_layer,
                ]))
            )
            resources.append(resource)
        return resources

    if action_index is None or not action_index.exists():
        return resources

    data = read_json(action_index)
    rows = data.get("rows", data if isinstance(data, list) else [])
    for row in rows:
        unity_path = normalize_path(row.get("unity_path", ""))
        if not unity_path:
            continue
        inferred = infer_resource_audio_metadata(row)
        resource = ResourceRow(
            name=str(row.get("name", "")),
            kind=str(row.get("kind", "")),
            unity_path=unity_path,
            system=str(row.get("system", "")),
            group=str(row.get("group", "")),
            preview_ready=str(row.get("preview_ready", "")),
            audio_priority=str(inferred.get("audio_priority", "")),
            audio_candidate_score=int(inferred.get("audio_candidate_score", 0) or 0),
            recommended_config_layer=str(inferred.get("recommended_config_layer", "")),
            audio_keyword_hits=str(inferred.get("audio_keyword_hits", "")),
            why_audio_candidate=str(inferred.get("why_audio_candidate", "")),
        )
        resource.tokens = compact_tokens(split_name_tokens(" ".join([resource.name, resource.unity_path, resource.system, resource.group])))
        resources.append(resource)
    return resources


def load_jira_issues(path: Path) -> List[JiraIssue]:
    data = read_json(path)
    raw_issues = data.get("issues", data if isinstance(data, list) else [])
    issues: List[JiraIssue] = []
    for raw in raw_issues:
        key = str(raw.get("key") or raw.get("id") or "").upper()
        if not key:
            continue
        issue = JiraIssue(
            key=key,
            summary=str(raw.get("summary", "")),
            description=str(raw.get("description", "")),
            status=str(raw.get("status", "")),
            assignee=str(raw.get("assignee", "")),
            reporter=str(raw.get("reporter", "")),
            creator=str(raw.get("creator", "")),
            fix_versions=str(raw.get("fix_versions", "") or raw.get("version_label", "")),
            system=str(raw.get("system", "")),
            design_area=str(raw.get("design_area", "")),
            design_doc=str(raw.get("design_doc", "")),
            dependency_label=str(raw.get("dependency_label", "")),
            audio_required=str(raw.get("audio_required", "")),
            ready_state=str(raw.get("ready_state", "")),
            sound_type=str(raw.get("sound_type", "")),
            url=str(raw.get("url", "")),
        )
        related_text = " ".join([
            str(raw.get("issue_links", "")),
            " ".join(parse_listish(raw.get("qa_design_issues"))),
            " ".join(parse_listish(raw.get("qa_design_details"))),
        ])
        issue.related_keys = compact_tokens(extract_jira_keys(related_text), limit=30)
        issue.qa_doc_refs = safe_join(parse_listish(raw.get("qa_doc_refs")), "\n", 1200)
        issue.tokens = compact_tokens(
            split_name_tokens(
                " ".join([
                    issue.key,
                    issue.summary,
                    issue.description[:2000],
                    issue.system,
                    issue.design_area,
                    issue.design_doc,
                    issue.dependency_label,
                    issue.sound_type,
                    issue.creator,
                    issue.reporter,
                    issue.assignee,
                    related_text,
                    issue.qa_doc_refs,
                ])
            ),
            limit=120,
        )
        issues.append(issue)
    return issues


def row_token_set(row: Dict[str, Any], fallback_fields: Sequence[str], limit: int = 140) -> set[str]:
    cached = row.get("_match_tokens")
    if isinstance(cached, set):
        return cached
    tokens = row.get("tokens", [])
    if isinstance(tokens, list):
        base = [str(x).lower() for x in tokens if str(x).strip()]
    else:
        base = []
    if not base:
        base = split_name_tokens(" ".join(str(row.get(field, "")) for field in fallback_fields))
    result = set(compact_tokens(base, limit=limit))
    row["_match_tokens"] = result
    return result


def issue_core_tokens(issue: JiraIssue) -> set[str]:
    return set(compact_tokens(split_name_tokens(" ".join([
        issue.summary,
        issue.description[:1200],
        issue.dependency_label,
        issue.sound_type,
    ])), limit=90))


def strong_match_tokens(tokens: Iterable[str]) -> List[str]:
    strong: List[str] = []
    for token in tokens:
        if token in GENERIC_MATCH_TOKENS:
            continue
        has_cjk = bool(re.search(r"[\u4e00-\u9fff]", token))
        if has_cjk and len(token) >= 3:
            strong.append(token)
        elif not has_cjk and len(token) >= 5:
            strong.append(token)
    return strong


def score_design_row_for_issue(issue: JiraIssue, row: Dict[str, Any]) -> Tuple[int, List[str]]:
    score = 0
    reasons: List[str] = []
    doc_path = normalize_path(str(row.get("doc_path", "")))
    full_path = normalize_path(str(row.get("full_path", "")))
    text = " ".join([
        str(row.get("doc_path", "")),
        str(row.get("title", "")),
        str(row.get("feature", "")),
        str(row.get("system", "")),
        str(row.get("sound_type", "")),
        str(row.get("reason", "")),
        str(row.get("evidence", ""))[:1200],
    ])
    if issue.key and issue.key in text:
        score += 120
        reasons.append("direct Jira key in design chunk")
    row_core = row_token_set(row, ["title", "feature", "evidence", "system", "sound_type"])
    overlap = sorted((issue_core_tokens(issue) & row_core) - STOP_TOKENS)
    weighted = strong_match_tokens(overlap)
    if weighted:
        add = min(70, len(weighted) * 10)
        score += add
        reasons.append("design token overlap: " + ", ".join(weighted[:8]))
    if score == 0:
        return 0, []
    if issue.design_doc:
        design_doc_lc = lower_path(issue.design_doc)
        if design_doc_lc and (design_doc_lc in lower_path(doc_path) or design_doc_lc in lower_path(full_path) or Path(design_doc_lc).name in lower_path(full_path)):
            score += 35
            reasons.append("same design doc")
    if issue.system and issue.system.lower() in str(row.get("system", "")).lower():
        score += 10
        reasons.append("same system")
    if issue.sound_type and issue.sound_type.lower() in str(row.get("sound_type", "")).lower():
        score += 8
        reasons.append("same sound type")
    try:
        audio_score = float(row.get("audio_score", 0) or 0)
    except (TypeError, ValueError):
        audio_score = 0
    if score and audio_score >= 0.5:
        score += 10
        reasons.append("audio-scored design requirement")
    return score, reasons


def score_qa_row_for_issue(issue: JiraIssue, row: Dict[str, Any]) -> Tuple[int, List[str]]:
    score = 0
    reasons: List[str] = []
    story_text = " ".join(parse_listish(row.get("story_keys")))
    if issue.key and issue.key in story_text:
        score += 130
        reasons.append("QA case story key")
    related = set(issue.related_keys)
    qa_keys = set(extract_jira_keys(story_text))
    if related & qa_keys:
        score += 90
        reasons.append("QA case linked by related Jira")
    text = " ".join([
        str(row.get("doc_path", "")),
        str(row.get("doc_title", "")),
        str(row.get("system", "")),
        str(row.get("feature_point", "")),
        str(row.get("test_point", "")),
        str(row.get("operation_steps", "")),
        str(row.get("expected_result", "")),
        str(row.get("design_refs", "")),
        str(row.get("raw", ""))[:1200],
    ])
    overlap = sorted((issue_core_tokens(issue) & row_token_set(row, ["doc_title", "feature_point", "test_point", "operation_steps", "expected_result", "raw"])) - STOP_TOKENS)
    weighted = strong_match_tokens(t for t in overlap if t != "checklist")
    if weighted:
        add = min(45, len(weighted) * 6)
        score += add
        reasons.append("QA token overlap: " + ", ".join(weighted[:8]))
    if score == 0:
        return 0, []
    if issue.system and issue.system.lower() in text.lower():
        score += 8
        reasons.append("same QA system")
    return score, reasons


def format_design_evidence(row: Dict[str, Any], score: int, reasons: List[str]) -> str:
    return safe_join([
        f"{row.get('doc_path', '')} | {row.get('locator', '')}",
        f"feature: {row.get('feature', '')}",
        f"score: {score}; " + safe_join(reasons, "; ", 300),
        str(row.get("evidence", ""))[:600],
    ], "\n", 1200)


def format_qa_evidence(row: Dict[str, Any], score: int, reasons: List[str]) -> str:
    return safe_join([
        f"{row.get('doc_path', '')} | {row.get('locator', '')}",
        f"case: {row.get('feature_point', '')} / {row.get('test_point', '')}",
        f"expect: {row.get('expected_result', '')}",
        f"score: {score}; " + safe_join(reasons, "; ", 300),
    ], "\n", 1000)


def enrich_issues_with_design_qa(issues: List[JiraIssue], design_index: Optional[Path], qa_index: Optional[Path]) -> Dict[str, Dict[str, Any]]:
    evidence_by_key: Dict[str, Dict[str, Any]] = {}
    design_rows: List[Dict[str, Any]] = []
    qa_rows: List[Dict[str, Any]] = []
    if design_index and design_index.exists():
        data = read_json(design_index)
        design_rows = data.get("requirements", [])
    if qa_index and qa_index.exists():
        data = read_json(qa_index)
        qa_rows = data.get("cases", [])

    for issue in issues:
        design_scored: List[Tuple[int, Dict[str, Any], List[str]]] = []
        qa_scored: List[Tuple[int, Dict[str, Any], List[str]]] = []
        for row in design_rows:
            score, reasons = score_design_row_for_issue(issue, row)
            if score >= 55:
                design_scored.append((score, row, reasons))
        for row in qa_rows:
            score, reasons = score_qa_row_for_issue(issue, row)
            if score >= 65:
                qa_scored.append((score, row, reasons))
        design_scored.sort(key=lambda x: x[0], reverse=True)
        qa_scored.sort(key=lambda x: x[0], reverse=True)
        top_design = design_scored[:3]
        top_qa = qa_scored[:3]
        issue.design_evidence = safe_join([format_design_evidence(row, score, reasons) for score, row, reasons in top_design], "\n\n", 2500)
        issue.qa_evidence = safe_join([format_qa_evidence(row, score, reasons) for score, row, reasons in top_qa], "\n\n", 2200)
        doc_token_source: List[str] = []
        for score, row, _ in top_design:
            if score >= 75:
                doc_token_source.extend([
                    str(row.get("doc_path", "")),
                    str(row.get("title", "")),
                    str(row.get("feature", "")),
                    str(row.get("system", "")),
                    str(row.get("sound_type", "")),
                    str(row.get("evidence", ""))[:600],
                ])
        for score, row, _ in top_qa:
            if score >= 90:
                doc_token_source.extend([
                    str(row.get("doc_path", "")),
                    str(row.get("doc_title", "")),
                    str(row.get("feature_point", "")),
                    str(row.get("test_point", "")),
                    str(row.get("expected_result", "")),
                ])
        issue.doc_tokens = compact_tokens(split_name_tokens(" ".join(doc_token_source)), limit=80)
        issue.tokens = compact_tokens([*issue.tokens, *issue.doc_tokens], limit=180)
        evidence_by_key[issue.key] = {
            "design_count": len(design_scored),
            "qa_count": len(qa_scored),
            "design_evidence": issue.design_evidence,
            "qa_evidence": issue.qa_evidence,
            "doc_tokens": issue.doc_tokens,
        }
    return evidence_by_key


def load_p4_learning(path: Optional[Path]) -> Tuple[List[P4Change], Dict[str, List[P4Change]]]:
    changes: Dict[str, P4Change] = {}
    by_path: Dict[str, List[P4Change]] = defaultdict(list)
    if path is None or not path.exists():
        return [], by_path
    data = read_json(path)
    for ex in data.get("examples", []):
        change_id = str(ex.get("change") or ex.get("p4_change") or "")
        if not change_id:
            continue
        change = changes.setdefault(change_id, P4Change(change=change_id))
        change.date = change.date or str(ex.get("change_date", ""))
        change.owner = change.owner or str(ex.get("change_owner", ""))
        change.summary = change.summary or str(ex.get("task_goal", ""))
        rel_path = normalize_path(str(ex.get("rel_path", "")))
        if rel_path and rel_path not in change.files:
            change.files.append(rel_path)
        for key in extract_jira_keys(change.summary):
            if key not in change.jira_keys:
                change.jira_keys.append(key)
        change.audio_related = True
        change.reason = change.reason or "p4 learning cache audio example"
    for change in changes.values():
        for f in change.files:
            by_path[lower_path(f)].append(change)
    return list(changes.values()), by_path


def parse_p4_changes(output: str) -> List[P4Change]:
    changes: List[P4Change] = []
    # Change 72219 on 2026/06/23 by USER012@client 'summary'
    pat = re.compile(r"^Change\s+(\d+)\s+on\s+(\S+)\s+by\s+(\S+)\s+'(.*)'$", re.MULTILINE)
    for m in pat.finditer(output):
        summary = m.group(4)
        change = P4Change(
            change=m.group(1),
            date=m.group(2),
            owner=m.group(3),
            summary=summary,
            jira_keys=extract_jira_keys(summary),
        )
        changes.append(change)
    return changes


def parse_p4_describe(output: str, change: P4Change) -> None:
    files: List[str] = []
    in_files = False
    for line in output.splitlines():
        if line.startswith("Affected files ..."):
            in_files = True
            continue
        if in_files:
            if not line.strip():
                continue
            # ... //depot/path#rev action
            m = re.match(r"\.\.\.\s+(//.+?)(?:#\d+)?\s+(\w+)", line.strip())
            if m:
                files.append(normalize_path(m.group(1)))
    change.files = files
    text = " ".join([change.summary, " ".join(files)])
    change.jira_keys = compact_tokens(extract_jira_keys(text), limit=20)
    audio_file = any(text_has_any(f, AUDIO_PATH_HINTS) or f.lower().endswith((".bnk", ".wem", ".wwu", ".wav")) for f in files)
    audio_text = text_has_any(change.summary, AUDIO_TEXT_HINTS)
    action_file = any(text_has_any(f, RESOURCE_ACTION_HINTS) and f.lower().endswith((".prefab", ".anim", ".playable")) for f in files)
    change.audio_related = bool(audio_file or audio_text or (change.jira_keys and action_file))
    reasons = []
    if audio_file:
        reasons.append("audio/Wwise file path")
    if audio_text:
        reasons.append("audio keyword in changelist summary")
    if action_file:
        reasons.append("action resource file in Jira-linked changelist")
    change.reason = "; ".join(reasons)


def collect_recent_p4_audio_changes(max_changes: int, since: str, scope: str, describe_limit: int) -> List[P4Change]:
    if max_changes <= 0:
        return []
    spec = f"{scope}@{since},@now" if since else scope
    code, out, err = p4_run(["changes", "-m", str(max_changes), spec], timeout=60)
    if code != 0:
        return []
    changes = parse_p4_changes(out)
    audio_changes: List[P4Change] = []
    for i, change in enumerate(changes[:describe_limit]):
        code, desc, _ = p4_run(["describe", "-s", change.change], timeout=30)
        if code != 0:
            continue
        parse_p4_describe(desc, change)
        if change.audio_related:
            audio_changes.append(change)
    return audio_changes


def build_path_change_index(changes: Iterable[P4Change]) -> Dict[str, List[P4Change]]:
    by_path: Dict[str, List[P4Change]] = defaultdict(list)
    for change in changes:
        for f in change.files:
            by_path[lower_path(f)].append(change)
    return by_path


def path_tail_variants(path: str) -> List[str]:
    p = lower_path(path)
    variants = [p]
    marker = "/client/targetproject/"
    if marker in p:
        variants.append(p.split(marker, 1)[1])
    marker2 = "/targetproject/"
    if marker2 in p:
        variants.append(p.split(marker2, 1)[1])
    if p.startswith("assets/"):
        variants.append(p)
    return compact_tokens(variants, limit=10)


def find_resource_p4_changes(resource: ResourceRow, path_index: Dict[str, List[P4Change]]) -> List[P4Change]:
    matches: List[P4Change] = []
    seen = set()
    variants = path_tail_variants(resource.unity_path)
    for p4_path, changes in path_index.items():
        for variant in variants:
            if variant and variant in p4_path:
                for change in changes:
                    if change.change not in seen:
                        seen.add(change.change)
                        matches.append(change)
                break
    return matches


def score_issue_for_resource(resource: ResourceRow, issue: JiraIssue, p4_changes: List[P4Change]) -> Tuple[int, List[str], List[str]]:
    score = 0
    semantic_score = 0
    direct_p4_score = 0
    reasons: List[str] = []
    evidence: List[str] = []
    resource_text = " ".join([resource.name, resource.unity_path, resource.system, resource.group, resource.audio_keyword_hits]).lower()
    issue_text = " ".join([
        issue.key,
        issue.summary,
        issue.description,
        issue.system,
        issue.design_area,
        issue.design_doc,
        issue.dependency_label,
        issue.sound_type,
        issue.design_evidence[:2000],
        issue.qa_evidence[:1200],
    ]).lower()

    p4_keys = set()
    p4_ids = []
    for ch in p4_changes:
        p4_ids.append(ch.change)
        p4_keys.update(ch.jira_keys)
    if issue.key in p4_keys:
        direct_p4_score += 100
        reasons.append("same P4 changelist edited this resource and mentions Jira key")
        evidence.append("P4:" + ",".join(p4_ids[:8]))

    exact_name = resource.name.lower()
    if exact_name and exact_name in issue_text:
        score += 90
        semantic_score += 90
        reasons.append("Jira text contains exact resource name")

    stem = Path(resource.name).stem.lower()
    if stem and len(stem) >= 6 and stem in issue_text and stem != exact_name:
        score += 60
        semantic_score += 60
        reasons.append("Jira text contains resource stem")

    if resource.system and resource.system.lower() in issue_text:
        score += 18
        semantic_score += 18
        reasons.append("resource system appears in Jira text")
    if resource.group and resource.group.lower() in issue_text:
        score += 18
        semantic_score += 18
        reasons.append("resource group appears in Jira text")

    rtoks = set(resource.tokens)
    itoks = set(issue.tokens)
    overlap = sorted((rtoks & itoks) - STOP_TOKENS)
    weighted_overlap = [t for t in overlap if len(t) >= 4 or re.search(r"[\u4e00-\u9fff]", t)]
    if weighted_overlap:
        add = min(45, len(weighted_overlap) * 8)
        score += add
        semantic_score += add
        reasons.append("token overlap: " + ", ".join(weighted_overlap[:8]))

    audio_resource = (
        resource.audio_priority in {"High", "Medium"}
        or text_has_any(resource_text, AUDIO_TEXT_HINTS)
        or text_has_any(resource_text, RESOURCE_ACTION_HINTS)
    )
    audio_issue = issue.audio_required.lower() in {"yes", "maybe", "true"} or text_has_any(issue_text, AUDIO_TEXT_HINTS)
    if audio_resource and audio_issue:
        score += 20
        reasons.append("both sides are audio-relevant")

    if issue.sound_type and any(t in resource_text for t in split_name_tokens(issue.sound_type)):
        score += 12
        semantic_score += 12
        reasons.append("sound type matches resource text")

    score += direct_p4_score
    if direct_p4_score and semantic_score == 0:
        # A mixed changelist can contain unrelated resource edits under one Jira key.
        # Keep it as useful evidence, but do not call it high confidence by itself.
        score = min(score, 95)
        reasons.append("P4 direct evidence only; mixed changelist risk, review manually")

    if score and issue.ready_state:
        evidence.append("Ready:" + issue.ready_state)
    if score and issue.status:
        evidence.append("Status:" + issue.status)

    return score, reasons, evidence


def confidence(score: int) -> str:
    if score >= 120:
        return "High"
    if score >= 75:
        return "Medium"
    if score >= 45:
        return "Low"
    return "Weak"


def safe_join(values: Iterable[str], sep: str = "; ", limit: int = 500) -> str:
    text = sep.join(v for v in values if v)
    return text[:limit]


def resource_public(resource: ResourceRow) -> Dict[str, Any]:
    return {
        "name": resource.name,
        "kind": resource.kind,
        "unity_path": resource.unity_path,
        "system": resource.system,
        "group": resource.group,
        "audio_priority": resource.audio_priority,
        "audio_candidate_score": resource.audio_candidate_score,
        "recommended_config_layer": resource.recommended_config_layer,
        "audio_keyword_hits": resource.audio_keyword_hits,
        "why_audio_candidate": resource.why_audio_candidate,
        "preview_ready": resource.preview_ready,
    }


def resource_type_sheet(resource: Dict[str, Any]) -> str:
    kind = str(resource.get("kind", ""))
    layer = str(resource.get("recommended_config_layer", ""))
    system = str(resource.get("system", ""))
    path = str(resource.get("unity_path", "")).lower()
    if "Timeline" in kind:
        return "Type_Timeline"
    if "VFX" in kind or "/vfx" in path or "/fx/" in path:
        return "Type_VFX"
    if layer == "UI Audio Config" or system == "UI" or "/ui/" in path:
        return "Type_UI"
    if kind == "AnimationClip":
        return "Type_Animation"
    if "Prefab" in kind:
        return "Type_Prefab"
    return "Type_Other"


def matrix_resource_bucket(resource: Dict[str, Any]) -> str:
    kind = str(resource.get("kind", ""))
    layer = str(resource.get("recommended_config_layer", ""))
    system = str(resource.get("system", ""))
    path = str(resource.get("unity_path", "")).lower()
    name = str(resource.get("name", "")).lower()
    if "Timeline" in kind or path.endswith(".playable") or name.startswith("tml_"):
        return "timelines"
    if kind == "AnimationClip" or path.endswith(".anim"):
        return "animations"
    if "VFX" in kind or "/vfx" in path or "/fx/" in path or "pfb_vfx" in name:
        return "vfx_prefabs"
    if "wwise" in path or "wwisebank" in path or path.endswith(".bnk"):
        return "wwise_assets"
    if path.endswith(".cs"):
        return "scripts"
    if "uisoundsavedata" in path or path.endswith((".xlsx", ".json", ".asset")) and ("config" in path or "sound" in path or "audio" in path):
        return "config_assets"
    if layer == "UI Audio Config" or system == "UI" or "/ui/" in path or name.startswith("pfb_ui"):
        return "ui_prefabs"
    if "Prefab" in kind or path.endswith(".prefab"):
        return "prefabs"
    return "other_resources"


def resource_cell_entry(link: Dict[str, Any]) -> str:
    resource = link.get("resource", {})
    name = str(resource.get("name", ""))
    path = str(resource.get("unity_path", ""))
    conf = str(link.get("confidence", ""))
    score = link.get("score", "")
    if name and path:
        return f"{name} ({conf},{score})\n{path}"
    return f"{path or name} ({conf},{score})"


def link_evidence_text(link: Dict[str, Any]) -> str:
    return " ".join([
        str(link.get("why", "")),
        str(link.get("evidence", "")),
        str(link.get("resource", {}).get("unity_path", "")),
        str(link.get("resource", {}).get("name", "")),
    ]).lower()


def link_decision(link: Dict[str, Any]) -> str:
    text = link_evidence_text(link)
    score = int(link.get("score", 0) or 0)
    conf = str(link.get("confidence", ""))
    if "jira-resource seed workbook" in text:
        return "Confirmed"
    if "same p4 changelist" in text and ("exact resource name" in text or "resource stem" in text):
        return "Confirmed"
    if score >= 120 and ("exact resource name" in text or "resource stem" in text):
        return "Confirmed"
    if score >= 120 and "token overlap" in text:
        return "Likely"
    if conf in {"High", "Medium"} and ("exact resource name" in text or "resource stem" in text):
        return "Likely"
    if conf == "High" and ("design" in text or "qa" in text or "p4" in text):
        return "Likely"
    return "ReviewOnly"


def is_strict_link(link: Dict[str, Any]) -> bool:
    return link_decision(link) in {"Confirmed", "Likely"}


def can_export_thumbnail(asset_path: str) -> bool:
    path = normalize_path(asset_path)
    if not path.lower().startswith("assets/"):
        return False
    ext = Path(path).suffix.lower()
    return ext in {".prefab", ".playable", ".anim", ".asset", ".controller", ".mat", ".png", ".jpg", ".jpeg", ".tga", ".fbx"}


def thumbnail_bucket_rank(bucket: str) -> int:
    order = {
        "ui_prefabs": 0,
        "timelines": 1,
        "vfx_prefabs": 2,
        "animations": 3,
        "prefabs": 4,
        "config_assets": 5,
        "other_resources": 6,
        "wwise_assets": 7,
        "scripts": 8,
    }
    return order.get(bucket, 9)


def build_thumbnail_candidates(links: Sequence[Dict[str, Any]], thumbnail_dir: Path, limit: int = 3) -> List[Dict[str, Any]]:
    confidence_rank = {"High": 0, "Medium": 1, "Low": 2, "Weak": 3, "": 4}
    candidates: List[Dict[str, Any]] = []
    seen: set[str] = set()
    for link in links:
        resource = link.get("resource", {})
        asset_path = normalize_path(str(resource.get("unity_path", "")))
        if not can_export_thumbnail(asset_path):
            continue
        path_key = asset_path.lower()
        if path_key in seen:
            continue
        seen.add(path_key)
        bucket = matrix_resource_bucket(resource)
        candidates.append({
            "asset_path": asset_path,
            "asset_name": str(resource.get("name", "")) or Path(asset_path).stem,
            "bucket": bucket,
            "confidence": str(link.get("confidence", "")),
            "score": int(link.get("score", 0) or 0),
            "thumbnail_file": str(thumbnail_path_for_asset(thumbnail_dir, asset_path)),
        })
    candidates.sort(key=lambda item: (
        confidence_rank.get(str(item.get("confidence", "")), 9),
        thumbnail_bucket_rank(str(item.get("bucket", ""))),
        -int(item.get("score", 0) or 0),
        str(item.get("asset_path", "")),
    ))
    return candidates[:limit]


def infer_seed_kind(path: str, name: str) -> str:
    text = " ".join([path, name]).lower()
    if text.endswith(".playable") or "/timeline/" in text or "tml_" in text:
        return "Timeline"
    if text.endswith(".anim"):
        return "AnimationClip"
    if "vfx" in text or "/fx/" in text:
        return "VFX Prefab"
    if text.endswith(".prefab"):
        return "Prefab"
    return "SeedResource"


def infer_seed_confidence(note: str) -> Tuple[str, int]:
    text = note or ""
    if "\u9ad8" in text:
        return "High", 130
    if "\u4e2d" in text:
        return "Medium", 85
    if "\u4f4e" in text:
        return "Low", 55
    return "Medium", 75


def extract_seed_paths(note: str) -> List[str]:
    text = (note or "").replace("\\", "/")
    matches = re.findall(r"Assets/GameProject/.*?\.(?:prefab|playable|anim|asset|bnk)", text, flags=re.IGNORECASE)
    cleaned: List[str] = []
    for item in matches:
        next_assets = item.find("Assets/GameProject/", 1)
        if next_assets > 0:
            item = item[:next_assets]
        cleaned.append(item.strip(" ;,\t\r\n"))
    return cleaned


def split_seed_resource_names(value: str) -> List[str]:
    if not value:
        return []
    names: List[str] = []
    for part in str(value).split("|"):
        for sub in part.split(","):
            item = sub.strip()
            if item:
                names.append(item)
    return names


def load_jira_resource_seed_links(paths: Sequence[Path]) -> Dict[str, List[Dict[str, Any]]]:
    seed_links: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for path in paths:
        if not path or not path.exists():
            continue
        wb = load_workbook(path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if idx == 1:
                    continue
                values = list(row) + [None, None, None]
                issue_text = str(values[0] or "")
                key = extract_jira_keys(issue_text)
                if not key:
                    continue
                jira_key = key[0]
                names = split_seed_resource_names(str(values[1] or ""))
                note = str(values[2] or "")
                paths_in_note = extract_seed_paths(note)
                if not names and paths_in_note:
                    names = [Path(p.replace("\\", "/")).name for p in paths_in_note]
                path_by_name: Dict[str, List[str]] = defaultdict(list)
                for resource_path in paths_in_note:
                    path_by_name[Path(resource_path.replace("\\", "/")).name.lower()].append(resource_path)

                conf, score = infer_seed_confidence(note)
                for name_idx, name in enumerate(names):
                    name_key = Path(name.replace("\\", "/")).name.lower()
                    resource_path = ""
                    if path_by_name.get(name_key):
                        resource_path = path_by_name[name_key][0]
                    elif name_idx < len(paths_in_note):
                        resource_path = paths_in_note[name_idx]
                    resource_name = Path((resource_path or name).replace("\\", "/")).stem or name
                    kind = infer_seed_kind(resource_path, name)
                    resource_text_lc = (resource_path or name).lower()
                    is_ui_resource = "/ui/" in resource_text_lc or Path(name.replace("\\", "/")).name.lower().startswith("pfb_ui")
                    resource = {
                        "name": resource_name,
                        "kind": kind,
                        "unity_path": normalize_path(resource_path or name),
                        "system": "UI" if is_ui_resource else "",
                        "group": "JiraResourceSeed",
                        "audio_priority": conf,
                        "recommended_config_layer": "UI Audio Config" if is_ui_resource else "Jira Resource Seed",
                    }
                    seed_links[jira_key].append({
                        "confidence": conf,
                        "score": score,
                        "resource": resource,
                        "jira": {"key": jira_key},
                        "why": f"Jira-resource seed workbook: {path.name}",
                        "evidence": note[:1000],
                    })
    return seed_links


def issue_audio_relevant(issue: JiraIssue) -> bool:
    if issue.audio_required.lower() in {"yes", "maybe", "true"}:
        return True
    hay = " ".join([issue.summary, issue.description, issue.sound_type, issue.design_area])
    return text_has_any(hay, AUDIO_TEXT_HINTS)


def build_jira_resource_matrix(issues: List[JiraIssue], jira_links: Dict[str, List[Dict[str, Any]]], thumbnail_dir: Path = DEFAULT_THUMBNAIL_DIR) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    confidence_rank = {"High": 0, "Medium": 1, "Low": 2, "Weak": 3, "": 4}
    decision_rank = {"Confirmed": 0, "Likely": 1, "ReviewOnly": 2, "NeedsEvidence": 3, "": 4}
    bucket_order = [
        "ui_prefabs",
        "timelines",
        "animations",
        "vfx_prefabs",
        "prefabs",
        "wwise_assets",
        "config_assets",
        "scripts",
        "other_resources",
    ]

    def bucket_cells_for(bucket_source_links: Sequence[Dict[str, Any]], strict: bool) -> Dict[str, str]:
        buckets: Dict[str, List[Dict[str, Any]]] = {name: [] for name in bucket_order}
        bucket_paths: Dict[str, set[str]] = {name: set() for name in bucket_order}
        for link in bucket_source_links:
            resource = link.get("resource", {})
            bucket = matrix_resource_bucket(resource)
            path = str(resource.get("unity_path", ""))
            if path in bucket_paths[bucket]:
                continue
            bucket_paths[bucket].add(path)
            buckets[bucket].append(link)

        bucket_cells: Dict[str, str] = {}
        for bucket, bucket_links in buckets.items():
            if strict:
                selected = bucket_links[:10]
                hidden = len(bucket_links) - len(selected)
                entries = [f"{link_decision(x)} | {resource_cell_entry(x)}" for x in selected]
                if hidden > 0:
                    entries.append(f"... +{hidden} more strict candidates")
            else:
                strong = [x for x in bucket_links if x.get("confidence") in {"High", "Medium"}]
                if strong:
                    selected = strong[:8]
                    hidden = len(strong) - len(selected)
                    weak_hidden = len(bucket_links) - len(strong)
                else:
                    selected = bucket_links[:5]
                    hidden = len(bucket_links) - len(selected)
                    weak_hidden = 0
                entries = [resource_cell_entry(x) for x in selected]
                if hidden > 0:
                    entries.append(f"... +{hidden} more candidates")
                if weak_hidden > 0:
                    entries.append(f"... +{weak_hidden} Low/Weak candidates hidden")
            bucket_cells[bucket] = safe_join(entries, "\n\n", 3000)
        return bucket_cells

    for issue in issues:
        links = sorted(jira_links.get(issue.key, []), key=lambda x: x.get("score", 0), reverse=True)
        if not links and not issue_audio_relevant(issue):
            continue
        for link in links:
            link["decision"] = link_decision(link)
        strict_links = [x for x in links if is_strict_link(x)]
        strict_links.sort(key=lambda x: (
            decision_rank.get(str(x.get("decision", "")), 9),
            confidence_rank.get(str(x.get("confidence", "")), 9),
            -int(x.get("score", 0) or 0),
        ))
        top = links[0] if links else None
        strict_top = strict_links[0] if strict_links else None
        bucket_cells = bucket_cells_for(links, strict=False)
        strict_bucket_cells = bucket_cells_for(strict_links, strict=True)

        rows.append({
            "confidence": top.get("confidence", "") if top else "",
            "best_score": top.get("score", 0) if top else 0,
            "resource_count": len({str(x.get("resource", {}).get("unity_path", "")) for x in links}),
            "decision": strict_top.get("decision", "NeedsEvidence") if strict_top else "NeedsEvidence",
            "strict_confidence": strict_top.get("confidence", "") if strict_top else "",
            "strict_best_score": strict_top.get("score", 0) if strict_top else 0,
            "strict_resource_count": len({str(x.get("resource", {}).get("unity_path", "")) for x in strict_links}),
            "thumbnail_candidates": build_thumbnail_candidates(links, thumbnail_dir, limit=3),
            "jira": {
                "key": issue.key,
                "summary": issue.summary,
                "status": issue.status,
                "ready_state": issue.ready_state,
                "system": issue.system,
                "creator": issue.creator,
                "assignee": issue.assignee,
                "design_doc": issue.design_doc,
                "audio_required": issue.audio_required,
                "sound_type": issue.sound_type,
                "url": issue.url,
                "design_evidence": issue.design_evidence,
                "qa_evidence": issue.qa_evidence,
            },
            "buckets": bucket_cells,
            "strict_buckets": strict_bucket_cells,
            "evidence": safe_join([x.get("evidence", "") for x in links[:5]], "\n", 1500),
            "why": safe_join([x.get("why", "") for x in links[:5]], "\n", 1500),
            "strict_evidence": safe_join([x.get("evidence", "") for x in strict_links[:5]], "\n", 1500),
            "strict_why": safe_join([x.get("why", "") for x in strict_links[:5]], "\n", 1500),
        })

    rows.sort(key=lambda x: (
        decision_rank.get(str(x.get("decision", "")), 9),
        -int(x.get("strict_resource_count", 0) or 0),
        confidence_rank.get(str(x.get("confidence", "")), 9),
        -int(x.get("best_score", 0) or 0),
        x.get("jira", {}).get("key", ""),
    ))
    return rows


def build_thumbnail_manifest_assets(matrix_rows: Sequence[Dict[str, Any]], assets_per_jira: int, max_jiras: int) -> List[Dict[str, Any]]:
    assets: List[Dict[str, Any]] = []
    if assets_per_jira <= 0 or max_jiras <= 0:
        return assets
    used_pairs: set[tuple[str, str]] = set()
    added_jiras = 0
    for row in matrix_rows:
        jira = row.get("jira", {})
        key = str(jira.get("key", ""))
        if not key:
            continue
        candidates = row.get("thumbnail_candidates", []) or []
        if not candidates:
            continue
        added_for_jira = 0
        for candidate in candidates:
            asset_path = normalize_path(str(candidate.get("asset_path", "")))
            pair = (key, asset_path.lower())
            if not asset_path or pair in used_pairs:
                continue
            used_pairs.add(pair)
            item = dict(candidate)
            item.update({
                "jira_key": key,
                "jira_summary": str(jira.get("summary", "")),
                "jira_status": str(jira.get("status", "")),
                "jira_ready": str(jira.get("ready_state", "")),
                "jira_url": str(jira.get("url", "")),
            })
            assets.append(item)
            added_for_jira += 1
            if added_for_jira >= assets_per_jira:
                break
        if added_for_jira:
            added_jiras += 1
        if added_jiras >= max_jiras:
            break
    return assets


def count_existing_thumbnail_images(assets: Sequence[Dict[str, Any]]) -> int:
    return sum(1 for item in assets if Path(str(item.get("thumbnail_file", ""))).exists())


def compact_excel_cell(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    if not value:
        return value
    text = value.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"\n+", " | ", text)
    text = re.sub(r"[ \t]+", " ", text)
    return text.strip()


def write_excel(path: Path, data: Dict[str, Any]) -> None:
    compact_no_image = int(data.get("summary", {}).get("thumbnail_images_found", 0) or 0) <= 0
    wb = Workbook()
    ws = wb.active
    ws.title = "README"
    thumbnail_note = (
        "No real thumbnail PNGs were found. Excel is written in compact no-image mode; JSON keeps full evidence."
        if compact_no_image
        else "Run Unity menu ProjectEF/Audio/Export Audio Resource Thumbnails, then rebuild this report to embed images."
    )
    for row in [
        ["ProjectEF Audio Resource <-> Jira Link Report"],
        ["Generated", data["generated_at"]],
        ["Resource index", data["inputs"].get("resource_index", "")],
        ["Jira cache", data["inputs"].get("jira_cache", "")],
        ["Design index", data["inputs"].get("design_index", "")],
        ["QA index", data["inputs"].get("qa_index", "")],
        ["P4 learning", data["inputs"].get("p4_learning", "")],
        ["Jira-resource seeds", "\n".join(data["inputs"].get("jira_resource_seeds", []))],
        ["Thumbnail manifest", data.get("thumbnail", {}).get("manifest", "")],
        ["Thumbnail directory", data.get("thumbnail", {}).get("dir", "")],
        ["Thumbnail note", thumbnail_note],
        ["Recent P4 audio changes", str(data["summary"].get("recent_p4_audio_changes", 0))],
        ["Notes", "Read-only report. Links are evidence-ranked suggestions, not final production truth."],
    ]:
        ws.append(row)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 120
    ws["A1"].font = Font(bold=True, size=14)

    def add_sheet(name: str, headers: List[str], rows: List[List[Any]]):
        sh = wb.create_sheet(name)
        sh.append(headers)
        for cell in sh[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        sh.row_dimensions[1].height = 22
        for row in rows:
            sh.append([compact_excel_cell(value) for value in row] if compact_no_image else row)
        sh.freeze_panes = "A2"
        sh.auto_filter.ref = sh.dimensions
        for idx, header in enumerate(headers, 1):
            width = 16
            if header in {
                "resource_path",
                "jira_summary",
                "top_jira_summary",
                "why",
                "evidence",
                "url",
                "top_jira_url",
                "design_doc",
                "asset_path",
                "all_jira_links",
                "ui_prefabs",
                "timelines",
                "animations",
                "vfx_prefabs",
                "prefabs",
                "wwise_assets",
                "config_assets",
                "scripts",
                "other_resources",
                "design_evidence",
                "qa_evidence",
                "thumbnail_asset",
                "thumbnail_file",
            }:
                width = 44 if compact_no_image else 60
            elif header == "thumbnail":
                width = 10 if compact_no_image else 18
            elif header in {"resource_name", "jira_key", "top_jira_key"}:
                width = 28
            elif header in {"top_jira_links", "top_resources"}:
                width = 50 if compact_no_image else 80
            column_letter = get_column_letter(idx)
            sh.column_dimensions[column_letter].width = width
            if compact_no_image and header in {"thumbnail", "thumbnail_asset", "thumbnail_file"}:
                sh.column_dimensions[column_letter].hidden = True
        header_index = {h: i + 1 for i, h in enumerate(headers)}
        for row_idx in range(2, sh.max_row + 1):
            if compact_no_image:
                sh.row_dimensions[row_idx].height = 18
            url = ""
            for url_header in ("url", "top_jira_url", "jira_url"):
                col = header_index.get(url_header)
                if col:
                    url = str(sh.cell(row=row_idx, column=col).value or "")
                    if url:
                        break
            if url:
                for link_header in ("jira_key", "jira_summary", "top_jira_key", "top_jira_summary"):
                    col = header_index.get(link_header)
                    if col:
                        cell = sh.cell(row=row_idx, column=col)
                        if cell.value:
                            cell.hyperlink = url
                            cell.style = "Hyperlink"
        for cells in sh.iter_rows(min_row=2, max_row=sh.max_row):
            conf = str(cells[0].value if cells else "")
            fill = None
            if conf == "Confirmed":
                fill = PatternFill("solid", fgColor="C6EFCE")
            elif conf == "Likely":
                fill = PatternFill("solid", fgColor="D9EAF7")
            elif conf == "NeedsEvidence":
                fill = PatternFill("solid", fgColor="D9D9D9")
            elif conf == "High":
                fill = PatternFill("solid", fgColor="FFC7CE")
            elif conf == "Medium":
                fill = PatternFill("solid", fgColor="FFEB9C")
            elif conf == "Low":
                fill = PatternFill("solid", fgColor="E2F0D9")
            if fill:
                fill_cells = cells[: min(4, len(cells))] if compact_no_image else cells
                for c in fill_cells:
                    c.fill = fill
            for c in cells:
                c.alignment = Alignment(
                    wrap_text=not compact_no_image,
                    vertical="center" if compact_no_image else "top",
                )
        return sh

    def embed_matrix_thumbnails(sh, headers: List[str], matrix_items: Sequence[Dict[str, Any]]) -> None:
        if compact_no_image or XLImage is None or not data.get("thumbnail", {}).get("embed_enabled", True):
            return
        header_index = {h: i + 1 for i, h in enumerate(headers)}
        thumb_col = header_index.get("thumbnail")
        if not thumb_col:
            return
        sh.column_dimensions[get_column_letter(thumb_col)].width = 20
        for row_offset, item in enumerate(matrix_items, start=2):
            candidates = item.get("thumbnail_candidates", []) or []
            if not candidates:
                continue
            image_path = Path(str(candidates[0].get("thumbnail_file", "")))
            if not image_path.exists():
                continue
            try:
                image = XLImage(str(image_path))
                image.width = 128
                image.height = 96
                sh.row_dimensions[row_offset].height = 78
                sh.add_image(image, f"{get_column_letter(thumb_col)}{row_offset}")
            except Exception:
                continue

    link_headers = [
        "confidence",
        "score",
        "resource_name",
        "resource_kind",
        "resource_priority",
        "recommended_layer",
        "resource_system",
        "resource_group",
        "resource_path",
        "jira_key",
        "jira_summary",
        "jira_status",
        "jira_ready",
        "jira_system",
        "jira_creator",
        "jira_assignee",
        "design_doc",
        "why",
        "evidence",
        "url",
    ]
    link_rows = []
    for item in data["links"]:
        link_rows.append([
            item["confidence"],
            item["score"],
            item["resource"]["name"],
            item["resource"]["kind"],
            item["resource"].get("audio_priority", ""),
            item["resource"].get("recommended_config_layer", ""),
            item["resource"].get("system", ""),
            item["resource"].get("group", ""),
            item["resource"]["unity_path"],
            item["jira"]["key"],
            item["jira"]["summary"],
            item["jira"].get("status", ""),
            item["jira"].get("ready_state", ""),
            item["jira"].get("system", ""),
            item["jira"].get("creator", ""),
            item["jira"].get("assignee", ""),
            item["jira"].get("design_doc", ""),
            item["why"],
            item["evidence"],
            item["jira"].get("url", ""),
        ])
    add_sheet("Resource_Jira_Links", link_headers, link_rows)

    catalog_headers = [
        "confidence",
        "best_score",
        "resource_name",
        "resource_kind",
        "resource_priority",
        "recommended_layer",
        "resource_system",
        "resource_group",
        "preview_ready",
        "keyword_hits",
        "resource_path",
        "top_jira_key",
        "top_jira_summary",
        "top_jira_status",
        "top_jira_ready",
        "top_jira_url",
        "all_jira_links",
        "why_audio_candidate",
        "why_jira_linked",
    ]

    def catalog_rows(items: List[Dict[str, Any]]) -> List[List[Any]]:
        rows: List[List[Any]] = []
        for item in items:
            r = item["resource"]
            top = item.get("top_link") or {}
            top_jira = top.get("jira", {}) if top else {}
            rows.append([
                item.get("best_confidence", ""),
                item.get("best_score", 0),
                r.get("name", ""),
                r.get("kind", ""),
                r.get("audio_priority", ""),
                r.get("recommended_config_layer", ""),
                r.get("system", ""),
                r.get("group", ""),
                r.get("preview_ready", ""),
                r.get("audio_keyword_hits", ""),
                r.get("unity_path", ""),
                top_jira.get("key", ""),
                top_jira.get("summary", ""),
                top_jira.get("status", ""),
                top_jira.get("ready_state", ""),
                top_jira.get("url", ""),
                item.get("all_jira_links", ""),
                r.get("why_audio_candidate", ""),
                item.get("why", ""),
            ])
        return rows

    resource_catalog = data.get("resource_catalog", [])
    add_sheet("All_Audio_Candidates", catalog_headers, catalog_rows(resource_catalog))
    type_order = ["Type_Timeline", "Type_Animation", "Type_Prefab", "Type_VFX", "Type_UI", "Type_Other"]
    for type_name in type_order:
        items = [item for item in resource_catalog if item.get("type_sheet") == type_name]
        if items:
            add_sheet(type_name, catalog_headers, catalog_rows(items))

    res_headers = [
        "confidence",
        "best_score",
        "resource_name",
        "resource_kind",
        "resource_priority",
        "recommended_layer",
        "resource_path",
        "top_jira_links",
        "why",
    ]
    res_rows = []
    for item in data["resource_summary"]:
        res_rows.append([
            item["best_confidence"],
            item["best_score"],
            item["resource"]["name"],
            item["resource"]["kind"],
            item["resource"].get("audio_priority", ""),
            item["resource"].get("recommended_config_layer", ""),
            item["resource"]["unity_path"],
            item["top_jira_links"],
            item["why"],
        ])
    add_sheet("By_Resource", res_headers, res_rows)

    matrix_headers = [
        "decision",
        "confidence",
        "best_score",
        "resource_count",
        "strict_confidence",
        "strict_best_score",
        "strict_resource_count",
        "thumbnail",
        "thumbnail_asset",
        "thumbnail_file",
        "jira_key",
        "jira_summary",
        "jira_status",
        "jira_ready",
        "jira_system",
        "creator",
        "assignee",
        "audio_required",
        "sound_type",
        "design_doc",
        "design_evidence",
        "qa_evidence",
        "ui_prefabs",
        "timelines",
        "animations",
        "vfx_prefabs",
        "prefabs",
        "wwise_assets",
        "config_assets",
        "scripts",
        "other_resources",
        "why",
        "evidence",
        "url",
    ]
    matrix_rows = []
    for item in data.get("jira_resource_matrix", []):
        jira = item.get("jira", {})
        buckets = item.get("buckets", {})
        thumbnail_candidates = item.get("thumbnail_candidates", []) or []
        thumbnail = thumbnail_candidates[0] if thumbnail_candidates else {}
        matrix_rows.append([
            item.get("decision", ""),
            item.get("confidence", ""),
            item.get("best_score", 0),
            item.get("resource_count", 0),
            item.get("strict_confidence", ""),
            item.get("strict_best_score", 0),
            item.get("strict_resource_count", 0),
            "image" if thumbnail and Path(str(thumbnail.get("thumbnail_file", ""))).exists() else "",
            thumbnail.get("asset_path", ""),
            thumbnail.get("thumbnail_file", ""),
            jira.get("key", ""),
            jira.get("summary", ""),
            jira.get("status", ""),
            jira.get("ready_state", ""),
            jira.get("system", ""),
            jira.get("creator", ""),
            jira.get("assignee", ""),
            jira.get("audio_required", ""),
            jira.get("sound_type", ""),
            jira.get("design_doc", ""),
            jira.get("design_evidence", ""),
            jira.get("qa_evidence", ""),
            buckets.get("ui_prefabs", ""),
            buckets.get("timelines", ""),
            buckets.get("animations", ""),
            buckets.get("vfx_prefabs", ""),
            buckets.get("prefabs", ""),
            buckets.get("wwise_assets", ""),
            buckets.get("config_assets", ""),
            buckets.get("scripts", ""),
            buckets.get("other_resources", ""),
            item.get("why", ""),
            item.get("evidence", ""),
            jira.get("url", ""),
        ])
    matrix_sheet = add_sheet("Jira_Resource_Matrix", matrix_headers, matrix_rows)
    embed_matrix_thumbnails(matrix_sheet, matrix_headers, data.get("jira_resource_matrix", []))

    strict_headers = [
        "decision",
        "strict_confidence",
        "strict_best_score",
        "strict_resource_count",
        "jira_key",
        "jira_summary",
        "jira_status",
        "jira_ready",
        "jira_system",
        "creator",
        "assignee",
        "audio_required",
        "sound_type",
        "design_doc",
        "design_evidence",
        "qa_evidence",
        "ui_prefabs",
        "timelines",
        "animations",
        "vfx_prefabs",
        "prefabs",
        "wwise_assets",
        "config_assets",
        "scripts",
        "other_resources",
        "why",
        "evidence",
        "url",
    ]
    strict_rows = []
    needs_rows = []
    for item in data.get("jira_resource_matrix", []):
        jira = item.get("jira", {})
        strict_buckets = item.get("strict_buckets", {})
        strict_count = int(item.get("strict_resource_count", 0) or 0)
        row = [
            item.get("decision", ""),
            item.get("strict_confidence", ""),
            item.get("strict_best_score", 0),
            strict_count,
            jira.get("key", ""),
            jira.get("summary", ""),
            jira.get("status", ""),
            jira.get("ready_state", ""),
            jira.get("system", ""),
            jira.get("creator", ""),
            jira.get("assignee", ""),
            jira.get("audio_required", ""),
            jira.get("sound_type", ""),
            jira.get("design_doc", ""),
            jira.get("design_evidence", ""),
            jira.get("qa_evidence", ""),
            strict_buckets.get("ui_prefabs", ""),
            strict_buckets.get("timelines", ""),
            strict_buckets.get("animations", ""),
            strict_buckets.get("vfx_prefabs", ""),
            strict_buckets.get("prefabs", ""),
            strict_buckets.get("wwise_assets", ""),
            strict_buckets.get("config_assets", ""),
            strict_buckets.get("scripts", ""),
            strict_buckets.get("other_resources", ""),
            item.get("strict_why", ""),
            item.get("strict_evidence", ""),
            jira.get("url", ""),
        ]
        if strict_count > 0:
            strict_rows.append(row)
        elif str(jira.get("audio_required", "")).lower() in {"yes", "maybe", "true"} or jira.get("design_evidence") or jira.get("qa_evidence"):
            needs_rows.append(row)
    add_sheet("Strict_Jira_Resource_Matrix", strict_headers, strict_rows)
    add_sheet("Needs_Resource_Evidence", strict_headers, needs_rows)

    thumb_headers = [
        "jira_key",
        "jira_summary",
        "bucket",
        "confidence",
        "score",
        "asset_name",
        "asset_path",
        "thumbnail_file",
        "image_found",
        "jira_url",
    ]
    thumb_rows = []
    for item in data.get("thumbnail", {}).get("assets", []):
        image_path = Path(str(item.get("thumbnail_file", "")))
        thumb_rows.append([
            item.get("jira_key", ""),
            item.get("jira_summary", ""),
            item.get("bucket", ""),
            item.get("confidence", ""),
            item.get("score", 0),
            item.get("asset_name", ""),
            item.get("asset_path", ""),
            item.get("thumbnail_file", ""),
            "Yes" if image_path.exists() else "No",
            item.get("jira_url", ""),
        ])
    add_sheet("Thumbnail_Manifest", thumb_headers, thumb_rows)

    jira_headers = [
        "confidence",
        "best_score",
        "jira_key",
        "jira_summary",
        "jira_status",
        "jira_ready",
        "jira_system",
        "creator",
        "assignee",
        "top_resources",
        "url",
    ]
    jira_rows = []
    for item in data["jira_summary"]:
        jira_rows.append([
            item["best_confidence"],
            item["best_score"],
            item["jira"]["key"],
            item["jira"]["summary"],
            item["jira"].get("status", ""),
            item["jira"].get("ready_state", ""),
            item["jira"].get("system", ""),
            item["jira"].get("creator", ""),
            item["jira"].get("assignee", ""),
            item["top_resources"],
            item["jira"].get("url", ""),
        ])
    add_sheet("By_Jira", jira_headers, jira_rows)

    p4_headers = ["change", "date", "owner", "jira_keys", "audio_related", "reason", "file_count", "summary"]
    p4_rows = []
    for ch in data["p4_audio_changes"]:
        p4_rows.append([
            ch.get("change", ""),
            ch.get("date", ""),
            ch.get("owner", ""),
            ", ".join(ch.get("jira_keys", [])),
            ch.get("audio_related", False),
            ch.get("reason", ""),
            len(ch.get("files", [])),
            ch.get("summary", ""),
        ])
    add_sheet("P4_Audio_History", p4_headers, p4_rows)

    unlinked_headers = [
        "resource_name",
        "kind",
        "priority",
        "recommended_layer",
        "system",
        "group",
        "path",
        "reason",
    ]
    unlinked_rows = []
    for item in data["unlinked_resources"]:
        r = item["resource"]
        unlinked_rows.append([
            r["name"],
            r["kind"],
            r.get("audio_priority", ""),
            r.get("recommended_config_layer", ""),
            r.get("system", ""),
            r.get("group", ""),
            r["unity_path"],
            item.get("reason", ""),
        ])
    add_sheet("Unlinked_Audio_Candidates", unlinked_headers, unlinked_rows)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> int:
    parser = argparse.ArgumentParser(description="Link audio-relevant ProjectEF resources to Jira issues using local caches and P4 metadata.")
    parser.add_argument("--action-index", type=Path, default=None)
    parser.add_argument("--audio-candidate-xlsx", type=Path, default=None)
    parser.add_argument("--use-action-index", action="store_true", help="Use the latest raw action index instead of a prebuilt audio candidate workbook.")
    parser.add_argument("--jira-cache", type=Path, default=DEFAULT_TOOLS_DIR / "audio_requirement_jira_issue_cache.json")
    parser.add_argument("--p4-learning", type=Path, default=DEFAULT_TOOLS_DIR / "p4_changelist_learning.json")
    parser.add_argument("--out-dir", type=Path, default=DEFAULT_REPORT_DIR)
    parser.add_argument("--min-score", type=int, default=45)
    parser.add_argument("--max-links-per-resource", type=int, default=5)
    parser.add_argument("--p4-since", default="2026/06/01")
    parser.add_argument("--p4-max-changes", type=int, default=0, help="Read-only recent P4 changelists to inspect. 0 disables live P4 scan.")
    parser.add_argument("--p4-describe-limit", type=int, default=300)
    parser.add_argument("--p4-scope", default=DEFAULT_P4_SCOPE)
    parser.add_argument("--jira-resource-seed-xlsx", type=Path, action="append", default=[], help="Optional Jira-to-resource seed workbook. One row should contain Jira text, resource names, and notes/paths.")
    parser.add_argument("--no-auto-jira-resource-seeds", action="store_true", help="Disable auto-loading known Jira-to-resource seed workbooks under the Wwise work folder.")
    parser.add_argument("--design-index", type=Path, default=DEFAULT_DESIGN_INDEX)
    parser.add_argument("--qa-index", type=Path, default=DEFAULT_QA_INDEX)
    parser.add_argument("--no-design-qa-evidence", action="store_true", help="Disable design/QA cache evidence enrichment.")
    parser.add_argument("--thumbnail-dir", type=Path, default=DEFAULT_THUMBNAIL_DIR)
    parser.add_argument("--thumbnail-assets-per-jira", type=int, default=1)
    parser.add_argument("--max-thumbnail-jiras", type=int, default=200)
    parser.add_argument("--no-embed-thumbnails", action="store_true", help="Do not embed existing thumbnail PNG files into the Excel matrix.")
    args = parser.parse_args()

    action_index = args.action_index or latest_file(DEFAULT_ACTION_INDEX_DIR, ["ProjectEF_ActionResourceIndex_*.json"])
    audio_xlsx = None
    if not args.use_action_index:
        audio_xlsx = args.audio_candidate_xlsx or latest_file(DEFAULT_ACTION_INDEX_DIR, ["ProjectEF_AudioCandidateResourceIndex_AudioWork_*.xlsx", "ProjectEF_AudioCandidateResourceIndex_Strict_*.xlsx"])
    if not audio_xlsx and not action_index:
        raise SystemExit("No resource index found. Run ProjectEF_ActionResourceIndex.py first.")
    if not args.jira_cache.exists():
        raise SystemExit(f"Jira cache not found: {args.jira_cache}")

    resources = load_resources(action_index, audio_xlsx)
    issues = load_jira_issues(args.jira_cache)
    doc_evidence_by_key: Dict[str, Dict[str, Any]] = {}
    if not args.no_design_qa_evidence:
        doc_evidence_by_key = enrich_issues_with_design_qa(issues, args.design_index, args.qa_index)
    learning_changes, learning_path_index = load_p4_learning(args.p4_learning)
    recent_changes = collect_recent_p4_audio_changes(args.p4_max_changes, args.p4_since, args.p4_scope, args.p4_describe_limit)
    seed_paths = list(args.jira_resource_seed_xlsx or [])
    if not args.no_auto_jira_resource_seeds:
        auto_seed = latest_file_recursive(DEFAULT_TOOLS_DIR.parent, ["ui_system_prefab_name_request_final_*_jira_links.xlsx"])
        if auto_seed and auto_seed not in seed_paths:
            seed_paths.append(auto_seed)
    seed_jira_links = load_jira_resource_seed_links(seed_paths)
    all_p4_changes_by_id = {ch.change: ch for ch in learning_changes}
    for ch in recent_changes:
        all_p4_changes_by_id[ch.change] = ch
    all_p4_changes = list(all_p4_changes_by_id.values())
    p4_path_index = build_path_change_index(all_p4_changes)
    # Keep old path index matches too in case p4 learning stores workspace-relative paths.
    for p, changes in learning_path_index.items():
        p4_path_index[p].extend(changes)

    issue_by_key = {issue.key: issue for issue in issues}
    links: List[Dict[str, Any]] = []
    resource_links: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    jira_links: Dict[str, List[Dict[str, Any]]] = defaultdict(list)

    # Restrict expensive scoring to resources that already look audio/action relevant.
    candidate_resources = [
        r for r in resources
        if r.audio_priority and r.audio_priority != "Reference"
        or text_has_any(" ".join([r.name, r.unity_path, r.audio_keyword_hits]), RESOURCE_ACTION_HINTS)
        or text_has_any(" ".join([r.name, r.unity_path, r.recommended_config_layer]), AUDIO_PATH_HINTS)
        or r.kind in {"Timeline", "Timeline Prefab", "VFX Prefab"}
    ]

    for resource in candidate_resources:
        p4_matches = find_resource_p4_changes(resource, p4_path_index)
        direct_keys = set()
        for ch in p4_matches:
            direct_keys.update(ch.jira_keys)

        scored: List[Tuple[int, JiraIssue, List[str], List[str]]] = []
        # Score direct Jira keys first, then all audio-required Jira issues with cheap token overlap.
        issue_pool = []
        for key in direct_keys:
            issue = issue_by_key.get(key)
            if issue:
                issue_pool.append(issue)
        rtoks = set(resource.tokens)
        for issue in issues:
            if issue in issue_pool:
                continue
            if issue.audio_required.lower() not in {"yes", "maybe", "true"} and not text_has_any(issue.summary + " " + issue.description, AUDIO_TEXT_HINTS):
                continue
            if rtoks & set(issue.tokens) or resource.system.lower() in (issue.system + " " + issue.design_area + " " + issue.summary).lower():
                issue_pool.append(issue)

        for issue in issue_pool:
            score, reasons, evidence = score_issue_for_resource(resource, issue, p4_matches)
            if score >= args.min_score:
                scored.append((score, issue, reasons, evidence))
        scored.sort(key=lambda x: x[0], reverse=True)
        for score, issue, reasons, evidence in scored[: args.max_links_per_resource]:
            link = {
                "confidence": confidence(score),
                "score": score,
                "resource": resource_public(resource),
                "jira": {
                    "key": issue.key,
                    "summary": issue.summary,
                    "status": issue.status,
                    "ready_state": issue.ready_state,
                    "system": issue.system,
                    "creator": issue.creator,
                    "assignee": issue.assignee,
                    "design_doc": issue.design_doc,
                    "url": issue.url,
                },
                "why": safe_join(reasons),
                "evidence": safe_join(evidence),
            }
            links.append(link)
            resource_links[resource.unity_path].append(link)
            jira_links[issue.key].append(link)

    links.sort(key=lambda x: ({"High": 0, "Medium": 1, "Low": 2, "Weak": 3}.get(x["confidence"], 9), -x["score"], x["resource"]["name"]))

    resource_summary = []
    for resource in candidate_resources:
        rs = sorted(resource_links.get(resource.unity_path, []), key=lambda x: x["score"], reverse=True)
        if rs:
            top = rs[0]
            resource_summary.append({
                "best_confidence": top["confidence"],
                "best_score": top["score"],
                "resource": resource_public(resource),
                "top_jira_links": safe_join([f"{x['jira']['key']}({x['confidence']},{x['score']}): {x['jira']['summary']}" for x in rs[:5]], "\n", 1500),
                "why": safe_join([x["why"] for x in rs[:3]], "\n", 1000),
            })
    resource_summary.sort(key=lambda x: ({"High": 0, "Medium": 1, "Low": 2, "Weak": 3}.get(x["best_confidence"], 9), -x["best_score"]))

    resource_catalog = []
    for resource in candidate_resources:
        rs = sorted(resource_links.get(resource.unity_path, []), key=lambda x: x["score"], reverse=True)
        top = rs[0] if rs else None
        rpub = resource_public(resource)
        item = {
            "best_confidence": top["confidence"] if top else "",
            "best_score": top["score"] if top else 0,
            "resource": rpub,
            "top_link": top,
            "all_jira_links": safe_join([f"{x['jira']['key']}({x['confidence']},{x['score']}): {x['jira']['summary']}" for x in rs[:8]], "\n", 2200),
            "why": safe_join([x["why"] for x in rs[:3]], "\n", 1000),
            "type_sheet": resource_type_sheet(rpub),
        }
        resource_catalog.append(item)
    resource_catalog.sort(
        key=lambda x: (
            {"Type_Timeline": 0, "Type_Animation": 1, "Type_Prefab": 2, "Type_VFX": 3, "Type_UI": 4, "Type_Other": 5}.get(x["type_sheet"], 9),
            {"High": 0, "Medium": 1, "Low": 2, "Weak": 3, "": 4}.get(x["best_confidence"], 9),
            -int(x["best_score"] or 0),
            {"High": 0, "Medium": 1, "Low": 2, "Reference": 3, "": 4}.get(x["resource"].get("audio_priority", ""), 9),
            x["resource"].get("name", ""),
        )
    )

    jira_summary = []
    for issue in issues:
        js = sorted(jira_links.get(issue.key, []), key=lambda x: x["score"], reverse=True)
        if js:
            top = js[0]
            jira_summary.append({
                "best_confidence": top["confidence"],
                "best_score": top["score"],
                "jira": top["jira"],
                "top_resources": safe_join([f"{x['resource']['name']}({x['confidence']},{x['score']}): {x['resource']['unity_path']}" for x in js[:8]], "\n", 2000),
            })
    jira_summary.sort(key=lambda x: ({"High": 0, "Medium": 1, "Low": 2, "Weak": 3}.get(x["best_confidence"], 9), -x["best_score"]))
    matrix_jira_links: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for key, values in jira_links.items():
        matrix_jira_links[key].extend(values)
    for key, values in seed_jira_links.items():
        matrix_jira_links[key].extend(values)
    jira_resource_matrix = build_jira_resource_matrix(issues, matrix_jira_links, args.thumbnail_dir)
    strict_matrix_rows = [row for row in jira_resource_matrix if int(row.get("strict_resource_count", 0) or 0) > 0]
    needs_resource_evidence_rows = [
        row for row in jira_resource_matrix
        if int(row.get("strict_resource_count", 0) or 0) == 0
        and (
            str(row.get("jira", {}).get("audio_required", "")).lower() in {"yes", "maybe", "true"}
            or row.get("jira", {}).get("design_evidence")
            or row.get("jira", {}).get("qa_evidence")
        )
    ]
    strict_decision_counts = dict(Counter(str(row.get("decision", "")) for row in jira_resource_matrix))

    linked_paths = set(resource_links.keys())
    unlinked_resources = []
    for resource in candidate_resources:
        if resource.unity_path in linked_paths:
            continue
        unlinked_resources.append({
            "resource": resource_public(resource),
            "reason": "No Jira/P4 evidence above threshold. Needs manual Jira keyword or owner confirmation.",
        })

    stamp = now_stamp()
    args.out_dir.mkdir(parents=True, exist_ok=True)
    args.thumbnail_dir.mkdir(parents=True, exist_ok=True)
    out_json = args.out_dir / f"ProjectEF_AudioResourceJiraLinks_{stamp}.json"
    out_xlsx = args.out_dir / f"ProjectEF_AudioResourceJiraLinks_{stamp}.xlsx"
    thumbnail_manifest_path = args.thumbnail_dir / f"ProjectEF_AudioResourceThumbnailManifest_{stamp}.json"
    thumbnail_assets = build_thumbnail_manifest_assets(
        jira_resource_matrix,
        assets_per_jira=max(0, int(args.thumbnail_assets_per_jira)),
        max_jiras=max(0, int(args.max_thumbnail_jiras)),
    )
    thumbnail_manifest = {
        "version": 1,
        "generated_at": _dt.datetime.now().isoformat(timespec="seconds"),
        "output_dir": str(args.thumbnail_dir),
        "source_report_json": str(out_json),
        "source_report_xlsx": str(out_xlsx),
        "assets": thumbnail_assets,
    }
    write_json(thumbnail_manifest_path, thumbnail_manifest)
    p4_changes_json = [
        {
            "change": ch.change,
            "date": ch.date,
            "owner": ch.owner,
            "summary": ch.summary,
            "jira_keys": ch.jira_keys,
            "audio_related": ch.audio_related,
            "reason": ch.reason,
            "files": ch.files[:200],
        }
        for ch in all_p4_changes
        if ch.audio_related
    ]
    report = {
        "version": 1,
        "generated_at": _dt.datetime.now().isoformat(timespec="seconds"),
        "inputs": {
            "resource_index": str(audio_xlsx or action_index),
            "jira_cache": str(args.jira_cache),
            "p4_learning": str(args.p4_learning) if args.p4_learning else "",
            "jira_resource_seeds": [str(p) for p in seed_paths],
            "design_index": str(args.design_index) if args.design_index else "",
            "qa_index": str(args.qa_index) if args.qa_index else "",
            "thumbnail_manifest": str(thumbnail_manifest_path),
        },
        "summary": {
            "resources_loaded": len(resources),
            "candidate_resources": len(candidate_resources),
            "catalog_resources": len(resource_catalog),
            "jira_issues_loaded": len(issues),
            "p4_learning_changes": len(learning_changes),
            "recent_p4_audio_changes": len(recent_changes),
            "jira_resource_seed_files": len(seed_paths),
            "jira_resource_seed_links": sum(len(v) for v in seed_jira_links.values()),
            "design_qa_evidence_issues": len([v for v in doc_evidence_by_key.values() if v.get("design_evidence") or v.get("qa_evidence")]),
            "links": len(links),
            "linked_resources": len(resource_summary),
            "linked_jira": len(jira_summary),
            "jira_resource_matrix_rows": len(jira_resource_matrix),
            "strict_jira_resource_matrix_rows": len(strict_matrix_rows),
            "needs_resource_evidence_rows": len(needs_resource_evidence_rows),
            "strict_decision_counts": strict_decision_counts,
            "thumbnail_manifest_assets": len(thumbnail_assets),
            "thumbnail_images_found": count_existing_thumbnail_images(thumbnail_assets),
            "unlinked_audio_candidates": len(unlinked_resources),
            "confidence_counts": dict(Counter(x["confidence"] for x in links)),
        },
        "thumbnail": {
            "dir": str(args.thumbnail_dir),
            "manifest": str(thumbnail_manifest_path),
            "assets": thumbnail_assets,
            "embed_enabled": not args.no_embed_thumbnails,
        },
        "links": links,
        "resource_catalog": resource_catalog,
        "resource_summary": resource_summary,
        "jira_summary": jira_summary,
        "jira_resource_matrix": jira_resource_matrix,
        "unlinked_resources": unlinked_resources,
        "p4_audio_changes": p4_changes_json,
    }
    write_json(out_json, report)
    write_excel(out_xlsx, report)
    print(json.dumps({"json": str(out_json), "xlsx": str(out_xlsx), "summary": report["summary"]}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
