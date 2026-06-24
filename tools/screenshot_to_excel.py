from __future__ import annotations

import ctypes
import hashlib
import os
import queue
import shutil
import threading
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional

import tkinter as tk
from tkinter import messagebox, ttk

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageGrab, ImageOps

try:
    from watchdog.events import FileSystemEventHandler
    from watchdog.observers import Observer
except Exception:  # pragma: no cover - the launcher checks this dependency.
    FileSystemEventHandler = object  # type: ignore[assignment]
    Observer = None  # type: ignore[assignment]


APP_TITLE = "截图转 Excel 留档"
IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".bmp", ".webp", ".tif", ".tiff"}
THUMB_MAX_SIZE = (420, 240)
DUPLICATE_WINDOW_SECONDS = 2.5
WORKBOOK_SHEET_NAME = "截图留档"
META_SHEET_NAME = "会话信息"
PREFERRED_WORKBOOK_NAME = "截图留档.xlsx"
DEFAULT_HEADERS = ["序号", "捕获时间", "来源", "尺寸", "原图路径", "Prefab名称", "备注", "缩略图"]
COLUMN_WIDTHS_BY_HEADER = {
    "序号": 8,
    "捕获时间": 21,
    "来源": 16,
    "尺寸": 14,
    "原图路径": 72,
    "Prefab名称": 24,
    "备注": 34,
    "缩略图": 54,
}


@dataclass
class CaptureRecord:
    index: int
    captured_at: datetime
    source: str
    image_path: Path
    width: int
    height: int
    pixel_hash: str
    original_path: Optional[Path] = None


def now_stamp(value: Optional[datetime] = None) -> str:
    value = value or datetime.now()
    return value.strftime("%Y%m%d_%H%M%S")


def display_time(value: datetime) -> str:
    return value.strftime("%Y-%m-%d %H:%M:%S")


def make_unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    parent = path.parent
    for number in range(2, 1000):
        candidate = parent / f"{stem}_{number}{suffix}"
        if not candidate.exists():
            return candidate
    raise RuntimeError(f"无法生成唯一文件名: {path}")


def safe_source_slug(source: str) -> str:
    result = []
    for char in source:
        if char.isalnum():
            result.append(char)
        elif char in ("_", "-"):
            result.append(char)
    slug = "".join(result).strip("_")
    return slug[:24] or "shot"


def normalize_image(image: Image.Image) -> Image.Image:
    image = ImageOps.exif_transpose(image)
    if image.mode == "RGB":
        return image.copy()
    if image.mode == "RGBA":
        background = Image.new("RGB", image.size, "white")
        background.paste(image, mask=image.getchannel("A"))
        return background
    return image.convert("RGB")


def pixel_hash_for(image: Image.Image) -> str:
    normalized = normalize_image(image)
    digest = hashlib.sha256()
    digest.update(f"{normalized.width}x{normalized.height}:RGB:".encode("ascii"))
    digest.update(normalized.tobytes())
    return digest.hexdigest()


def clipboard_sequence_number() -> int:
    try:
        user32 = ctypes.windll.user32
        user32.GetClipboardSequenceNumber.restype = ctypes.c_uint
        return int(user32.GetClipboardSequenceNumber())
    except Exception:
        return int(time.time() * 1000)


def default_watch_dirs() -> list[Path]:
    home = Path.home()
    bases = [home]

    for env_name in ("OneDrive", "OneDriveConsumer", "OneDriveCommercial"):
        value = os.environ.get(env_name)
        if value:
            bases.append(Path(value))

    candidates: list[Path] = []
    for base in bases:
        candidates.extend(
            [
                base / "Pictures" / "Screenshots",
                base / "Pictures" / "ScreenShots",
                base / "Pictures" / "屏幕截图",
                base / "Pictures" / "截图",
                base / "图片" / "Screenshots",
                base / "图片" / "屏幕截图",
                base / "图片" / "截图",
            ]
        )

    candidates.extend(
        [
            home / "Videos" / "Captures",
            home / "Videos" / "捕获",
            home / "视频" / "Captures",
            home / "视频" / "捕获",
        ]
    )

    unique: list[Path] = []
    seen: set[str] = set()
    for path in candidates:
        key = str(path).lower()
        if key not in seen:
            unique.append(path)
            seen.add(key)
    return unique


class ScreenshotFileHandler(FileSystemEventHandler):  # type: ignore[misc]
    def __init__(self, session: "ScreenshotSession") -> None:
        self.session = session

    def on_created(self, event) -> None:  # noqa: ANN001 - watchdog event type.
        self._handle(event)

    def on_moved(self, event) -> None:  # noqa: ANN001 - watchdog event type.
        target = getattr(event, "dest_path", None)
        if target:
            self.session.capture_file_later(Path(target), "截图目录")

    def on_modified(self, event) -> None:  # noqa: ANN001 - watchdog event type.
        self._handle(event)

    def _handle(self, event) -> None:  # noqa: ANN001 - watchdog event type.
        if getattr(event, "is_directory", False):
            return
        self.session.capture_file_later(Path(event.src_path), "截图目录")


class ScreenshotSession:
    def __init__(self, workspace_root: Path, event_queue: "queue.Queue[tuple[str, object]]") -> None:
        self.workspace_root = workspace_root
        self.output_root = workspace_root / "Reports" / "ScreenshotCapture"
        self.start_dt = datetime.now()
        self.start_epoch = time.time()
        self.event_queue = event_queue

        self.stop_event = threading.Event()
        self.records: list[CaptureRecord] = []
        self.capture_lock = threading.Lock()
        self.recent_hash_times: dict[str, float] = {}
        self.claimed_source_paths: set[str] = set()
        self.source_lock = threading.Lock()

        self.temp_image_dir: Optional[Path] = None
        self.running_workbook_path: Optional[Path] = None
        self.saved_record_count = 0
        self.next_index = 1
        self.watch_dirs = default_watch_dirs()
        self.watched_dirs: set[str] = set()
        self.observer = Observer() if Observer else None
        self.threads: list[threading.Thread] = []
        self.running_workbook_path = self._ensure_running_workbook_path()
        self.next_index = self._next_index_from_workbook(self.running_workbook_path)

    @property
    def has_records(self) -> bool:
        with self.capture_lock:
            return bool(self.records)

    def start(self) -> None:
        self._start_watchers_for_existing_dirs()

        if self.observer:
            try:
                self.observer.start()
            except Exception as exc:
                self.event_queue.put(("log", f"截图目录监听启动失败: {exc}"))

        self._start_thread(self._clipboard_loop, "clipboard-watch")
        self._start_thread(self._watch_dir_discovery_loop, "screenshot-dir-discovery")

    def _start_thread(self, target, name: str) -> None:  # noqa: ANN001 - threading target.
        thread = threading.Thread(target=target, name=name, daemon=True)
        thread.start()
        self.threads.append(thread)

    def _start_watchers_for_existing_dirs(self) -> None:
        if not self.observer:
            return
        for path in self.watch_dirs:
            if not path.exists() or not path.is_dir():
                continue
            key = str(path.resolve()).lower()
            if key in self.watched_dirs:
                continue
            try:
                self.observer.schedule(ScreenshotFileHandler(self), str(path), recursive=False)
                self.watched_dirs.add(key)
                self.event_queue.put(("watching", str(path)))
            except Exception as exc:
                self.event_queue.put(("log", f"无法监听截图目录 {path}: {exc}"))

    def _watch_dir_discovery_loop(self) -> None:
        while not self.stop_event.wait(3.0):
            self._start_watchers_for_existing_dirs()
            self._scan_recent_files()

    def _scan_recent_files(self) -> None:
        cutoff = self.start_epoch - 1.0
        for folder in self.watch_dirs:
            if not folder.exists() or not folder.is_dir():
                continue
            try:
                for item in folder.iterdir():
                    if not item.is_file() or item.suffix.lower() not in IMAGE_SUFFIXES:
                        continue
                    try:
                        if item.stat().st_mtime >= cutoff:
                            self.capture_file_later(item, "截图目录")
                    except OSError:
                        continue
            except OSError:
                continue

    def _clipboard_loop(self) -> None:
        last_sequence = clipboard_sequence_number()
        while not self.stop_event.wait(0.45):
            current_sequence = clipboard_sequence_number()
            if current_sequence == last_sequence:
                continue
            last_sequence = current_sequence
            self.capture_clipboard()

    def capture_clipboard(self) -> None:
        if self.stop_event.is_set():
            return

        payload = None
        for _ in range(4):
            try:
                payload = ImageGrab.grabclipboard()
                break
            except Exception:
                time.sleep(0.15)

        if isinstance(payload, Image.Image):
            self.capture_image(payload, "剪贴板")
            return

        if isinstance(payload, list):
            for item in payload:
                path = Path(item)
                if path.suffix.lower() in IMAGE_SUFFIXES:
                    self.capture_file_later(path, "剪贴板文件")

    def capture_file_later(self, path: Path, source: str) -> None:
        if self.stop_event.is_set():
            return
        if path.suffix.lower() not in IMAGE_SUFFIXES:
            return
        thread = threading.Thread(
            target=self.capture_file,
            args=(path, source),
            name="capture-file",
            daemon=True,
        )
        thread.start()

    def capture_file(self, path: Path, source: str) -> None:
        try:
            resolved = str(path.resolve()).lower()
        except OSError:
            return

        with self.source_lock:
            if resolved in self.claimed_source_paths:
                return
            self.claimed_source_paths.add(resolved)

        try:
            if not self._wait_for_file_ready(path):
                with self.source_lock:
                    self.claimed_source_paths.discard(resolved)
                return

            if self.stop_event.is_set():
                with self.source_lock:
                    self.claimed_source_paths.discard(resolved)
                return

            with Image.open(path) as image:
                self.capture_image(image.copy(), source, original_path=path)
        except Exception as exc:
            with self.source_lock:
                self.claimed_source_paths.discard(resolved)
            self.event_queue.put(("log", f"跳过无法读取的截图 {path}: {exc}"))

    def _wait_for_file_ready(self, path: Path) -> bool:
        last_size = -1
        stable_count = 0
        for _ in range(20):
            if self.stop_event.is_set():
                return False
            try:
                size = path.stat().st_size
                with path.open("rb"):
                    pass
            except OSError:
                time.sleep(0.15)
                continue
            if size == last_size and size > 0:
                stable_count += 1
                if stable_count >= 2:
                    return True
            else:
                stable_count = 0
                last_size = size
            time.sleep(0.15)
        return path.exists()

    def capture_image(self, image: Image.Image, source: str, original_path: Optional[Path] = None) -> bool:
        if self.stop_event.is_set():
            return False

        captured_at = datetime.now()
        normalized = normalize_image(image)
        pixel_hash = pixel_hash_for(normalized)

        with self.capture_lock:
            if self.stop_event.is_set():
                return False

            monotonic_now = time.monotonic()
            last_seen = self.recent_hash_times.get(pixel_hash)
            if last_seen and monotonic_now - last_seen < DUPLICATE_WINDOW_SECONDS:
                return False
            self.recent_hash_times[pixel_hash] = monotonic_now
            self._prune_recent_hashes(monotonic_now)

            temp_dir = self._ensure_temp_image_dir()
            index = self.next_index
            self.next_index += 1

            filename = f"shot_{index:04d}_{captured_at.strftime('%Y%m%d_%H%M%S_%f')[:-3]}_{safe_source_slug(source)}.png"
            image_path = temp_dir / filename
            normalized.save(image_path, format="PNG", optimize=False)

            record = CaptureRecord(
                index=index,
                captured_at=captured_at,
                source=source,
                image_path=image_path,
                width=normalized.width,
                height=normalized.height,
                pixel_hash=pixel_hash,
                original_path=original_path,
            )
            self.records.append(record)
            self._write_running_workbook_locked()

        self.event_queue.put(("captured", record))
        return True

    def _prune_recent_hashes(self, monotonic_now: float) -> None:
        stale = [
            image_hash
            for image_hash, seen_at in self.recent_hash_times.items()
            if monotonic_now - seen_at > 60.0
        ]
        for image_hash in stale:
            self.recent_hash_times.pop(image_hash, None)

    def _ensure_temp_image_dir(self) -> Path:
        if self.temp_image_dir is None:
            self.output_root.mkdir(parents=True, exist_ok=True)
            workbook_path = self._ensure_running_workbook_path()
            temp_name = f"{workbook_path.stem}_images"
            self.temp_image_dir = self.output_root / temp_name
            self.temp_image_dir.mkdir(parents=True, exist_ok=True)
        return self.temp_image_dir

    def _find_existing_workbook_path(self) -> Path:
        self.output_root.mkdir(parents=True, exist_ok=True)
        preferred = self.output_root / PREFERRED_WORKBOOK_NAME
        if preferred.exists():
            return preferred
        candidates = [
            path
            for path in self.output_root.glob("*.xlsx")
            if path.is_file() and not path.name.startswith("~$")
        ]
        if candidates:
            return max(candidates, key=lambda path: path.stat().st_mtime)
        return preferred

    def _ensure_running_workbook_path(self) -> Path:
        if self.running_workbook_path is None:
            self.running_workbook_path = self._find_existing_workbook_path()
        return self.running_workbook_path

    def _next_index_from_workbook(self, workbook_path: Path) -> int:
        if not workbook_path.exists():
            return 1
        try:
            wb = load_workbook(workbook_path, read_only=True, data_only=True)
            try:
                ws = wb[WORKBOOK_SHEET_NAME] if WORKBOOK_SHEET_NAME in wb.sheetnames else wb.active
                rows = ws.iter_rows(min_row=1, max_row=1, values_only=True)
                headers = [str(value).strip() if value is not None else "" for value in next(rows, [])]
                index_col = headers.index("序号") + 1 if "序号" in headers else 1
                max_index = 0
                for (value,) in ws.iter_rows(min_row=2, min_col=index_col, max_col=index_col, values_only=True):
                    if value is None:
                        continue
                    try:
                        max_index = max(max_index, int(value))
                    except (TypeError, ValueError):
                        continue
                return max_index + 1 if max_index else 1
            finally:
                wb.close()
        except Exception as exc:
            self.event_queue.put(("log", f"读取已有 Excel 序号失败，将从 1 开始: {exc}"))
            return 1

    def _write_running_workbook_locked(self) -> None:
        if self.temp_image_dir is None or not self.records:
            return
        workbook_path = self._ensure_running_workbook_path()
        pending_records = self.records[self.saved_record_count :]
        if not pending_records:
            return
        try:
            self._append_records_to_workbook(
                workbook_path,
                pending_records,
                self.temp_image_dir,
            )
            self.saved_record_count += len(pending_records)
        except Exception as exc:
            self.event_queue.put(("log", f"Excel 追加失败，截图原图已保存；请确认 Excel 未被打开占用: {exc}"))

    def stop(self) -> None:
        self.stop_event.set()
        if self.observer:
            try:
                self.observer.stop()
                self.observer.join(timeout=2.0)
            except Exception:
                pass

    def finalize(self) -> Optional[Path]:
        # Catch the last Win+Shift+S/PrintScreen clipboard image before stopping.
        self.capture_clipboard()
        self.stop()
        with self.capture_lock:
            records = list(self.records)
        if not records:
            return None

        with self.capture_lock:
            self._write_running_workbook_locked()
        return self._ensure_running_workbook_path()

    def _append_records_to_workbook(
        self,
        workbook_path: Path,
        records: list[CaptureRecord],
        image_dir: Path,
    ) -> None:
        created = not workbook_path.exists()
        if created:
            wb = Workbook()
            ws = wb.active
            ws.title = WORKBOOK_SHEET_NAME
        else:
            wb = load_workbook(workbook_path)
            ws = wb[WORKBOOK_SHEET_NAME] if WORKBOOK_SHEET_NAME in wb.sheetnames else wb.active

        header_map = self._ensure_headers(ws)
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="D9E2F3")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for header, width in COLUMN_WIDTHS_BY_HEADER.items():
            col_index = header_map.get(header)
            if col_index:
                ws.column_dimensions[get_column_letter(col_index)].width = width

        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False

        thumb_dir = image_dir / "_thumbs"
        thumb_dir.mkdir(exist_ok=True)

        for record in records:
            row_index = ws.max_row + 1
            size_text = f"{record.width} x {record.height}"
            source_text = record.source
            if record.original_path:
                source_text = f"{record.source}"

            values_by_header = {
                "序号": record.index,
                "捕获时间": display_time(record.captured_at),
                "来源": source_text,
                "尺寸": size_text,
                "原图路径": str(record.image_path),
                "备注": f"原始来源: {record.original_path}" if record.original_path else "",
                "缩略图": "",
            }
            for header, value in values_by_header.items():
                col_index = header_map.get(header)
                if not col_index:
                    continue
                cell = ws.cell(row=row_index, column=col_index, value=value)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if header in {"序号", "来源", "尺寸"}:
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            for col_index in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_index, column=col_index)
                cell.border = border
                if not cell.alignment:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

            path_col = header_map.get("原图路径")
            if path_col:
                path_cell = ws.cell(row=row_index, column=path_col)
                try:
                    path_cell.hyperlink = record.image_path.resolve().as_uri()
                    path_cell.style = "Hyperlink"
                except ValueError:
                    pass

            thumb_path, thumb_width, thumb_height = self._make_thumbnail(record.image_path, thumb_dir)
            excel_image = ExcelImage(str(thumb_path))
            excel_image.width = thumb_width
            excel_image.height = thumb_height
            thumb_col = header_map.get("缩略图")
            if thumb_col:
                ws.add_image(excel_image, f"{get_column_letter(thumb_col)}{row_index}")
            ws.row_dimensions[row_index].height = max(120, int(thumb_height * 0.75) + 12)

        ws.row_dimensions[1].height = 24
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        if created and META_SHEET_NAME not in wb.sheetnames:
            self._create_meta_sheet(wb, image_dir)

        workbook_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(workbook_path)
        wb.close()

    def _ensure_headers(self, ws) -> dict[str, int]:  # noqa: ANN001 - openpyxl worksheet.
        if ws.max_row < 1:
            ws.append(DEFAULT_HEADERS)
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
        if not any(headers):
            for col_index, header in enumerate(DEFAULT_HEADERS, start=1):
                ws.cell(row=1, column=col_index, value=header)
            headers = list(DEFAULT_HEADERS)

        header_map = {header: index for index, header in enumerate(headers, start=1) if header}
        for header in DEFAULT_HEADERS:
            if header in header_map:
                continue
            col_index = ws.max_column + 1
            ws.cell(row=1, column=col_index, value=header)
            header_map[header] = col_index
        return header_map

    def _create_meta_sheet(self, wb, image_dir: Path) -> None:  # noqa: ANN001 - openpyxl workbook.
        meta = wb.create_sheet(META_SHEET_NAME)
        meta.append(["项目", "内容"])
        meta.append(["会话文件名", self._ensure_running_workbook_path().name])
        meta.append(["首次创建时间", display_time(self.start_dt)])
        meta.append(["图片目录", str(image_dir)])
        meta.append(["说明", "本 Excel 现在为持续追加模式；关闭并重新打开工具后会继续写入同一个文件。Prefab名称 和其他手工新增列会保留。"])
        meta.column_dimensions["A"].width = 18
        meta.column_dimensions["B"].width = 90
        for row in meta.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in meta[1]:
            cell.fill = header_fill
            cell.font = header_font

    def _make_thumbnail(self, image_path: Path, thumb_dir: Path) -> tuple[Path, int, int]:
        thumb_path = thumb_dir / f"{image_path.stem}_thumb.png"
        with Image.open(image_path) as image:
            thumb = normalize_image(image)
            thumb.thumbnail(THUMB_MAX_SIZE, Image.Resampling.LANCZOS)
            thumb.save(thumb_path, format="PNG", optimize=True)
        return thumb_path, thumb.width, thumb.height


class ScreenshotToExcelApp(tk.Tk):
    def __init__(self, workspace_root: Path) -> None:
        super().__init__()
        self.workspace_root = workspace_root
        self.event_queue: "queue.Queue[tuple[str, object]]" = queue.Queue()
        self.session = ScreenshotSession(workspace_root, self.event_queue)
        self.finalizing = False

        self.title(APP_TITLE)
        self.geometry("820x520")
        self.minsize(760, 460)
        self.configure(bg="#F6F8FB")

        self._build_ui()
        self.protocol("WM_DELETE_WINDOW", self.stop_and_close)

        self.session.start()
        self.after(200, self._drain_events)

    def _build_ui(self) -> None:
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass
        style.configure("TFrame", background="#F6F8FB")
        style.configure("TLabel", background="#F6F8FB", foreground="#172033", font=("Microsoft YaHei UI", 10))
        style.configure("Title.TLabel", font=("Microsoft YaHei UI", 16, "bold"), foreground="#172033")
        style.configure("Count.TLabel", font=("Microsoft YaHei UI", 24, "bold"), foreground="#1F4E79")
        style.configure("TButton", font=("Microsoft YaHei UI", 10), padding=(12, 6))
        style.configure("Treeview", rowheight=28, font=("Microsoft YaHei UI", 9))
        style.configure("Treeview.Heading", font=("Microsoft YaHei UI", 9, "bold"))

        outer = ttk.Frame(self, padding=18)
        outer.pack(fill=tk.BOTH, expand=True)

        top = ttk.Frame(outer)
        top.pack(fill=tk.X)

        title_area = ttk.Frame(top)
        title_area.pack(side=tk.LEFT, fill=tk.X, expand=True)
        ttk.Label(title_area, text=APP_TITLE, style="Title.TLabel").pack(anchor=tk.W)
        self.status_var = tk.StringVar(value=f"正在采集，开始时间：{display_time(self.session.start_dt)}")
        ttk.Label(title_area, textvariable=self.status_var).pack(anchor=tk.W, pady=(6, 0))

        count_area = ttk.Frame(top)
        count_area.pack(side=tk.RIGHT, padx=(16, 0))
        self.count_var = tk.StringVar(value="0")
        ttk.Label(count_area, textvariable=self.count_var, style="Count.TLabel").pack(anchor=tk.E)
        ttk.Label(count_area, text="已捕获截图").pack(anchor=tk.E)

        info = ttk.Frame(outer)
        info.pack(fill=tk.X, pady=(16, 10))
        output_text = f"输出位置：{self.session.output_root}"
        ttk.Label(info, text=output_text).pack(anchor=tk.W)
        ttk.Label(info, text="支持 PrintScreen / Win+Shift+S 剪贴板截图、Win+PrintScreen 截图目录、Xbox Game Bar Captures。").pack(anchor=tk.W, pady=(4, 0))
        ttk.Label(info, text="截图会持续追加到同一个 Excel；关闭再打开工具后继续往后写，手工新增列会保留。").pack(anchor=tk.W, pady=(4, 0))

        table_frame = ttk.Frame(outer)
        table_frame.pack(fill=tk.BOTH, expand=True, pady=(8, 12))

        columns = ("index", "time", "source", "size", "path")
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings", selectmode="browse")
        self.tree.heading("index", text="#")
        self.tree.heading("time", text="捕获时间")
        self.tree.heading("source", text="来源")
        self.tree.heading("size", text="尺寸")
        self.tree.heading("path", text="保存路径")
        self.tree.column("index", width=52, anchor=tk.CENTER, stretch=False)
        self.tree.column("time", width=150, anchor=tk.W, stretch=False)
        self.tree.column("source", width=100, anchor=tk.CENTER, stretch=False)
        self.tree.column("size", width=95, anchor=tk.CENTER, stretch=False)
        self.tree.column("path", width=380, anchor=tk.W, stretch=True)

        scrollbar = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        buttons = ttk.Frame(outer)
        buttons.pack(fill=tk.X)
        ttk.Button(buttons, text="打开输出目录", command=self.open_output_dir).pack(side=tk.LEFT)
        ttk.Button(buttons, text="停止并保存 Excel", command=self.stop_and_close).pack(side=tk.RIGHT)

    def _drain_events(self) -> None:
        while True:
            try:
                kind, payload = self.event_queue.get_nowait()
            except queue.Empty:
                break

            if kind == "captured" and isinstance(payload, CaptureRecord):
                self._add_record(payload)
            elif kind == "watching":
                self.status_var.set(f"正在采集，开始时间：{display_time(self.session.start_dt)}")
            elif kind == "log":
                self.status_var.set(str(payload))

        if not self.finalizing:
            self.after(200, self._drain_events)

    def _add_record(self, record: CaptureRecord) -> None:
        self.count_var.set(str(record.index))
        size_text = f"{record.width} x {record.height}"
        self.tree.insert(
            "",
            tk.END,
            values=(
                record.index,
                display_time(record.captured_at),
                record.source,
                size_text,
                str(record.image_path),
            ),
        )
        self.tree.yview_moveto(1.0)
        workbook_text = ""
        if self.session.running_workbook_path:
            workbook_text = f"；Excel：{self.session.running_workbook_path.name}"
        self.status_var.set(f"已捕获第 {record.index} 张截图{workbook_text}")

    def open_output_dir(self) -> None:
        self.session.output_root.mkdir(parents=True, exist_ok=True)
        os.startfile(str(self.session.output_root))

    def stop_and_close(self) -> None:
        if self.finalizing:
            return
        self.finalizing = True
        self.status_var.set("正在停止监听并保存 Excel...")
        self.update_idletasks()

        try:
            workbook_path = self.session.finalize()
        except Exception as exc:
            messagebox.showerror(APP_TITLE, f"保存 Excel 失败：\n{exc}")
            self.destroy()
            return

        if workbook_path:
            messagebox.showinfo(APP_TITLE, f"已保存 Excel：\n{workbook_path}")
        else:
            messagebox.showinfo(APP_TITLE, "这次没有捕获到新截图，Excel 未变化。")
        self.destroy()


def main() -> int:
    workspace_root = Path(__file__).resolve().parents[1]
    app = ScreenshotToExcelApp(workspace_root)
    app.mainloop()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
